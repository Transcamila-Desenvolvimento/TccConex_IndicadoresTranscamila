import io
import re
from datetime import date
from decimal import Decimal
from django.db import transaction
from django.db.models import Q, Sum, Avg
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from apps.accounts.mixins import ModuleScopedViewMixin
from apps.audit.services import record_audit

from .models import (
    Colaborador,
    LoteMovimentacaoRH,
    MovimentacaoColaborador,
    InconsistenciaColaborador,
    CargoMapping,
    ColaboradorPJ,
)
from .serializers import (
    ColaboradorSerializer,
    LoteMovimentacaoRHSerializer,
    MovimentacaoColaboradorSerializer,
    InconsistenciaColaboradorSerializer,
    CargoMappingSerializer,
    ColaboradorPJSerializer,
)
from .import_service import (
    import_movimentacao_mensal,
    import_movimentacao_lote_completo,
    importar_historico_salarial_completo,
)
from .utils import definir_categoria_colaborador


def _parse_emails(value) -> list[str]:
    """Aceita lista (`EmailTagsInput`) ou string separada por vírgula/ponto-e-vírgula."""
    if value is None:
        return []
    if isinstance(value, list):
        parts = value
    else:
        parts = re.split(r'[,;]+', str(value))
    return [p.strip() for p in parts if p and '@' in p]


def _meses_afastado(cpf: str, lote_atual: 'LoteMovimentacaoRH') -> int:
    """Conta há quantos meses consecutivos (incluindo o mês do lote atual) o
    colaborador aparece com situação de afastamento. Não existe um campo de
    "data de início do afastamento" na planilha mensal, então a contagem é
    feita retroagindo mês a mês pelos lotes já importados até encontrar uma
    lacuna: lote inexistente, colaborador ausente naquele mês, ou situação
    diferente de afastado."""
    meses = 0
    mes, ano = lote_atual.mes, lote_atual.ano
    while True:
        lote_mes = LoteMovimentacaoRH.objects.filter(mes=mes, ano=ano).first()
        if not lote_mes:
            break
        registro = lote_mes.colaboradores.filter(cpf=cpf).first()
        if not registro or not registro.situacao or 'AFASTADO' not in registro.situacao.upper():
            break
        meses += 1
        mes -= 1
        if mes == 0:
            mes = 12
            ano -= 1
    return meses


class LoteMovimentacaoRHViewSet(ModuleScopedViewMixin, viewsets.ModelViewSet):
    permission_module = 'RH'
    serializer_class = LoteMovimentacaoRHSerializer
    queryset = LoteMovimentacaoRH.objects.all()

    @action(detail=False, methods=['post'])
    def prepare_import(self, request):
        mes = int(request.data.get('mes', timezone.now().month))
        ano = int(request.data.get('ano', timezone.now().year))
        
        batch, created = LoteMovimentacaoRH.objects.get_or_create(
            mes=mes,
            ano=ano,
            defaults={'usuario': request.user}
        )
        
        action_name = 'rh.lote.criado' if created else 'rh.lote.reutilizado'
        record_audit(
            request.user,
            action_name,
            f'Lote de movimentação para {mes:02d}/{ano} {"criado" if created else "reutilizado"}.'
        )
        
        return Response(LoteMovimentacaoRHSerializer(batch).data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def importar_arquivo(self, request):
        mes = int(request.data.get('mes'))
        ano = int(request.data.get('ano'))
        arquivo = request.FILES.get('arquivo')

        if not arquivo:
            return Response({'error': 'Arquivo Excel não fornecido.'}, status=status.HTTP_400_BAD_REQUEST)

        # Ler os bytes ANTES de salvar no FileField: o Storage do Django consome
        # todo o stream do upload ao salvar (chunks() percorre até o EOF), então um
        # `arquivo.read()` posterior retornaria vazio e quebraria o parsing do Excel.
        file_bytes = arquivo.read()
        arquivo.seek(0)

        lote, _ = LoteMovimentacaoRH.objects.update_or_create(
            mes=mes, ano=ano,
            defaults={'usuario': request.user, 'arquivo': arquivo}
        )

        try:
            res = import_movimentacao_mensal(lote, file_bytes, request.user)
            record_audit(
                request.user,
                'rh.movimentacao.importada',
                f'Planilha de movimentação importada para o período {mes:02d}/{ano}.'
            )
            return Response(res, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def importar_lote(self, request):
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            return Response({'error': 'Arquivo Excel não fornecido.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            res = import_movimentacao_lote_completo(arquivo.read(), request.user)
            record_audit(
                request.user,
                'rh.lote_completo.importado',
                f'Planilha multi-período de movimentações importada com sucesso ({res.get("lotes_count")} períodos).'
            )
            return Response(res, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def exportar_modelo(self, request):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lote_Movimentacoes"
        
        colunas = [
            'C.P.F.', 'Nome', 'Salario', 'Filial', 'Estado', 
            'Desc. Cargo', 'Data Admis.', 'Data Nasc.', 'Mes', 'Ano'
        ]
        ws.append(colunas)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=modelo_importacao_lote_rh.xlsx'
        return response

    @action(detail=True, methods=['post'])
    def enviar_email(self, request, pk=None):
        lote = self.get_object()
        to_emails = _parse_emails(request.data.get('to') or request.data.get('email'))
        cc_emails = _parse_emails(request.data.get('cc') or request.data.get('emailCopia'))

        if not to_emails:
            return Response({'error': 'Informe ao menos um destinatário.'}, status=status.HTTP_400_BAD_REQUEST)

        # 1. Compilar estatísticas para o e-mail
        cpfs_desconsiderados = set(Colaborador.objects.filter(desconsiderado=True).values_list('cpf', flat=True))
        colaboradores = lote.colaboradores.exclude(cpf__in=cpfs_desconsiderados)
        
        mes_ant = 12 if lote.mes == 1 else lote.mes - 1
        ano_ant = lote.ano - 1 if lote.mes == 1 else lote.ano
        lote_ant = LoteMovimentacaoRH.objects.filter(mes=mes_ant, ano=ano_ant).first()
        
        cpfs_atuais = set(colaboradores.values_list('cpf', flat=True))
        cpfs_anteriores = set()
        if lote_ant:
            colaboradores_ant = lote_ant.colaboradores.exclude(cpf__in=cpfs_desconsiderados)
            cpfs_anteriores = set(colaboradores_ant.values_list('cpf', flat=True))
            
        cpfs_novos = cpfs_atuais - cpfs_anteriores
        cpfs_desligados = cpfs_anteriores - cpfs_atuais
        
        filiais = colaboradores.values_list('filial', flat=True).distinct()
        relatorio_filiais = []
        novos_detalhes = []
        desligados_detalhes = []
        
        hoje = date.today()
        total_idades_geral = []
        total_tempos_geral = []

        def format_tenure(months_count):
            months_rounded = round(months_count)
            anos = int(months_rounded // 12)
            meses = int(months_rounded % 12)
            partes = []
            if anos > 0:
                partes.append(f"{anos} {'ano' if anos == 1 else 'anos'}")
            if meses > 0:
                partes.append(f"{meses} {'mês' if meses == 1 else 'meses'}")
            return " e ".join(partes) if partes else "0 meses"

        def format_currency_br(val):
            return f"R$ {val:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

        for f_nome in filiais:
            colabs_f = colaboradores.filter(filial=f_nome)
            qty = colabs_f.count()
            payroll = colabs_f.aggregate(Sum('salario'))['salario__sum'] or Decimal('0')
            
            novos_f = colabs_f.filter(cpf__in=cpfs_novos)
            for n in novos_f:
                novos_detalhes.append({
                    'nome': n.nome.upper(),
                    'filial': n.filial or 'N/I',
                    'cargo': n.funcao or 'N/I'
                })
                
            des_count = 0
            if lote_ant:
                des_objs = lote_ant.colaboradores.filter(filial=f_nome, cpf__in=cpfs_desligados)
                des_count = des_objs.count()
                for d in des_objs:
                    desligados_detalhes.append({
                        'nome': d.nome.upper(),
                        'filial': d.filial or 'N/I',
                        'cargo': d.funcao or 'N/I'
                    })
                    
            idades = []
            tempos = []
            for c in colabs_f:
                if c.data_nascimento:
                    idade = hoje.year - c.data_nascimento.year - ((hoje.month, hoje.day) < (c.data_nascimento.month, c.data_nascimento.day))
                    idades.append(idade)
                    total_idades_geral.append(idade)
                if c.data_admissao:
                    tempo_anos = (hoje - c.data_admissao).days / 365.25
                    tempo_meses = tempo_anos * 12
                    tempos.append(tempo_meses)
                    total_tempos_geral.append(tempo_meses)
                    
            avg_age = sum(idades) / len(idades) if idades else 0
            avg_tenure = sum(tempos) / len(tempos) if tempos else 0
            
            relatorio_filiais.append({
                'nome': f_nome or 'Não Informada',
                'qty': qty,
                'total_payroll': payroll,
                'total_payroll_str': format_currency_br(payroll),
                'avg_age': round(avg_age, 1),
                'avg_tenure_str': format_tenure(avg_tenure),
                'novos': novos_f.count(),
                'desligados': des_count
            })

        total_colabs = colaboradores.count()
        total_adm = colaboradores.filter(categoria='ADMINISTRATIVO').count()
        total_oper = colaboradores.filter(categoria='OPERACIONAL').count()
        total_mot_ativos = colaboradores.filter(categoria='MOTORISTA').exclude(situacao__icontains='AFASTADO').count()
        total_mot_afastados = colaboradores.filter(categoria='MOTORISTA', situacao__icontains='AFASTADO').count()
        total_folha = colaboradores.aggregate(Sum('salario'))['salario__sum'] or Decimal('0')
        avg_age_g = sum(total_idades_geral) / len(total_idades_geral) if total_idades_geral else 0
        avg_tenure_g = sum(total_tempos_geral) / len(total_tempos_geral) if total_tempos_geral else 0

        # Alterações (inconsistências)
        alteracoes = lote.inconsistencias.exclude(cpf__in=cpfs_desconsiderados).order_by('nome')

        # Afastados
        afastados_objs = colaboradores.filter(situacao__icontains='AFASTADO').order_by('filial', 'nome')
        afastados_por_filial = {}
        for a in afastados_objs:
            f_key = a.filial or 'Não Informada'
            if f_key not in afastados_por_filial:
                afastados_por_filial[f_key] = []
            meses_afastado = _meses_afastado(a.cpf, lote)
            afastados_por_filial[f_key].append({
                'nome': a.nome.upper(),
                'cargo': a.funcao or 'N/I',
                'situacao': a.situacao or 'AFASTADO',
                'tempo_afastado_str': format_tenure(meses_afastado) if meses_afastado else None,
            })

        # 2. Gerar Anexo Excel
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        excel_buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alterações"
        
        header_fill = PatternFill(start_color="118CC4", end_color="118CC4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        headers = ["Funcionário", "CPF", "Tipo", "Valor Anterior", "Valor Atual", "Data", "Justificativa"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        periodo_lote = date(lote.ano, lote.mes, 1)

        for alt in alteracoes:
            ws.append([
                alt.nome.upper(),
                alt.cpf,
                alt.get_tipo_display(),
                alt.valor_anterior or "",
                alt.valor_atual or "",
                periodo_lote,
                alt.justificativa or ""
            ])
            ws.cell(row=ws.max_row, column=6).number_format = 'MM/YYYY'
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # 3. Disparar e-mail
        from django.core.mail import EmailMessage
        from django.template.loader import render_to_string
        
        meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        meses_abrev = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        mes_nome = meses[lote.mes - 1]

        # Histórico da folha de pagamento dos últimos 5 meses, para o gráfico do e-mail
        lotes_recentes = list(
            LoteMovimentacaoRH.objects.filter(
                Q(ano__lt=lote.ano) | Q(ano=lote.ano, mes__lte=lote.mes)
            ).order_by('-ano', '-mes')[:5]
        )
        lotes_recentes.reverse()

        folha_historico = []
        for lr in lotes_recentes:
            colabs_lr = lr.colaboradores.exclude(cpf__in=cpfs_desconsiderados)
            payroll_lr = colabs_lr.aggregate(Sum('salario'))['salario__sum'] or Decimal('0')
            folha_historico.append({
                'label': f'{meses_abrev[lr.mes - 1]}/{str(lr.ano)[-2:]}',
                'valor': payroll_lr,
                'valor_str': format_currency_br(payroll_lr),
                'atual': lr.id == lote.id,
            })

        from .chart_service import gerar_grafico_evolucao_folha
        grafico_folha_bytes = gerar_grafico_evolucao_folha(folha_historico) if len(folha_historico) > 1 else b''

        context = {
            'lote': lote,
            'relatorio': relatorio_filiais,
            'geral': {
                'total_colabs': total_colabs,
                'total_adm': total_adm,
                'total_oper': total_oper,
                'total_mot_ativos': total_mot_ativos,
                'total_mot_afastados': total_mot_afastados,
                'total_folha_str': format_currency_br(total_folha),
                'avg_age': round(avg_age_g, 1),
                'avg_tenure_str': format_tenure(avg_tenure_g)
            },
            'novos_detalhes': novos_detalhes,
            'desligados_detalhes': desligados_detalhes,
            'alteracoes': alteracoes,
            'afastados_por_filial': afastados_por_filial,
            'tem_grafico_folha': bool(grafico_folha_bytes),
            'ano': lote.ano,
            'mes_nome': mes_nome
        }
        
        html_body = render_to_string('rh/email_movimentacao.html', context)

        from .pdf_service import gerar_pdf_relatorio_movimentacoes
        pdf_bytes = gerar_pdf_relatorio_movimentacoes(context)

        email_obj = EmailMessage(
            subject=f'Relatório de Movimentação RH - {mes_nome}/{lote.ano}',
            body=html_body,
            from_email=None,
            to=to_emails,
            cc=cc_emails,
        )
        email_obj.content_subtype = "html"
        email_obj.attach(
            f'Relatorio_Alteracoes_{lote.mes:02d}_{lote.ano}.xlsx',
            excel_buffer.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        email_obj.attach(
            f'Relatorio_Movimentacoes_{lote.mes:02d}_{lote.ano}.pdf',
            pdf_bytes,
            'application/pdf'
        )

        if grafico_folha_bytes:
            from email.mime.image import MIMEImage
            grafico_img = MIMEImage(grafico_folha_bytes, _subtype='png')
            grafico_img.add_header('Content-ID', '<grafico_folha>')
            grafico_img.add_header('Content-Disposition', 'inline', filename='grafico_folha.png')
            email_obj.attach(grafico_img)

        try:
            email_obj.send(fail_silently=False)
            record_audit(
                request.user,
                'rh.email.enviado',
                f'Relatório de movimentação do período {lote.mes:02d}/{lote.ano} enviado por e-mail para {", ".join(to_emails)}.'
            )
            return Response({'success': True, 'message': f'E-mail enviado para {", ".join(to_emails)}.'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Falha ao enviar e-mail: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MovimentacaoColaboradorViewSet(ModuleScopedViewMixin, viewsets.ReadOnlyModelViewSet):
    permission_module = 'RH'
    serializer_class = MovimentacaoColaboradorSerializer
    queryset = MovimentacaoColaborador.objects.all()

    def get_queryset(self):
        qs = super().get_queryset()
        lote_id = self.request.query_params.get('loteId')
        if lote_id:
            qs = qs.filter(lote_id=lote_id)
        filial = self.request.query_params.get('filial')
        if filial:
            qs = qs.filter(filial=filial)
        categoria = self.request.query_params.get('categoria')
        if categoria:
            qs = qs.filter(categoria=categoria)
        situacao = self.request.query_params.get('situacao')
        if situacao:
            qs = qs.filter(situacao=situacao)
            
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(nome__icontains=search) | Q(cpf__icontains=search) | Q(funcao__icontains=search))
            
        # Ocultar desconsiderados
        cpfs_desconsiderados = set(Colaborador.objects.filter(desconsiderado=True).values_list('cpf', flat=True))
        qs = qs.exclude(cpf__in=cpfs_desconsiderados)
        return qs.order_by('nome')

    @action(detail=False, methods=['get'])
    def dashboard_summary(self, request):
        mes = request.query_params.get('mes')
        ano = request.query_params.get('ano')
        lote_id = request.query_params.get('loteId')

        if lote_id:
            lote = get_object_or_404(LoteMovimentacaoRH, id=lote_id)
        elif mes and ano:
            lote = LoteMovimentacaoRH.objects.filter(mes=int(mes), ano=int(ano)).first()
        else:
            lote = LoteMovimentacaoRH.objects.first()

        lotes_disponiveis = LoteMovimentacaoRHSerializer(LoteMovimentacaoRH.objects.all(), many=True).data

        if not lote:
            return Response({
                'lote': None,
                'lotesDisponiveis': lotes_disponiveis,
                'resumoFiliais': [],
                'novos': [],
                'desligados': [],
                'alteracoes': [],
                'totais': {
                    'totalColaboradores': 0,
                    'admitidos': 0,
                    'desligados': 0,
                    'alteracoes': 0,
                    'payroll': Decimal('0'),
                    'mediaIdade': 0,
                    'mediaTempo': 0,
                }
            })

        # Excluir desconsiderados
        cpfs_desconsiderados = set(Colaborador.objects.filter(desconsiderado=True).values_list('cpf', flat=True))
        colaboradores = lote.colaboradores.exclude(cpf__in=cpfs_desconsiderados)
        cpfs_atuais = set(colaboradores.values_list('cpf', flat=True))

        # Buscar lote anterior para comparação
        mes_ant = 12 if lote.mes == 1 else lote.mes - 1
        ano_ant = lote.ano - 1 if lote.mes == 1 else lote.ano
        lote_ant = LoteMovimentacaoRH.objects.filter(mes=mes_ant, ano=ano_ant).first()
        
        cpfs_anteriores = set()
        colaboradores_ant = MovimentacaoColaborador.objects.none()
        if lote_ant:
            colaboradores_ant = lote_ant.colaboradores.exclude(cpf__in=cpfs_desconsiderados)
            cpfs_anteriores = set(colaboradores_ant.values_list('cpf', flat=True))

        # 1. Admitidos (Atuais - Anteriores)
        cpfs_novos = cpfs_atuais - cpfs_anteriores
        novos_qs = colaboradores.filter(cpf__in=cpfs_novos).order_by('nome')
        novos_data = MovimentacaoColaboradorSerializer(novos_qs, many=True).data

        # 2. Desligados (Anteriores - Atuais)
        cpfs_desligados = cpfs_anteriores - cpfs_atuais
        desligados_qs = colaboradores_ant.filter(cpf__in=cpfs_desligados).order_by('nome')
        desligados_data = MovimentacaoColaboradorSerializer(desligados_qs, many=True).data

        # 3. Alterações (Inconsistências registradas no lote)
        alteracoes_qs = lote.inconsistencias.exclude(cpf__in=cpfs_desconsiderados).order_by('nome')
        alteracoes_data = InconsistenciaColaboradorSerializer(alteracoes_qs, many=True).data

        # 4. Resumo por Filial
        hoje = date.today()
        filiais = colaboradores.values_list('filial', flat=True).distinct()
        resumo_filiais = []
        idades_geral = []
        tempos_geral = []

        for f in filiais:
            colabs_f = colaboradores.filter(filial=f)
            total = colabs_f.count()
            
            # Idades e tempos
            idades = []
            tempos = []
            for c in colabs_f:
                if c.data_nascimento:
                    idade = hoje.year - c.data_nascimento.year - ((hoje.month, hoje.day) < (c.data_nascimento.month, c.data_nascimento.day))
                    idades.append(idade)
                    idades_geral.append(idade)
                if c.data_admissao:
                    tempo = (hoje - c.data_admissao).days / 365.25
                    tempos.append(tempo)
                    tempos_geral.append(tempo)
                    
            avg_age = sum(idades) / len(idades) if idades else 0
            avg_tenure = sum(tempos) / len(tempos) if tempos else 0
            
            novos_count = colabs_f.filter(cpf__in=cpfs_novos).count()
            desligados_count = colaboradores_ant.filter(filial=f, cpf__in=cpfs_desligados).count()
            payroll = colabs_f.aggregate(Sum('salario'))['salario__sum'] or Decimal('0')

            resumo_filiais.append({
                'filial': f or 'Não Informada',
                'total': total,
                'payroll': payroll,
                'mediaIdade': round(avg_age, 1),
                'mediaTempo': round(avg_tenure, 1),
                'novos': novos_count,
                'desligados': desligados_count
            })

        avg_age_geral = sum(idades_geral) / len(idades_geral) if idades_geral else 0
        avg_tenure_geral = sum(tempos_geral) / len(tempos_geral) if tempos_geral else 0
        payroll_geral = colaboradores.aggregate(Sum('salario'))['salario__sum'] or Decimal('0')

        totais = {
            'totalColaboradores': colaboradores.count(),
            'admitidos': len(novos_data),
            'desligados': len(desligados_data),
            'alteracoes': len(alteracoes_data),
            'payroll': payroll_geral,
            'mediaIdade': round(avg_age_geral, 1),
            'mediaTempo': round(avg_tenure_geral, 1),
        }

        return Response({
            'lote': LoteMovimentacaoRHSerializer(lote).data,
            'lotesDisponiveis': lotes_disponiveis,
            'resumoFiliais': resumo_filiais,
            'novos': novos_data,
            'desligados': desligados_data,
            'alteracoes': alteracoes_data,
            'totais': totais
        })

    @action(detail=False, methods=['get'])
    def exportar_relatorio(self, request):
        """Gera um Excel com todas as ocorrências do período: contratações, demissões e alterações
        (cargo/salário), com a filial e o tipo de ocorrência de cada colaborador."""
        mes = request.query_params.get('mes')
        ano = request.query_params.get('ano')
        lote_id = request.query_params.get('loteId')

        if lote_id:
            lote = get_object_or_404(LoteMovimentacaoRH, id=lote_id)
        elif mes and ano:
            lote = LoteMovimentacaoRH.objects.filter(mes=int(mes), ano=int(ano)).first()
        else:
            lote = LoteMovimentacaoRH.objects.first()

        if not lote:
            return Response({'error': 'Nenhum lote de movimentação encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        cpfs_desconsiderados = set(Colaborador.objects.filter(desconsiderado=True).values_list('cpf', flat=True))
        colaboradores = lote.colaboradores.exclude(cpf__in=cpfs_desconsiderados)
        cpfs_atuais = set(colaboradores.values_list('cpf', flat=True))

        mes_ant = 12 if lote.mes == 1 else lote.mes - 1
        ano_ant = lote.ano - 1 if lote.mes == 1 else lote.ano
        lote_ant = LoteMovimentacaoRH.objects.filter(mes=mes_ant, ano=ano_ant).first()

        cpfs_anteriores = set()
        colaboradores_ant = MovimentacaoColaborador.objects.none()
        if lote_ant:
            colaboradores_ant = lote_ant.colaboradores.exclude(cpf__in=cpfs_desconsiderados)
            cpfs_anteriores = set(colaboradores_ant.values_list('cpf', flat=True))

        cpfs_novos = cpfs_atuais - cpfs_anteriores
        cpfs_desligados = cpfs_anteriores - cpfs_atuais

        novos_qs = colaboradores.filter(cpf__in=cpfs_novos).order_by('filial', 'nome')
        desligados_qs = colaboradores_ant.filter(cpf__in=cpfs_desligados).order_by('filial', 'nome')
        alteracoes_qs = lote.inconsistencias.exclude(cpf__in=cpfs_desconsiderados).order_by('nome')
        colaboradores_por_cpf = {c.cpf: c for c in colaboradores}

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Movimentações'

        headers = ['Filial', 'Colaborador', 'CPF', 'Ocorrência', 'Cargo', 'Valor Anterior', 'Valor Atual', 'Justificativa']
        ws.append(headers)
        header_fill = PatternFill(start_color='118CC4', end_color='118CC4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for c in novos_qs:
            ws.append([c.filial or 'Não Informada', c.nome.upper(), c.cpf, 'Contratação', c.funcao or '-', '-', '-', '-'])

        for c in desligados_qs:
            ws.append([c.filial or 'Não Informada', c.nome.upper(), c.cpf, 'Demissão', c.funcao or '-', '-', '-', '-'])

        for a in alteracoes_qs:
            colab_atual = colaboradores_por_cpf.get(a.cpf)
            filial = colab_atual.filial if colab_atual else None
            cargo = colab_atual.funcao if colab_atual else None
            ws.append([
                filial or 'Não Informada',
                a.nome.upper(),
                a.cpf,
                a.get_tipo_display(),
                cargo or '-',
                a.valor_anterior or '-',
                a.valor_atual or '-',
                a.justificativa or '-',
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        record_audit(
            request.user,
            'rh.relatorio.exportado',
            f'Relatório de movimentações exportado para o período {lote.mes:02d}/{lote.ano}.'
        )

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Relatorio_Movimentacoes_{lote.mes:02d}_{lote.ano}.xlsx'
        return response

    @action(detail=False, methods=['get'])
    def buscar_comparar(self, request):
        term = request.query_params.get('term', '')
        if not term:
            return Response({'results': []})

        colaboradores_query = MovimentacaoColaborador.objects.filter(
            Q(nome__icontains=term) | Q(funcao__icontains=term) | Q(filial__icontains=term)
        ).values('cpf', 'nome', 'funcao', 'filial').order_by('nome')
        
        seen_cpfs = set()
        results = []
        
        for c in colaboradores_query:
            if c['cpf'] not in seen_cpfs:
                results.append({
                    'cpf': c['cpf'],
                    'nome': c['nome'],
                    'cargo': c['funcao'] or '-',
                    'filial': c['filial'] or '-'
                })
                seen_cpfs.add(c['cpf'])
            if len(results) >= 20:
                break
                
        return Response({'results': results})

    @action(detail=False, methods=['get'])
    def dados_comparacao(self, request):
        cpfs = request.query_params.getlist('cpfs[]')
        if len(cpfs) != 2:
            return Response({'error': 'Selecione exatamente dois colaboradores.'}, status=status.HTTP_400_BAD_REQUEST)
            
        data_final = {}
        for cpf in cpfs:
            historico = MovimentacaoColaborador.objects.filter(cpf=cpf).select_related('lote').order_by('lote__ano', 'lote__mes')
            
            labels = []
            valores = []
            for h in historico:
                label = f"{h.lote.mes:02d}/{h.lote.ano}"
                labels.append(label)
                valores.append(float(h.salario or 0))
                
            colab_info = MovimentacaoColaborador.objects.filter(cpf=cpf).first()
            nome = colab_info.nome if colab_info else "Desconhecido"
            
            data_final[cpf] = {
                'nome': nome,
                'labels': labels,
                'valores': valores
            }
            
        return Response(data_final)


class ColaboradorPJViewSet(ModuleScopedViewMixin, viewsets.ModelViewSet):
    permission_module = 'RH'
    serializer_class = ColaboradorPJSerializer
    queryset = ColaboradorPJ.objects.all()

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(nome__icontains=search) | Q(cpf__icontains=search) | Q(cargo__icontains=search))
        return qs


class CargoMappingViewSet(ModuleScopedViewMixin, viewsets.ModelViewSet):
    permission_module = 'RH'
    serializer_class = CargoMappingSerializer
    queryset = CargoMapping.objects.all()

    def get_queryset(self):
        qs = super().get_queryset()
        status_filter = self.request.query_params.get('status')
        search = self.request.query_params.get('search')
        
        if search:
            qs = qs.filter(cargo__icontains=search)
            
        if status_filter == 'pendente':
            qs = qs.filter(categoria__isnull=True)
        elif status_filter == 'definido':
            qs = qs.filter(categoria__isnull=False)
            
        return qs

    def perform_create(self, serializer):
        with transaction.atomic():
            mapping = serializer.save()
            # Replicar em cascata
            cargo_nome = mapping.cargo
            nova_cat = mapping.categoria
            Colaborador.objects.filter(cargo=cargo_nome).update(categoria=nova_cat)
            MovimentacaoColaborador.objects.filter(funcao=cargo_nome).update(categoria=nova_cat)

    def perform_update(self, serializer):
        with transaction.atomic():
            mapping = serializer.save()
            # Replicar em cascata
            cargo_nome = mapping.cargo
            nova_cat = mapping.categoria
            Colaborador.objects.filter(cargo=cargo_nome).update(categoria=nova_cat)
            MovimentacaoColaborador.objects.filter(funcao=cargo_nome).update(categoria=nova_cat)


class ColaboradorViewSet(ModuleScopedViewMixin, viewsets.ModelViewSet):
    permission_module = 'RH'
    serializer_class = ColaboradorSerializer
    queryset = Colaborador.objects.all()

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(nome_completo__icontains=search) | Q(cpf__icontains=search) | Q(matricula__icontains=search))
            
        desconsiderados_only = self.request.query_params.get('desconsiderados')
        if desconsiderados_only == 'true':
            qs = qs.filter(desconsiderado=True)
            
        # Ordenar desconsiderados no topo se for listagem geral
        return qs.order_by('-desconsiderado', 'nome_completo')

    @action(detail=True, methods=['post'])
    def toggle_desconsiderar(self, request, pk=None):
        colab = self.get_object()
        colab.desconsiderado = not colab.desconsiderado
        colab.save()
        
        record_audit(
            request.user,
            'rh.colaborador.toggle_desconsiderar',
            f'Colaborador {colab.nome_completo} (CPF: {colab.cpf}) marcado como {"desconsiderado" if colab.desconsiderado else "considerado"} nos indicadores.'
        )
        
        return Response({
            'success': True,
            'desconsiderado': colab.desconsiderado
        }, status=status.HTTP_200_OK)


class HistoricoSalarialViewSet(ModuleScopedViewMixin, viewsets.ReadOnlyModelViewSet):
    permission_module = 'RH'
    serializer_class = InconsistenciaColaboradorSerializer
    queryset = InconsistenciaColaborador.objects.filter(tipo='salario')

    def get_queryset(self):
        qs = super().get_queryset().select_related('lote').order_by('-lote__ano', '-lote__mes', 'nome')
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(nome__icontains=search) | Q(cpf__icontains=search))
        return qs

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def importar_historico(self, request):
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            return Response({'error': 'Arquivo Excel não fornecido.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            res = importar_historico_salarial_completo(arquivo.read(), request.user)
            record_audit(
                request.user,
                'rh.historico_salarial.importado',
                f'Planilha de histórico salarial importada com sucesso. ({res.get("imported_count")} reajustes inseridos).'
            )
            return Response(res, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class InconsistenciaColaboradorViewSet(ModuleScopedViewMixin, viewsets.ModelViewSet):
    permission_module = 'RH'
    serializer_class = InconsistenciaColaboradorSerializer
    queryset = InconsistenciaColaborador.objects.all()

    @action(detail=True, methods=['post'])
    def salvar_justificativa(self, request, pk=None):
        inc = self.get_object()
        justificativa = request.data.get('justificativa', '')
        
        inc.justificativa = justificativa
        inc.save()
        
        record_audit(
            request.user,
            'rh.inconsistencia.justificada',
            f'Justificativa de alteração para {inc.nome} (CPF: {inc.cpf}) atualizada.'
        )
        
        return Response({'success': True, 'justificativa': inc.justificativa}, status=status.HTTP_200_OK)
