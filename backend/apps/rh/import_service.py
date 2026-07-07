import re
import unicodedata
from io import BytesIO
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction
from django.db.models import Q

from .models import (
    Colaborador,
    LoteMovimentacaoRH,
    MovimentacaoColaborador,
    InconsistenciaColaborador,
    CargoMapping,
    ColaboradorPJ,
)
from .utils import definir_categoria_colaborador


def _normalize_header(value) -> str:
    text = str(value or '').strip()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if ord(c) < 128 or c.isalnum() or c.isspace())
    return re.sub(r'\s+', ' ', text.lower().strip())


def find_column_index(headers, aliases) -> int | None:
    """Retorna o índice da coluna que corresponde a uma lista de aliases."""
    normalized_headers = [_normalize_header(h) for h in headers]
    
    # 1. Busca exata
    for alias in aliases:
        norm_alias = _normalize_header(alias)
        if norm_alias in normalized_headers:
            return normalized_headers.index(norm_alias)
            
    # 2. Busca parcial (se contiver a string do alias)
    for idx, header in enumerate(normalized_headers):
        for alias in aliases:
            norm_alias = _normalize_header(alias)
            if len(norm_alias) > 2 and norm_alias in header:
                return idx
                
    return None


def clean_cpf(val) -> str | None:
    if val is None or val == '':
        return None
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = re.sub(r'\D', '', s)
    return s if s else None


def clean_money(val) -> Decimal:
    if val is None or val == '':
        return Decimal('0')
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
        
    s = str(val).replace('R$', '').replace(' ', '').strip()
    if not s:
        return Decimal('0')
        
    # Identificar padrão BR (1.234,56) ou US (1,234.56)
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            # Padrão BR
            s = s.replace('.', '').replace(',', '.')
        else:
            # Padrão US
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
        
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal('0')


def parse_date(val) -> date | None:
    if val is None or val == '':
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
        
    s_val = str(val).strip()
    if not re.search(r'\d', s_val):
        return None
        
    # Tentar vários formatos comuns
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s_val, fmt).date()
        except ValueError:
            continue
            
    # Fallback para parsing genérico do pandas/openpyxl
    try:
        # Se for timestamp float do Excel
        if s_val.isdigit():
            # Apenas um número inteiro representando dias desde 1900
            # Para simplificar, tentamos converter para float e ver
            pass
        return datetime.fromisoformat(s_val.split(' ')[0]).date()
    except:
        return None


def import_movimentacao_mensal(lote: LoteMovimentacaoRH, file_bytes: bytes, user) -> dict:
    """Processa a importação de uma planilha para um lote mensal."""
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    
    # Encontrar a aba que contenha CPFs
    best_sheet_name = workbook.sheetnames[0]
    best_sheet = workbook[best_sheet_name]
    best_score = 0
    
    cpf_aliases = ['c.p.f.', 'cpf', 'fis_cpf', 'identificador']
    
    for name in workbook.sheetnames:
        sheet = workbook[name]
        sample = []
        for i, row in enumerate(sheet.iter_rows(max_row=5, values_only=True)):
            sample.extend(row or [])
        flat = ' '.join(_normalize_header(c) for c in sample)
        score = sum(1 for alias in cpf_aliases if _normalize_header(alias) in flat)
        if score > best_score:
            best_score = score
            best_sheet = sheet
            best_sheet_name = name
            
    rows = list(best_sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"A aba '{best_sheet_name}' está vazia.")
        
    # Encontrar linha de cabeçalho
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        if any(any(_normalize_header(alias) in _normalize_header(cell) for alias in cpf_aliases) for cell in (row or []) if cell):
            header_idx = i
            break
            
    headers = [str(c).strip() if c is not None else f"Col{idx}" for idx, c in enumerate(rows[header_idx])]
    
    # Mapeamento de colunas
    col_cpf = find_column_index(headers, cpf_aliases)
    col_nome = find_column_index(headers, ['nome complet', 'nome', 'nome completo', 'funcionario', 'nome do colaborador'])
    col_salario = find_column_index(headers, ['salario', 'salário', 'remuneracao', 'valor'])
    col_filial = find_column_index(headers, ['filial', 'unidade', 'empresa', 'estabelecimento'])
    col_cargo = find_column_index(headers, ['desc. cargo', 'cargo', 'funcao', 'função'])
    col_situacao = find_column_index(headers, ['situacao', 'situação', 'status', 'status do funcionario'])
    col_estado = find_column_index(headers, ['estado', 'uf', 'u.f.'])
    col_admissao = find_column_index(headers, ['data admis.', 'admissão', 'data admissão', 'dt. adm.'])
    col_nasc = find_column_index(headers, ['data nasc.', 'nascimento', 'data nascimento', 'dt. nasc.'])
    col_matricula = find_column_index(headers, ['matricula', 'matrícula', 'chapa'])
    col_depto = find_column_index(headers, ['depto', 'departamento', 'desc. depto.'])
    col_lider = find_column_index(headers, ['lider', 'líder', 'nome lider', 'nome do lider'])
    col_sexo = find_column_index(headers, ['sexo', 'genero', 'gênero'])
    col_escolaridade = find_column_index(headers, ['escolaridade', 'instrucao', 'instrução'])
    col_telefone = find_column_index(headers, ['telefone', 'fone', 'celular'])
    col_empresa = find_column_index(headers, ['empresa', 'nome empresa', 'razao social'])

    if col_cpf is None:
        raise ValueError(f"Não foi possível localizar a coluna de CPF na planilha (Aba: {best_sheet_name}).")

    with transaction.atomic():
        # Limpar dados antigos do lote
        MovimentacaoColaborador.objects.filter(lote=lote).delete()
        InconsistenciaColaborador.objects.filter(lote=lote).delete()
        
        # Buscar lote anterior para comparação de alterações (Salário/Cargo)
        lote_ant = LoteMovimentacaoRH.objects.filter(
            Q(ano__lt=lote.ano) | Q(ano=lote.ano, mes__lt=lote.mes)
        ).order_by('-ano', '-mes').first()
        
        colabs_ant_dict = {}
        if lote_ant:
            colabs_ant_dict = {c.cpf: c for c in lote_ant.colaboradores.all()}
            
        novos_objs = []
        cpfs_importados = set()
        
        for row in rows[header_idx + 1:]:
            # Ignorar linhas vazias
            if not row or all(c is None for c in row):
                continue
                
            cpf = clean_cpf(row[col_cpf])
            if not cpf:
                continue
                
            cpfs_importados.add(cpf)
            
            nome = str(row[col_nome] or '').strip() if col_nome is not None else ""
            salario = clean_money(row[col_salario]) if col_salario is not None else Decimal('0')
            cargo = str(row[col_cargo] or '').strip() if col_cargo is not None else ""
            filial = str(row[col_filial] or '').strip() if col_filial is not None else ""
            situacao = str(row[col_situacao] or '').strip() if col_situacao is not None else ""
            uf = str(row[col_estado] or '').strip() if col_estado is not None else ""
            dt_adm = parse_date(row[col_admissao]) if col_admissao is not None else None
            dt_nasc = parse_date(row[col_nasc]) if col_nasc is not None else None
            
            # Cadastro Geral (Colaborador)
            colab_defaults = {
                'nome_completo': nome,
                'filial': filial,
                'cargo': cargo,
                'situacao': situacao,
                'data_admissao': dt_adm,
                'data_nascimento': dt_nasc,
                'categoria': definir_categoria_colaborador(cargo)
            }
            if col_matricula is not None and row[col_matricula]:
                colab_defaults['matricula'] = str(row[col_matricula]).strip().split('.')[0]
            else:
                colab_defaults['matricula'] = cpf[:8] # Fallback matricula
                
            if col_depto is not None and row[col_depto]:
                colab_defaults['departamento'] = str(row[col_depto]).strip()
            if col_lider is not None and row[col_lider]:
                colab_defaults['nome_lider'] = str(row[col_lider]).strip()
            if col_sexo is not None and row[col_sexo]:
                val_sexo = str(row[col_sexo]).strip().upper()
                colab_defaults['sexo'] = 'M' if val_sexo.startswith('M') else 'F' if val_sexo.startswith('F') else None
            if col_escolaridade is not None and row[col_escolaridade]:
                colab_defaults['escolaridade'] = str(row[col_escolaridade]).strip()
            if col_telefone is not None and row[col_telefone]:
                colab_defaults['telefone'] = str(row[col_telefone]).strip()
            if col_empresa is not None and row[col_empresa]:
                colab_defaults['empresa'] = str(row[col_empresa]).strip()
                
            Colaborador.objects.update_or_create(
                cpf=cpf,
                defaults=colab_defaults
            )
            
            # MovimentacaoColaborador
            mov = MovimentacaoColaborador(
                lote=lote,
                filial=filial,
                nome=nome,
                situacao=situacao,
                uf_estado=uf,
                funcao=cargo,
                data_admissao=dt_adm,
                data_nascimento=dt_nasc,
                cpf=cpf,
                salario=salario,
                categoria=definir_categoria_colaborador(cargo)
            )
            novos_objs.append(mov)
            
            # Inconsistências
            if cpf in colabs_ant_dict:
                ant = colabs_ant_dict[cpf]
                # Alteração salarial
                if float(salario) != float(ant.salario or 0):
                    InconsistenciaColaborador.objects.create(
                        lote=lote,
                        cpf=cpf,
                        nome=nome,
                        tipo='salario',
                        valor_anterior=f"R$ {ant.salario:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                        valor_atual=f"R$ {salario:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                        justificativa="Atualização de dissídio ou mérito"
                    )
                # Alteração de cargo
                cargo_ant = str(ant.funcao or "").strip().upper()
                cargo_novo = cargo.upper()
                if cargo_ant and cargo_novo and cargo_ant != cargo_novo:
                    InconsistenciaColaborador.objects.create(
                        lote=lote,
                        cpf=cpf,
                        nome=nome,
                        tipo='cargo',
                        valor_anterior=ant.funcao,
                        valor_atual=cargo,
                        justificativa="Alteração de Cargo/Promoção"
                    )

        if novos_objs:
            MovimentacaoColaborador.objects.bulk_create(novos_objs)
            
        # Injetar PJs cadastrados e ativos no sistema
        pjs_para_injetar = ColaboradorPJ.objects.filter(ativo=True).exclude(cpf__in=cpfs_importados)
        pjs_objs = []
        for pj in pjs_para_injetar:
            pjs_objs.append(MovimentacaoColaborador(
                lote=lote,
                filial=pj.filial,
                nome=pj.nome,
                situacao='ATIVO (PJ)',
                funcao=pj.cargo,
                data_admissao=pj.data_admissao,
                data_nascimento=pj.data_nascimento,
                cpf=pj.cpf,
                salario=pj.salario,
                categoria=definir_categoria_colaborador(pj.cargo)
            ))
            
        if pjs_objs:
            MovimentacaoColaborador.objects.bulk_create(pjs_objs)
            
        return {
            'success': True,
            'imported': len(novos_objs),
            'pjs_injected': len(pjs_objs),
            'total': len(novos_objs) + len(pjs_objs),
            'aba': best_sheet_name,
        }


def import_movimentacao_lote_completo(file_bytes: bytes, user) -> dict:
    """Processa a importação de uma planilha com múltiplos períodos."""
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = workbook.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("A planilha de lote está vazia.")
        
    headers = [str(c).strip() if c is not None else f"Col{idx}" for idx, c in enumerate(rows[0])]
    
    col_cpf = find_column_index(headers, ['c.p.f.', 'cpf', 'fis_cpf', 'identificador'])
    col_mes = find_column_index(headers, ['mes', 'mês', 'mes_referencia'])
    col_ano = find_column_index(headers, ['ano', 'ano_referencia'])
    
    if col_cpf is None or col_mes is None or col_ano is None:
        raise ValueError("A planilha de lote deve conter as colunas de CPF, Mes e Ano.")
        
    col_nome = find_column_index(headers, ['nome complet', 'nome', 'nome completo', 'funcionario'])
    col_salario = find_column_index(headers, ['salario', 'salário', 'remuneracao', 'valor'])
    col_filial = find_column_index(headers, ['filial', 'unidade', 'empresa'])
    col_cargo = find_column_index(headers, ['desc. cargo', 'cargo', 'funcao', 'função'])
    col_estado = find_column_index(headers, ['estado', 'uf', 'u.f.'])
    col_admissao = find_column_index(headers, ['data admis.', 'admissão', 'data admissão'])
    col_nasc = find_column_index(headers, ['data nasc.', 'nascimento', 'data nascimento'])

    # Organizar linhas por período (ano, mes)
    periodos = {}
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
            
        m_val = row[col_mes]
        a_val = row[col_ano]
        if m_val is None or a_val is None:
            continue
            
        try:
            mes = int(m_val)
            ano = int(a_val)
        except (ValueError, TypeError):
            continue
            
        key = (ano, mes)
        if key not in periodos:
            periodos[key] = []
        periodos[key].append(row)
        
    # Processar cada período cronologicamente para consistência
    ordenados = sorted(periodos.keys())
    lotes_processados = 0
    registros_total = 0
    
    with transaction.atomic():
        for (ano, mes) in ordenados:
            lote, _ = LoteMovimentacaoRH.objects.update_or_create(
                mes=mes, ano=ano,
                defaults={'usuario': user}
            )
            
            # Limpar dados antigos do lote
            MovimentacaoColaborador.objects.filter(lote=lote).delete()
            InconsistenciaColaborador.objects.filter(lote=lote).delete()
            
            lote_ant = LoteMovimentacaoRH.objects.filter(
                Q(ano__lt=ano) | Q(ano=ano, mes__lt=mes)
            ).order_by('-ano', '-mes').first()
            
            colabs_ant_dict = {}
            if lote_ant:
                colabs_ant_dict = {c.cpf: c for c in lote_ant.colaboradores.all()}
                
            novos_objs = []
            cpfs_importados = set()
            
            for row in periodos[(ano, mes)]:
                cpf = clean_cpf(row[col_cpf])
                if not cpf:
                    continue
                cpfs_importados.add(cpf)
                
                nome = str(row[col_nome] or '').strip() if col_nome is not None else ""
                salario = clean_money(row[col_salario]) if col_salario is not None else Decimal('0')
                cargo = str(row[col_cargo] or '').strip() if col_cargo is not None else ""
                filial = str(row[col_filial] or '').strip() if col_filial is not None else ""
                uf = str(row[col_estado] or '').strip() if col_estado is not None else ""
                dt_adm = parse_date(row[col_admissao]) if col_admissao is not None else None
                dt_nasc = parse_date(row[col_nasc]) if col_nasc is not None else None
                
                # Ficha de Colaborador
                Colaborador.objects.get_or_create(
                    cpf=cpf,
                    defaults={
                        'nome_completo': nome,
                        'filial': filial,
                        'cargo': cargo,
                        'situacao': 'SITUACAO NORMAL',
                        'data_admissao': dt_adm,
                        'data_nascimento': dt_nasc,
                        'categoria': definir_categoria_colaborador(cargo)
                    }
                )
                
                # MovimentacaoColaborador
                mov = MovimentacaoColaborador(
                    lote=lote,
                    filial=filial,
                    nome=nome,
                    situacao='SITUACAO NORMAL',
                    uf_estado=uf,
                    funcao=cargo,
                    data_admissao=dt_adm,
                    data_nascimento=dt_nasc,
                    cpf=cpf,
                    salario=salario,
                    categoria=definir_categoria_colaborador(cargo)
                )
                novos_objs.append(mov)
                
                # Inconsistências
                if cpf in colabs_ant_dict:
                    ant = colabs_ant_dict[cpf]
                    # Salário
                    if float(salario) != float(ant.salario or 0):
                        InconsistenciaColaborador.objects.create(
                            lote=lote,
                            cpf=cpf,
                            nome=nome,
                            tipo='salario',
                            valor_anterior=f"R$ {ant.salario:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                            valor_atual=f"R$ {salario:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                            justificativa="Ajuste salarial de lote"
                        )
                    # Cargo
                    cargo_ant = str(ant.funcao or "").strip().upper()
                    cargo_novo = cargo.upper()
                    if cargo_ant and cargo_novo and cargo_ant != cargo_novo:
                        InconsistenciaColaborador.objects.create(
                            lote=lote,
                            cpf=cpf,
                            nome=nome,
                            tipo='cargo',
                            valor_anterior=ant.funcao,
                            valor_atual=cargo,
                            justificativa="Alteração de Cargo de lote"
                        )
                        
            if novos_objs:
                MovimentacaoColaborador.objects.bulk_create(novos_objs)
                
            # Injetar PJs
            pjs_para_injetar = ColaboradorPJ.objects.filter(ativo=True).exclude(cpf__in=cpfs_importados)
            pjs_objs = []
            for pj in pjs_para_injetar:
                pjs_objs.append(MovimentacaoColaborador(
                    lote=lote,
                    filial=pj.filial,
                    nome=pj.nome,
                    situacao='ATIVO (PJ)',
                    funcao=pj.cargo,
                    data_admissao=pj.data_admissao,
                    data_nascimento=pj.data_nascimento,
                    cpf=pj.cpf,
                    salario=pj.salario,
                    categoria=definir_categoria_colaborador(pj.cargo)
                ))
                
            if pjs_objs:
                MovimentacaoColaborador.objects.bulk_create(pjs_objs)
                
            lotes_processados += 1
            registros_total += (len(novos_objs) + len(pjs_objs))
            
    return {
        'success': True,
        'lotes_count': lotes_processados,
        'records_count': registros_total
    }


def importar_historico_salarial_completo(file_bytes: bytes, user) -> dict:
    """Importa histórico de alterações salariais a partir de planilha personalizada."""
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    
    # Geralmente os históricos estão na segunda aba, se houver
    aba_idx = 1 if len(workbook.sheetnames) > 1 else 0
    ws = workbook[workbook.sheetnames[aba_idx]]
    rows = list(ws.iter_rows(values_only=True))
    
    records = []
    
    for row in rows:
        if not row or all(c is None for c in row):
            continue
            
        matricula = str(row[0]).strip().split('.')[0] if row[0] is not None else ""
        if not matricula.isdigit():
            continue
            
        try:
            nome = str(row[1] or '').strip()
            dt_val = row[2]
            dt = parse_date(dt_val)
            tipo_alteracao = str(row[3] or '').strip()
            valor = clean_money(row[6])
            
            if dt and valor > 0:
                records.append({
                    'matricula': matricula,
                    'nome': nome,
                    'data': dt,
                    'tipo': tipo_alteracao,
                    'valor': valor
                })
        except:
            continue
            
    if not records:
        raise ValueError("Nenhum registro válido de histórico salarial encontrado. Verifique se os dados estão na segunda aba.")
        
    # Ordenar por matricula e data
    records.sort(key=lambda x: (x['matricula'], x['data']))
    
    count_created = 0
    last_salaries = {} # matricula -> valor
    
    with transaction.atomic():
        for rec in records:
            mat = rec['matricula']
            val_atual = rec['valor']
            val_ant = last_salaries.get(mat, Decimal('0'))
            
            last_salaries[mat] = val_atual
            
            # Localizar CPF
            colab = Colaborador.objects.filter(matricula=mat).first()
            if not colab:
                colab = Colaborador.objects.filter(nome_completo__icontains=rec['nome']).first()
                
            cpf = colab.cpf if colab else f"MAT-{mat}"
            
            lote, _ = LoteMovimentacaoRH.objects.get_or_create(
                mes=rec['data'].month,
                ano=rec['data'].year,
                defaults={'usuario': user}
            )
            
            v_atual_str = f"R$ {val_atual:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            v_ant_str = f"R$ {val_ant:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if val_ant > 0 else "-"
            
            # Evitar duplicados exatos
            if not InconsistenciaColaborador.objects.filter(lote=lote, cpf=cpf, tipo='salario', valor_atual=v_atual_str).exists():
                InconsistenciaColaborador.objects.create(
                    lote=lote,
                    cpf=cpf,
                    nome=rec['nome'],
                    tipo='salario',
                    valor_anterior=v_ant_str,
                    valor_atual=v_atual_str,
                    justificativa=rec['tipo']
                )
                count_created += 1
                
    return {
        'success': True,
        'imported_count': count_created
    }
