from io import BytesIO

import openpyxl
from django.contrib.auth import get_user_model
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIClient

from apps.accounts.tests import auth_headers
from datetime import date
from decimal import Decimal

from apps.rh.models import (
    LoteMovimentacaoRH,
    MovimentacaoColaborador,
    InconsistenciaColaborador,
    ColaboradorPJ,
    ColaboradorPJHistorico,
)
from apps.rh.pj_sync_service import sync_pj_nos_lotes
from apps.rh.import_service import clean_cpf, _pick_sheet_for_lote
from apps.rh.views import _meses_afastado
import openpyxl
from io import BytesIO

User = get_user_model()


def _build_xlsx_bytes(rows) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class ImportarMovimentacaoMensalTests(TestCase):
    """Regressão: importar_arquivo salvava o upload no FileField antes de lê-lo,
    o que esvaziava o stream (Storage.save() consome o arquivo até o EOF) e fazia
    o openpyxl falhar com 'File is not a zip file' ao tentar processar bytes vazios."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='rh_import_tests',
            password='rh123',
            role_id='2',
            environments=['RH'],
            filiais={},
        )
        self.xlsx_bytes = _build_xlsx_bytes([
            ['C.P.F.', 'Nome', 'Salario', 'Filial', 'Desc. Cargo', 'Situacao'],
            ['123.456.789-00', 'Fulano de Tal', '2500.50', 'Ibiporã (Matriz)', 'Motorista', 'ATIVO'],
            ['987.654.321-00', 'Ciclana de Souza', '3200.00', 'Rondonópolis', 'Analista Administrativo', 'ATIVO'],
        ])

    def test_importar_arquivo_processa_planilha_e_persiste_movimentacoes(self):
        upload = SimpleUploadedFile(
            'ativos.xlsx',
            self.xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            '/api/rh/lotes/importar_arquivo/',
            data={'mes': 11, 'ano': 2026, 'arquivo': upload},
            format='multipart',
            **auth_headers(self.user, 'RH'),
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['imported'], 2)

        lote = LoteMovimentacaoRH.objects.get(mes=11, ano=2026)
        self.assertEqual(MovimentacaoColaborador.objects.filter(lote=lote).count(), 2)
        # O arquivo enviado deve ter sido salvo corretamente no storage (não vazio).
        self.assertTrue(lote.arquivo)
        self.assertGreater(lote.arquivo.size, 0)

    def test_importar_arquivo_sem_arquivo_retorna_erro_400(self):
        response = self.client.post(
            '/api/rh/lotes/importar_arquivo/',
            data={'mes': 11, 'ano': 2026},
            format='multipart',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('error', response.data)


class EnviarEmailMovimentacaoTests(TestCase):
    """Regressão: o modal de e-mail do RH passou a enviar `to`/`cc` como listas
    (autocomplete de contatos do Google via EmailTagsInput), igual ao relatório
    gerencial do Financeiro. O endpoint precisa aceitar múltiplos destinatários."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='rh_email_tests',
            password='rh123',
            role_id='2',
            environments=['RH'],
            filiais={},
        )
        self.lote = LoteMovimentacaoRH.objects.create(mes=11, ano=2026, usuario=self.user, data_importacao=timezone.now())
        MovimentacaoColaborador.objects.create(
            lote=self.lote,
            filial='Ibiporã (Matriz)',
            nome='Fulano de Tal',
            situacao='ATIVO',
            funcao='Motorista',
            cpf='12345678900',
            salario='2500.50',
            categoria='MOTORISTA',
        )

    def test_enviar_email_com_multiplos_destinatarios(self):
        response = self.client.post(
            f'/api/rh/lotes/{self.lote.id}/enviar_email/',
            data={
                'to': ['destino1@transcamila.com.br', 'destino2@transcamila.com.br'],
                'cc': ['copia@transcamila.com.br'],
            },
            format='json',
            **auth_headers(self.user, 'RH'),
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(len(mail.outbox), 1)
        sent = mail.outbox[0]
        self.assertEqual(sent.to, ['destino1@transcamila.com.br', 'destino2@transcamila.com.br'])
        self.assertEqual(sent.cc, ['copia@transcamila.com.br'])

        self.assertEqual(len(sent.attachments), 2)
        nomes_anexos = {nome for nome, _conteudo, _mimetype in sent.attachments}
        self.assertIn(f'Relatorio_Alteracoes_{self.lote.mes:02d}_{self.lote.ano}.xlsx', nomes_anexos)
        self.assertIn(f'Relatorio_Movimentacoes_{self.lote.mes:02d}_{self.lote.ano}.pdf', nomes_anexos)

        pdf_nome, pdf_conteudo, pdf_mimetype = next(
            a for a in sent.attachments if a[0] == f'Relatorio_Movimentacoes_{self.lote.mes:02d}_{self.lote.ano}.pdf'
        )
        self.assertEqual(pdf_mimetype, 'application/pdf')
        self.assertTrue(pdf_conteudo.startswith(b'%PDF'))

    def test_enviar_email_com_historico_anexa_grafico_folha_inline(self):
        """Com mais de um lote no histórico, o e-mail deve anexar o gráfico de evolução
        da folha como imagem inline (Content-ID), pois clientes de e-mail (Gmail, etc.)
        bloqueiam imagens em base64 embutidas diretamente no HTML."""
        lote_anterior = LoteMovimentacaoRH.objects.create(mes=10, ano=2026, usuario=self.user, data_importacao=timezone.now())
        MovimentacaoColaborador.objects.create(
            lote=lote_anterior,
            filial='Ibiporã (Matriz)',
            nome='Fulano de Tal',
            situacao='ATIVO',
            funcao='Motorista',
            cpf='12345678900',
            salario='2400.00',
            categoria='MOTORISTA',
        )

        response = self.client.post(
            f'/api/rh/lotes/{self.lote.id}/enviar_email/',
            data={'to': ['destino1@transcamila.com.br']},
            format='json',
            **auth_headers(self.user, 'RH'),
        )

        self.assertEqual(response.status_code, 200, response.data)
        sent = mail.outbox[0]

        from email.mime.image import MIMEImage
        imagens_inline = [a for a in sent.attachments if isinstance(a, MIMEImage)]
        self.assertEqual(len(imagens_inline), 1)
        self.assertEqual(imagens_inline[0]['Content-ID'], '<grafico_folha>')
        self.assertIn('inline', imagens_inline[0]['Content-Disposition'])
        self.assertIn('cid:grafico_folha', sent.body)

    def test_enviar_email_sem_destinatario_retorna_erro_400(self):
        response = self.client.post(
            f'/api/rh/lotes/{self.lote.id}/enviar_email/',
            data={'to': [], 'cc': []},
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('error', response.data)

    def test_enviar_email_quadro_pessoal_exibe_grupos_situacao(self):
        MovimentacaoColaborador.objects.create(
            lote=self.lote,
            filial='Rondonópolis',
            nome='Afastado Temporario',
            situacao='AFASTADO TEMP.',
            funcao='Auxiliar',
            cpf='11111111111',
            salario='1800.00',
            categoria='OPERACIONAL',
        )
        MovimentacaoColaborador.objects.create(
            lote=self.lote,
            filial='Ibiporã (Matriz)',
            nome='Em Ferias',
            situacao='FERIAS',
            funcao='Analista',
            cpf='22222222222',
            salario='3000.00',
            categoria='ADMINISTRATIVO',
        )

        response = self.client.post(
            f'/api/rh/lotes/{self.lote.id}/enviar_email/',
            data={'to': ['destino@transcamila.com.br']},
            format='json',
            **auth_headers(self.user, 'RH'),
        )

        self.assertEqual(response.status_code, 200, response.data)
        body = mail.outbox[0].body
        self.assertIn('Total de Colaboradores', body)
        self.assertIn('personnel-label">Administrativo', body)
        self.assertIn('personnel-label">Operacional', body)
        self.assertIn('personnel-label">Motorista', body)
        self.assertIn('<strong>1</strong> ativos', body)
        self.assertIn('<strong>1</strong> afastados', body)
        self.assertIn('<strong>1</strong> em f', body)
        self.assertIn('personnel-value text-blue">3<', body)
        self.assertIn('personnel-value text-blue">1<', body)


class MesesAfastadoTests(TestCase):
    """A planilha mensal não guarda a data em que o afastamento começou, então
    o tempo de afastamento é inferido andando mês a mês pelos lotes anteriores
    enquanto o mesmo CPF continuar com situação de afastamento."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='rh_afastado_tests', password='rh123', role_id='2', environments=['RH'], filiais={},
        )
        self.cpf = '44444444444'

    def _criar_lote(self, mes, ano, situacao):
        lote = LoteMovimentacaoRH.objects.create(mes=mes, ano=ano, usuario=self.user, data_importacao=timezone.now())
        MovimentacaoColaborador.objects.create(
            lote=lote, filial='Ibiporã (Matriz)', nome='Colaborador Afastado', situacao=situacao,
            funcao='Auxiliar', cpf=self.cpf, salario='1800.00', categoria='OPERACIONAL',
        )
        return lote

    def test_conta_meses_consecutivos_afastado_incluindo_mes_atual(self):
        self._criar_lote(2, 2026, 'ATIVO')
        self._criar_lote(3, 2026, 'AFASTADO TEMP.')
        self._criar_lote(4, 2026, 'AFASTADO TEMP.')
        lote_atual = self._criar_lote(5, 2026, 'AFASTADO INSS')

        self.assertEqual(_meses_afastado(self.cpf, lote_atual), 3)

    def test_para_quando_encontra_lote_sem_o_colaborador(self):
        self._criar_lote(4, 2026, 'AFASTADO TEMP.')
        lote_atual = self._criar_lote(5, 2026, 'AFASTADO TEMP.')
        # Sem lote de 03/2026: a contagem deve parar em 2 (mês atual + anterior).

        self.assertEqual(_meses_afastado(self.cpf, lote_atual), 2)

    def test_para_quando_colaborador_estava_ativo_no_mes_anterior(self):
        self._criar_lote(4, 2026, 'ATIVO')
        lote_atual = self._criar_lote(5, 2026, 'AFASTADO TEMP.')

        self.assertEqual(_meses_afastado(self.cpf, lote_atual), 1)

    def test_email_exibe_tempo_de_afastamento_do_colaborador(self):
        self._criar_lote(3, 2026, 'AFASTADO TEMP.')
        self._criar_lote(4, 2026, 'AFASTADO TEMP.')
        lote_atual = self._criar_lote(5, 2026, 'AFASTADO TEMP.')

        client = APIClient()
        response = client.post(
            f'/api/rh/lotes/{lote_atual.id}/enviar_email/',
            data={'to': ['destino@transcamila.com.br']},
            format='json',
            **auth_headers(self.user, 'RH'),
        )

        self.assertEqual(response.status_code, 200, response.data)
        sent = mail.outbox[0]
        self.assertIn('há 3 meses', sent.body)


class DashboardSummaryTotaisGeraisTests(TestCase):
    """Regressão: o card 'Total Geral' do dashboard depende de `totais.mediaIdade`,
    `totais.mediaTempo` e `totais.payroll`, agregados a partir de todas as filiais."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='rh_dashboard_tests',
            password='rh123',
            role_id='2',
            environments=['RH'],
            filiais={},
        )
        self.lote = LoteMovimentacaoRH.objects.create(mes=11, ano=2026, usuario=self.user, data_importacao=timezone.now())
        MovimentacaoColaborador.objects.create(
            lote=self.lote, filial='Ibiporã (Matriz)', nome='Fulano', situacao='ATIVO',
            funcao='Motorista', cpf='11111111111', salario='2000.00', categoria='MOTORISTA',
        )
        MovimentacaoColaborador.objects.create(
            lote=self.lote, filial='Rondonópolis', nome='Ciclana', situacao='ATIVO',
            funcao='Analista', cpf='22222222222', salario='4000.00', categoria='ADMINISTRATIVO',
        )

    def test_totais_geral_soma_todas_as_filiais(self):
        response = self.client.get(
            f'/api/rh/movimentacoes/dashboard_summary/?loteId={self.lote.id}',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(response.status_code, 200, response.data)
        totais = response.data['totais']
        self.assertEqual(totais['totalColaboradores'], 2)
        self.assertEqual(float(totais['payroll']), 6000.00)
        self.assertIn('mediaIdade', totais)
        self.assertIn('mediaTempo', totais)


class ExportarRelatorioMovimentacoesTests(TestCase):
    """O botão 'Exportar Relatório de Movimentações' deve gerar um Excel com uma linha
    por ocorrência (contratação, demissão, alteração), incluindo a filial de cada uma."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='rh_export_tests', password='rh123', role_id='2', environments=['RH'], filiais={},
        )
        self.lote_anterior = LoteMovimentacaoRH.objects.create(mes=10, ano=2026, usuario=self.user, data_importacao=timezone.now())
        MovimentacaoColaborador.objects.create(
            lote=self.lote_anterior, filial='Rondonópolis', nome='Colaborador Desligado', situacao='ATIVO',
            funcao='Auxiliar', cpf='33333333333', salario='1800.00', categoria='OPERACIONAL',
        )

        self.lote = LoteMovimentacaoRH.objects.create(mes=11, ano=2026, usuario=self.user, data_importacao=timezone.now())
        MovimentacaoColaborador.objects.create(
            lote=self.lote, filial='Ibiporã (Matriz)', nome='Colaborador Novo', situacao='ATIVO',
            funcao='Motorista', cpf='11111111111', salario='2000.00', categoria='MOTORISTA',
            data_admissao='2026-11-03',
        )
        InconsistenciaColaborador.objects.create(
            lote=self.lote, cpf='11111111111', nome='Colaborador Novo', tipo='salario',
            valor_anterior='2000.00', valor_atual='2500.00',
        )

    def test_exportar_relatorio_retorna_planilha_com_ocorrencias(self):
        response = self.client.get(
            f'/api/rh/movimentacoes/exportar_relatorio/?loteId={self.lote.id}',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], ('Filial', 'Colaborador', 'CPF', 'Ocorrência', 'Cargo', 'Valor Anterior', 'Valor Atual', 'Justificativa'))
        ocorrencias = {(row[0], row[1], row[3]) for row in rows[1:]}
        self.assertIn(('Ibiporã (Matriz)', 'COLABORADOR NOVO', 'Contratação'), ocorrencias)
        self.assertIn(('Rondonópolis', 'COLABORADOR DESLIGADO', 'Demissão'), ocorrencias)
        self.assertIn(('Ibiporã (Matriz)', 'COLABORADOR NOVO', 'Aumento de Salário'), ocorrencias)

    def test_exportar_relatorio_sem_lote_retorna_404(self):
        LoteMovimentacaoRH.objects.all().delete()
        response = self.client.get(
            '/api/rh/movimentacoes/exportar_relatorio/',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(response.status_code, 404)


class CleanCpfImportTests(TestCase):
    def test_clean_cpf_exige_11_digitos(self):
        self.assertEqual(clean_cpf('141.684.779-08'), '14168477908')
        self.assertIsNone(clean_cpf('0'))
        self.assertIsNone(clean_cpf('31716298850574713'))  # PIS/outro id
        self.assertIsNone(clean_cpf(''))
        self.assertIsNone(clean_cpf(None))

    def test_pick_sheet_prioriza_mes_e_ano_do_lote(self):
        wb = openpyxl.Workbook()
        # remove default
        default = wb.active
        wb.remove(default)
        for name, header in [
            ('ABRIL 2015', ['CPF', 'Nome', 'Salario']),
            ('JAN 2026', ['CPF', 'Nome', 'Salario']),
            ('ABRIL', ['CPF', 'Nome', 'Salario']),
            ('GRAFICOS', ['x']),
        ]:
            ws = wb.create_sheet(name)
            ws.append(header)
            ws.append(['123.456.789-00', 'Fulano', 1000])
        buffer = BytesIO()
        wb.save(buffer)
        loaded = openpyxl.load_workbook(BytesIO(buffer.getvalue()), read_only=True, data_only=True)

        _, abr = _pick_sheet_for_lote(loaded, LoteMovimentacaoRH(mes=4, ano=2026))
        self.assertEqual(abr, 'ABRIL')
        _, jan = _pick_sheet_for_lote(loaded, LoteMovimentacaoRH(mes=1, ano=2026))
        self.assertEqual(jan, 'JAN 2026')


class PjSyncNosLotesTests(TestCase):
    """PJ deve ser projetado nos lotes importados conforme admissão/demissão e histórico."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='rh_pj_sync_tests',
            password='rh123',
            role_id='2',
            environments=['RH'],
            filiais={},
        )
        self.lotes = {}
        for mes in (1, 2, 3, 4):
            self.lotes[mes] = LoteMovimentacaoRH.objects.create(
                mes=mes, ano=2026, usuario=self.user, data_importacao=timezone.now(),
            )

    def _cpfs_no_lote(self, mes):
        return set(
            MovimentacaoColaborador.objects.filter(lote=self.lotes[mes]).values_list('cpf', flat=True)
        )

    def test_pj_so_entra_a_partir_da_admissao(self):
        response = self.client.post(
            '/api/rh/pjs/',
            data={
                'nome': 'PJ Consultor',
                'cpf': '555.555.555-55',
                'salario': '5000.00',
                'filial': 'Ibiporã (Matriz)',
                'cargo': 'Consultor',
                'dataAdmissao': '2026-03-10',
                'ativo': True,
            },
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(response.status_code, 201, response.data)

        self.assertNotIn('555.555.555-55', self._cpfs_no_lote(1))
        self.assertNotIn('555.555.555-55', self._cpfs_no_lote(2))
        self.assertIn('555.555.555-55', self._cpfs_no_lote(3))
        self.assertIn('555.555.555-55', self._cpfs_no_lote(4))

        mov_mar = MovimentacaoColaborador.objects.get(lote=self.lotes[3], cpf='555.555.555-55')
        self.assertEqual(mov_mar.situacao, 'ATIVO (PJ)')
        self.assertEqual(Decimal(mov_mar.salario), Decimal('5000.00'))

    def test_historico_salarial_aplica_por_competencia(self):
        pj = ColaboradorPJ.objects.create(
            nome='PJ Histórico',
            cpf='666.666.666-66',
            salario=Decimal('4000.00'),
            filial='Ibiporã (Matriz)',
            cargo='Analista',
            data_admissao=date(2026, 1, 1),
            ativo=True,
        )
        ColaboradorPJHistorico.objects.create(pj=pj, ano=2026, mes=3, salario=Decimal('4500.00'))
        sync_pj_nos_lotes(pj)

        self.assertEqual(
            Decimal(MovimentacaoColaborador.objects.get(lote=self.lotes[2], cpf=pj.cpf).salario),
            Decimal('4000.00'),
        )
        self.assertEqual(
            Decimal(MovimentacaoColaborador.objects.get(lote=self.lotes[3], cpf=pj.cpf).salario),
            Decimal('4500.00'),
        )
        self.assertEqual(
            Decimal(MovimentacaoColaborador.objects.get(lote=self.lotes[4], cpf=pj.cpf).salario),
            Decimal('4500.00'),
        )

    def test_demissao_remove_meses_posteriores(self):
        response = self.client.post(
            '/api/rh/pjs/',
            data={
                'nome': 'PJ Temporário',
                'cpf': '777.777.777-77',
                'salario': '3000.00',
                'filial': 'Rondonópolis',
                'cargo': 'Apoio',
                'dataAdmissao': '2026-01-01',
                'dataDemissao': '2026-02-28',
                'ativo': True,
            },
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(response.status_code, 201, response.data)

        self.assertIn('777.777.777-77', self._cpfs_no_lote(1))
        self.assertIn('777.777.777-77', self._cpfs_no_lote(2))
        self.assertNotIn('777.777.777-77', self._cpfs_no_lote(3))
        self.assertNotIn('777.777.777-77', self._cpfs_no_lote(4))

    def test_nao_sobrescreve_linha_clt_do_mesmo_cpf(self):
        MovimentacaoColaborador.objects.create(
            lote=self.lotes[3],
            filial='Ibiporã (Matriz)',
            nome='CLT Existente',
            situacao='ATIVO',
            funcao='Motorista',
            cpf='888.888.888-88',
            salario='2000.00',
            categoria='MOTORISTA',
        )
        response = self.client.post(
            '/api/rh/pjs/',
            data={
                'nome': 'PJ Mesmo CPF',
                'cpf': '888.888.888-88',
                'salario': '9000.00',
                'filial': 'Ibiporã (Matriz)',
                'cargo': 'Consultor',
                'dataAdmissao': '2026-01-01',
                'ativo': True,
            },
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(response.status_code, 201, response.data)

        mov = MovimentacaoColaborador.objects.get(lote=self.lotes[3], cpf='888.888.888-88')
        self.assertEqual(mov.situacao, 'ATIVO')
        self.assertEqual(mov.nome, 'CLT Existente')
        self.assertEqual(Decimal(mov.salario), Decimal('2000.00'))
        # Em meses sem CLT, o PJ entra normalmente
        self.assertIn('888.888.888-88', self._cpfs_no_lote(1))
        self.assertEqual(
            MovimentacaoColaborador.objects.get(lote=self.lotes[1], cpf='888.888.888-88').situacao,
            'ATIVO (PJ)',
        )

    def test_cargo_pj_aparece_no_mapeamento_e_categoria_propaga(self):
        """Cargo do PJ deve entrar na classificação e, ao mapear, atualizar a movimentação."""
        create = self.client.post(
            '/api/rh/pjs/',
            data={
                'nome': 'PJ Cargo',
                'cpf': '121.212.212-12',
                'salario': '2000.00',
                'filial': 'Ibiporã (Matriz)',
                'cargo': 'Cargo Teste PJ',
                'dataAdmissao': '2026-01-01',
                'ativo': True,
            },
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(create.status_code, 201, create.data)

        cargos = self.client.get(
            '/api/rh/cargos/?status=pendente',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(cargos.status_code, 200, cargos.data)
        nomes = {c['cargo'] for c in cargos.data}
        self.assertIn('CARGO TESTE PJ', nomes)

        mapping_id = next(c['id'] for c in cargos.data if c['cargo'] == 'CARGO TESTE PJ')
        patch = self.client.patch(
            f'/api/rh/cargos/{mapping_id}/',
            data={'categoria': 'ADMINISTRATIVO'},
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(patch.status_code, 200, patch.data)

        mov = MovimentacaoColaborador.objects.get(lote=self.lotes[3], cpf='121.212.212-12')
        self.assertEqual(mov.categoria, 'ADMINISTRATIVO')
        self.assertEqual(mov.funcao, 'CARGO TESTE PJ')

    def test_api_historico_crud_e_resync(self):
        create = self.client.post(
            '/api/rh/pjs/',
            data={
                'nome': 'PJ API Hist',
                'cpf': '999.999.999-99',
                'salario': '1000.00',
                'filial': 'Ibiporã (Matriz)',
                'cargo': 'Analista',
                'dataAdmissao': '2026-01-01',
                'ativo': True,
            },
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(create.status_code, 201, create.data)
        pj_id = create.data['id']

        hist = self.client.post(
            f'/api/rh/pjs/{pj_id}/historico/',
            data={'ano': 2026, 'mes': 3, 'salario': '1500.00'},
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(hist.status_code, 201, hist.data)
        self.assertEqual(
            Decimal(MovimentacaoColaborador.objects.get(lote=self.lotes[3], cpf='999.999.999-99').salario),
            Decimal('1500.00'),
        )

        hid = hist.data['id']
        patch = self.client.patch(
            f'/api/rh/pjs/{pj_id}/historico/{hid}/',
            data={'salario': '1600.00'},
            format='json',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(patch.status_code, 200, patch.data)
        self.assertEqual(
            Decimal(MovimentacaoColaborador.objects.get(lote=self.lotes[3], cpf='999.999.999-99').salario),
            Decimal('1600.00'),
        )

        delete = self.client.delete(
            f'/api/rh/pjs/{pj_id}/historico/{hid}/',
            **auth_headers(self.user, 'RH'),
        )
        self.assertEqual(delete.status_code, 204)
        self.assertEqual(
            Decimal(MovimentacaoColaborador.objects.get(lote=self.lotes[3], cpf='999.999.999-99').salario),
            Decimal('1000.00'),
        )
