"""Importação de protocolos de envio a partir de planilha .xlsx.

Usado pela API (admin) e pelo management command `import_protocolos`.
"""

from __future__ import annotations

import re
from collections import OrderedDict
from datetime import date, datetime
from io import BytesIO

import openpyxl
from django.db import transaction

from apps.faturamento.constants import EXPEDICAO_VALUES
from apps.faturamento.models import ClienteProtocolo, ProtocoloEnvio
from apps.faturamento.services import (
    combinar_expedicoes,
    gerar_numero_sequencial,
    notas_fiscais_duplicadas,
    parse_notas_fiscais,
)

_CANDIDATOS_DATA = [
    'data de envio', 'data envio', 'data', 'date', 'dt envio', 'dt',
]
_CANDIDATOS_NF = [
    'nota fiscal', 'nota(s) fiscal(is)', 'notas fiscais', 'nf', 'nfs',
    'numero nf', 'numero nota', 'notas',
]
_CANDIDATOS_EXPEDICAO = ['expedicao', 'expedição', 'transportadora', 'transp', 'transporte']
_CANDIDATOS_FILIAL = [
    'filial', 'filial cliente', 'filial do cliente', 'unidade', 'loja',
    'branch', 'nome filial',
]
_CANDIDATOS_ANO = ['ano', 'year', 'ano protocolo']
_CANDIDATOS_NUMERO = [
    'numero protocolo', 'nº protocolo', 'n protocol', 'num protocolo',
    'nr protocolo', 'numero', 'protocolo',
]
_CANDIDATOS_CLIENTE = ['cliente', 'nome cliente', 'razao social', 'razão social']


class ProtocoloImportError(Exception):
    """Erro de estrutura/validação da planilha (não gravar)."""


def _norm(s: str) -> str:
    s = s.strip().lower()
    replacements = {
        'á': 'a', 'ã': 'a', 'â': 'a', 'à': 'a',
        'é': 'e', 'ê': 'e', 'í': 'i', 'ó': 'o', 'ô': 'o',
        'õ': 'o', 'ú': 'u', 'ü': 'u', 'ç': 'c', 'ñ': 'n',
        'º': 'o', 'ª': 'a',
    }
    for k, v in replacements.items():
        s = s.replace(k, v)
    return re.sub(r'\s+', ' ', s)


def _detectar_coluna(headers: list[str], candidatos: list[str]) -> str | None:
    normed = {_norm(h): h for h in headers if h}
    for c in candidatos:
        if c in normed:
            return normed[c]
    return None


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_nf(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _parse_int(value) -> int | None:
    if value is None or value == '':
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _match_expedicao(value: str) -> list[str]:
    if not value:
        return []
    matched = []
    for ev in EXPEDICAO_VALUES:
        if _norm(ev) in _norm(value) or _norm(value) in _norm(ev):
            matched.append(ev)
    return matched[:2]


def _filiais_index(cliente: ClienteProtocolo) -> dict[str, str]:
    """Mapa nome-normalizado → nome oficial cadastrado no cliente."""
    return {_norm(f.nome): f.nome for f in cliente.filiais.all()}


def _resolver_filial(
    raw: str,
    filiais_por_norm: dict[str, str],
) -> tuple[str | None, str | None]:
    """Resolve o nome da filial. Retorna (nome_resolvido, aviso_opcional)."""
    value = (raw or '').strip()
    if not value:
        return None, None
    key = _norm(value)
    if key in filiais_por_norm:
        return filiais_por_norm[key], None
    for cand, nome in filiais_por_norm.items():
        if key in cand or cand in key:
            return nome, None
    if filiais_por_norm:
        return value, f'filial "{value}" não cadastrada para o cliente — gravada como informada'
    return value, None


def _build_notas_filiais_from_raw(
    notas: list[str],
    filial_raw: str,
    filiais_por_norm: dict[str, str],
) -> tuple[dict[str, str], list[str]]:
    """Monta {nf: filial} a partir da coluna Filial (modo linha-a-linha).

    - Um valor → aplica a todas as NFs da linha
    - Vários valores separados por vírgula → associa na ordem das NFs
    """
    avisos: list[str] = []
    value = (filial_raw or '').strip()
    if not value or not notas:
        return {}, avisos

    partes = [p.strip() for p in value.split(',') if p.strip()]
    mapping: dict[str, str] = {}

    if len(partes) == 1:
        nome, aviso = _resolver_filial(partes[0], filiais_por_norm)
        if aviso:
            avisos.append(aviso)
        if nome:
            for nf in notas:
                mapping[nf] = nome
        return mapping, avisos

    for i, nf in enumerate(notas):
        if i >= len(partes):
            break
        nome, aviso = _resolver_filial(partes[i], filiais_por_norm)
        if aviso:
            avisos.append(f'NF {nf}: {aviso}')
        if nome:
            mapping[nf] = nome

    if len(partes) != len(notas):
        avisos.append(
            f'quantidade de filiais ({len(partes)}) difere das NFs ({len(notas)}); '
            'associação feita na ordem disponível'
        )
    return mapping, avisos


def _corrigir_data_ano(data: date, ano: int | None) -> tuple[date, str | None]:
    if ano is None or data.year == ano:
        return data, None
    corrigida = data.replace(year=ano)
    aviso = (
        f'Data {data.isoformat()} ajustada para {corrigida.isoformat()} '
        f'para bater com o Ano={ano} do protocolo'
    )
    return corrigida, aviso


def _load_worksheet(file_bytes: bytes, sheet: str = ''):
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ProtocoloImportError(f'Erro ao abrir o arquivo: {exc}') from exc

    sheet_name = (sheet or '').strip()
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ProtocoloImportError(
                f'Aba "{sheet_name}" não encontrada. Abas disponíveis: {", ".join(wb.sheetnames)}'
            )
        return wb[sheet_name]
    return wb.active


def _read_headers(ws) -> tuple[list[str], int]:
    headers_row = None
    header_row_num = 1
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [c.value for c in row]
        if any(v is not None for v in values):
            headers_row = [str(c).strip() if c is not None else '' for c in values]
            header_row_num = row[0].row
            break

    if not headers_row:
        raise ProtocoloImportError('Não foi possível detectar cabeçalho no arquivo.')

    while headers_row and not headers_row[-1]:
        headers_row.pop()
    return headers_row, header_row_num


def _resolve_columns(
    headers_row: list[str],
    *,
    col_data: str = '',
    col_nf: str = '',
    col_expedicao: str = '',
    col_filial: str = '',
) -> dict:
    resolved_data = (col_data or '').strip() or _detectar_coluna(headers_row, _CANDIDATOS_DATA)
    resolved_nf = (col_nf or '').strip() or _detectar_coluna(headers_row, _CANDIDATOS_NF)
    resolved_exp = (col_expedicao or '').strip() or _detectar_coluna(headers_row, _CANDIDATOS_EXPEDICAO)
    resolved_filial = (col_filial or '').strip() or _detectar_coluna(headers_row, _CANDIDATOS_FILIAL)
    resolved_ano = _detectar_coluna(headers_row, _CANDIDATOS_ANO)
    resolved_numero = _detectar_coluna(headers_row, _CANDIDATOS_NUMERO)

    if not resolved_data:
        raise ProtocoloImportError(
            f'Coluna de data não detectada. Colunas: {headers_row}'
        )
    if not resolved_nf:
        raise ProtocoloImportError(
            f'Coluna de nota fiscal não detectada. Colunas: {headers_row}'
        )

    return {
        'data': resolved_data,
        'notaFiscal': resolved_nf,
        'expedicao': resolved_exp,
        'filial': resolved_filial,
        'ano': resolved_ano,
        'numero': resolved_numero,
    }


def _agrupar_por_numero(raw_rows: list[dict]) -> list[dict]:
    grupos: OrderedDict[tuple[int, int], dict] = OrderedDict()
    for row in raw_rows:
        ano = row['ano']
        numero = row['numero']
        if ano is None or numero is None:
            continue
        key = (ano, numero)
        if key not in grupos:
            grupos[key] = {
                'ano': ano,
                'numero': numero,
                'datas': [],
                'nfs': [],
                'nf_filiais': {},  # nf -> filial_raw (primeira ocorrência)
                'exp_raw': '',
                'avisos': [],
            }
        g = grupos[key]
        if row['data']:
            g['datas'].append(row['data'])
        if row['nf_raw']:
            g['nfs'].append(row['nf_raw'])
            filial_raw = (row.get('filial_raw') or '').strip()
            if filial_raw and row['nf_raw'] not in g['nf_filiais']:
                g['nf_filiais'][row['nf_raw']] = filial_raw
        if row['exp_raw'] and not g['exp_raw']:
            g['exp_raw'] = row['exp_raw']

    protocolos = []
    for (ano, numero), g in grupos.items():
        if not g['datas']:
            protocolos.append({
                'label': f'{ano}-{numero:04d}',
                'data': None,
                'numero': numero,
                'nf_raw': ', '.join(g['nfs']),
                'exp_raw': g['exp_raw'],
                'filial_raw': '',
                'nf_filiais': dict(g['nf_filiais']),
                'avisos': g['avisos'],
            })
            continue

        data = min(g['datas'])
        data, aviso = _corrigir_data_ano(data, ano)
        if aviso:
            g['avisos'].append(aviso)
        datas_unicas = sorted(set(g['datas']))
        if len(datas_unicas) > 1:
            g['avisos'].append(
                f'Múltiplas datas no protocolo ({", ".join(d.isoformat() for d in datas_unicas)}); '
                f'usando {data.isoformat()}'
            )

        nfs_unicas: list[str] = []
        vistas: set[str] = set()
        for nf in g['nfs']:
            if nf not in vistas:
                vistas.add(nf)
                nfs_unicas.append(nf)

        protocolos.append({
            'label': f'{ano}-{numero:04d}',
            'data': data,
            'numero': numero,
            'nf_raw': ', '.join(nfs_unicas),
            'exp_raw': g['exp_raw'],
            'filial_raw': '',
            'nf_filiais': {nf: g['nf_filiais'][nf] for nf in nfs_unicas if nf in g['nf_filiais']},
            'avisos': g['avisos'],
        })
    return protocolos


def _linhas_como_protocolos(raw_rows: list[dict]) -> list[dict]:
    protocolos = []
    for i, row in enumerate(raw_rows, start=1):
        protocolos.append({
            'label': f'linha/{i}',
            'data': row['data'],
            'numero': None,
            'nf_raw': row['nf_raw'],
            'exp_raw': row['exp_raw'],
            'filial_raw': row.get('filial_raw') or '',
            'nf_filiais': {},
            'avisos': [],
        })
    return protocolos


def _parse_raw_rows(ws, headers_row: list[str], header_row_num: int, mapping: dict) -> list[dict]:
    idx_data = headers_row.index(mapping['data'])
    idx_nf = headers_row.index(mapping['notaFiscal'])
    idx_exp = (
        headers_row.index(mapping['expedicao'])
        if mapping['expedicao'] and mapping['expedicao'] in headers_row
        else None
    )
    idx_filial = (
        headers_row.index(mapping['filial'])
        if mapping.get('filial') and mapping['filial'] in headers_row
        else None
    )
    idx_ano = (
        headers_row.index(mapping['ano'])
        if mapping['ano'] and mapping['ano'] in headers_row
        else None
    )
    idx_numero = (
        headers_row.index(mapping['numero'])
        if mapping['numero'] and mapping['numero'] in headers_row
        else None
    )

    raw_rows = []
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        if all(v is None for v in row):
            continue
        raw_rows.append({
            'data': _parse_date(row[idx_data] if idx_data < len(row) else None),
            'nf_raw': _parse_nf(row[idx_nf] if idx_nf < len(row) else None),
            'exp_raw': (
                str(row[idx_exp]).strip()
                if idx_exp is not None and idx_exp < len(row) and row[idx_exp]
                else ''
            ),
            'filial_raw': (
                str(row[idx_filial]).strip()
                if idx_filial is not None and idx_filial < len(row) and row[idx_filial]
                else ''
            ),
            'ano': _parse_int(row[idx_ano] if idx_ano is not None and idx_ano < len(row) else None),
            'numero': _parse_int(
                row[idx_numero] if idx_numero is not None and idx_numero < len(row) else None
            ),
        })
    return raw_rows


def build_protocolo_import_template() -> bytes:
    """Gera planilha de referência (.xlsx) no formato oficial de importação."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Aba Importação (formato oficial) ──────────────────────────────────────
    ws = wb.active
    ws.title = 'Importação'

    # Ordem alinhada ao modelo operacional usado pela equipe.
    headers = [
        'Ano',
        'Numero Protocolo',
        'Expedição',
        'Data de envio',
        'Cliente',
        'Nota Fiscal',
    ]
    header_fill = PatternFill(fill_type='solid', fgColor='118CC4')
    header_font = Font(bold=True, color='FFFFFF')
    example_fill = PatternFill(fill_type='solid', fgColor='F0F9FF')
    note_font = Font(italic=True, color='64748B', size=10)

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    # Exemplos: Expedição pode ficar vazia; mesmo Ano+Numero = mesmo protocolo.
    exemplos = [
        (2025, 1, '', '04/02/2025', 'Ascenza brasil ltda.', '44'),
        (2025, 1, '', '04/02/2025', 'Ascenza brasil ltda.', '34'),
        (2025, 1, '', '04/02/2025', 'Ascenza brasil ltda.', '36'),
        (2025, 2, 'Transcamila Ibiporã', '10/02/2025', 'Ascenza brasil ltda.', '100'),
        (2026, 1, '', '15/01/2026', 'Cliente Exemplo', '2001'),
        (2026, 1, '', '15/01/2026', 'Cliente Exemplo', '2002'),
    ]
    for row in exemplos:
        ws.append(list(row))
        for col in range(1, len(headers) + 1):
            ws.cell(ws.max_row, col).fill = example_fill

    ws.append([])
    ws.append([
        'Os exemplos acima são apenas referência — apague-os e preencha com os dados reais '
        'antes de importar. Expedição é opcional (pode ficar em branco). Filial também é '
        'opcional: se precisar, adicione uma coluna "Filial".'
    ])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
    ws.cell(ws.max_row, 1).font = note_font
    ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True)

    widths = [8, 18, 24, 14, 28, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Aba Instruções ────────────────────────────────────────────────────────
    wi = wb.create_sheet('Instruções')
    wi['A1'] = 'Como preencher a planilha de importação de protocolos'
    wi['A1'].font = Font(bold=True, size=14, color='0F172A')
    wi.merge_cells('A1:B1')

    instrucoes = [
        ('', ''),
        ('Colunas do modelo', ''),
        ('Ano', 'Ano do protocolo (ex.: 2025). Usado com "Numero Protocolo" para agrupar NFs.'),
        (
            'Numero Protocolo',
            'Número sequencial do protocolo. Linhas com o mesmo Ano + Numero Protocolo '
            'viram um único protocolo (várias NFs).',
        ),
        (
            'Expedição',
            'Opcional. Pode ficar em branco. Valores reconhecidos quando informados: '
            'Transcamila Ibiporã, Transcamila Barueri, Transcamila Paranaguá, '
            'Transcamila Rondonópolis.',
        ),
        (
            'Data de envio',
            'Obrigatória. Formatos aceitos: DD/MM/AAAA, AAAA-MM-DD, DD-MM-AAAA '
            '(ou data nativa do Excel).',
        ),
        (
            'Cliente',
            'Informativo na planilha. Na tela de importação, selecione o cliente no sistema '
            '(a coluna não substitui essa seleção).',
        ),
        (
            'Nota Fiscal',
            'Obrigatória. Uma NF por linha (recomendado) ou várias separadas por vírgula. '
            'Máximo de 72 por protocolo.',
        ),
        ('', ''),
        ('Coluna extra opcional', ''),
        (
            'Filial',
            'Opcional. Só use se o cliente tiver filiais cadastradas. Pode ficar ausente '
            'ou em branco — a importação não exige filial.',
        ),
        ('', ''),
        ('Dicas', ''),
        (
            'Expedição / Filial',
            'Nunca são obrigatórias na importação. Se o cliente estiver marcado como '
            '"requer expedição" ou "exige filial", a falta gera apenas aviso.',
        ),
        (
            'Agrupamento',
            'Com Ano + Numero Protocolo preenchidos, várias linhas do mesmo grupo '
            'formam um protocolo. Sem essas colunas, cada linha vira um protocolo.',
        ),
        (
            'Dry-run',
            'Na tela de importação, use "Simular importação" para validar sem gravar no banco.',
        ),
    ]

    row_num = 3
    section_font = Font(bold=True, size=12, color='118CC4')
    label_font = Font(bold=True, color='334155')
    for titulo, texto in instrucoes:
        if titulo and not texto:
            wi.cell(row_num, 1, titulo).font = section_font
        elif titulo:
            wi.cell(row_num, 1, titulo).font = label_font
            wi.cell(row_num, 2, texto).alignment = Alignment(wrap_text=True, vertical='top')
        row_num += 1

    wi.column_dimensions['A'].width = 28
    wi.column_dimensions['B'].width = 90
    for r in range(3, row_num):
        wi.row_dimensions[r].height = 36

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def inspect_protocolo_workbook(file_bytes: bytes, *, sheet: str = '') -> dict:
    ws = _load_worksheet(file_bytes, sheet)
    headers_row, _ = _read_headers(ws)
    mapping = {
        'data': _detectar_coluna(headers_row, _CANDIDATOS_DATA),
        'notaFiscal': _detectar_coluna(headers_row, _CANDIDATOS_NF),
        'expedicao': _detectar_coluna(headers_row, _CANDIDATOS_EXPEDICAO),
        'filial': _detectar_coluna(headers_row, _CANDIDATOS_FILIAL),
        'ano': _detectar_coluna(headers_row, _CANDIDATOS_ANO),
        'numero': _detectar_coluna(headers_row, _CANDIDATOS_NUMERO),
        'cliente': _detectar_coluna(headers_row, _CANDIDATOS_CLIENTE),
    }
    grouping_mode = (
        'grouped' if mapping['ano'] and mapping['numero'] else 'row_by_row'
    )
    return {
        'sheetName': ws.title,
        'columns': headers_row,
        'detectedMapping': mapping,
        'groupingMode': grouping_mode,
    }


def import_protocolos_from_workbook(
    file_bytes: bytes,
    *,
    cliente: ClienteProtocolo,
    dry_run: bool = False,
    skip_duplicatas: bool = False,
    sheet: str = '',
    col_data: str = '',
    col_nf: str = '',
    col_expedicao: str = '',
    col_filial: str = '',
    usuario=None,
    usuario_nome: str = 'Importação Excel',
    file_name: str = '',
) -> dict:
    """Importa protocolos a partir dos bytes de um .xlsx.

    Expedição e filial são opcionais: se houver colunas na planilha, são gravadas.
    NFs/números já existentes são ignorados com aviso (não derrubam o lote).
    Flags ``cliente.requer_expedicao`` / ``exige_filial`` não bloqueiam a importação.

    Retorna dict camelCase pronto para a API. Em falha estrutural, levanta
    ProtocoloImportError. Em falha de negócio com rollback, `success=False`.
    """
    _ = skip_duplicatas  # mantido por compatibilidade da API; duplicatas sempre são ignoradas
    ws = _load_worksheet(file_bytes, sheet)
    headers_row, header_row_num = _read_headers(ws)
    mapping = _resolve_columns(
        headers_row,
        col_data=col_data,
        col_nf=col_nf,
        col_expedicao=col_expedicao,
        col_filial=col_filial,
    )
    raw_rows = _parse_raw_rows(ws, headers_row, header_row_num, mapping)

    agrupar = bool(mapping['ano'] and mapping['numero'])
    grouping_mode = 'grouped' if agrupar else 'row_by_row'
    protocolos = _agrupar_por_numero(raw_rows) if agrupar else _linhas_como_protocolos(raw_rows)

    criados = 0
    ignorados = 0
    erros: list[dict] = []
    avisos: list[dict] = []
    max_numero = cliente.ultimo_numero_protocolo
    filiais_por_norm = _filiais_index(cliente)

    # Expedição e filial são sempre opcionais na importação — sem avisos de
    # "cliente requer…". O cadastro manual no formulário continua respeitando
    # os flags do cliente.

    with transaction.atomic():
        for proto in protocolos:
            label = proto['label']
            for aviso in proto.get('avisos') or []:
                avisos.append({'label': label, 'message': aviso})

            if not proto['data']:
                erros.append({'label': label, 'message': 'data inválida ou ausente'})
                continue
            if not proto['nf_raw']:
                erros.append({'label': label, 'message': 'nota fiscal ausente'})
                continue

            try:
                notas = parse_notas_fiscais(proto['nf_raw'])
            except ValueError as exc:
                erros.append({'label': label, 'message': str(exc)})
                continue

            duplicadas = notas_fiscais_duplicadas(cliente=cliente, notas=notas)
            if duplicadas:
                msg = (
                    f'NFs já cadastradas para "{cliente.nome}": '
                    f'{", ".join(duplicadas)}'
                )
                # Duplicatas nunca derrubam o lote: remove NFs repetidas ou
                # ignora o protocolo se não sobrar nenhuma.
                notas = [nf for nf in notas if nf not in set(duplicadas)]
                if not notas:
                    avisos.append({
                        'label': label,
                        'message': f'ignorado — {msg}',
                    })
                    ignorados += 1
                    continue
                avisos.append({
                    'label': label,
                    'message': f'{msg} — removidas do protocolo',
                })

            # ── Expedição (opcional) ──────────────────────────────────────────
            expedicoes: list[str] = []
            if proto.get('exp_raw'):
                expedicoes = _match_expedicao(proto['exp_raw'])
                if not expedicoes:
                    avisos.append({
                        'label': label,
                        'message': f'expedição "{proto["exp_raw"]}" não reconhecida — ignorada',
                    })
            expedicao_combinada = combinar_expedicoes(expedicoes) if expedicoes else None

            # ── Filiais por NF (opcional) ─────────────────────────────────────
            notas_filiais: dict[str, str] = {}
            nf_filiais_raw = proto.get('nf_filiais') or {}
            if nf_filiais_raw:
                for nf in notas:
                    raw_filial = nf_filiais_raw.get(nf) or ''
                    if not raw_filial:
                        continue
                    nome, aviso = _resolver_filial(raw_filial, filiais_por_norm)
                    if aviso:
                        avisos.append({'label': label, 'message': f'NF {nf}: {aviso}'})
                    if nome:
                        notas_filiais[nf] = nome
            elif proto.get('filial_raw'):
                built, build_avisos = _build_notas_filiais_from_raw(
                    notas, proto['filial_raw'], filiais_por_norm,
                )
                notas_filiais = built
                for msg in build_avisos:
                    avisos.append({'label': label, 'message': msg})

            nota_normalizada = ', '.join(notas)
            numero_seq = proto['numero']

            if numero_seq is not None:
                if ProtocoloEnvio.objects.filter(
                    cliente=cliente, numero_sequencial=numero_seq,
                ).exists():
                    msg = f'número sequencial {numero_seq} já existe para "{cliente.nome}"'
                    avisos.append({'label': label, 'message': f'ignorado — {msg}'})
                    ignorados += 1
                    continue

            if not dry_run:
                if numero_seq is None:
                    numero_seq = gerar_numero_sequencial(cliente)
                ProtocoloEnvio.objects.create(
                    data=proto['data'],
                    cliente=cliente,
                    nota_fiscal=nota_normalizada,
                    numero_sequencial=numero_seq,
                    notas_filiais=notas_filiais,
                    expedicao=expedicao_combinada,
                    usuario=usuario,
                    usuario_nome=usuario_nome,
                )
                max_numero = max(max_numero, numero_seq)
            else:
                if numero_seq is not None:
                    max_numero = max(max_numero, numero_seq)

            criados += 1

        if not dry_run and criados and max_numero > cliente.ultimo_numero_protocolo:
            cliente.ultimo_numero_protocolo = max_numero
            cliente.save(update_fields=['ultimo_numero_protocolo'])

        # Só erros estruturais (data/NF inválida) fazem rollback do lote.
        rolled_back = bool(erros and not dry_run)
        if rolled_back:
            transaction.set_rollback(True)
            criados = 0

    success = not erros

    result = {
        'success': success,
        'dryRun': dry_run,
        'fileName': file_name or '',
        'clienteId': cliente.pk,
        'clienteNome': cliente.nome,
        'sheetName': ws.title,
        'groupingMode': grouping_mode,
        'detectedMapping': mapping,
        'created': criados,
        'ignored': ignorados,
        'warnings': avisos,
        'errors': erros,
    }
    if not success and not dry_run:
        result['detail'] = (
            'Nenhum protocolo foi importado por erros na planilha. '
            'Corrija datas/NFs inválidas e tente novamente.'
        )
    return result
