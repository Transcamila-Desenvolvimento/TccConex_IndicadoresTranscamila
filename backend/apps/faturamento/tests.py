from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient

from apps.accounts.tests import auth_headers
from apps.faturamento.models import ClienteProtocolo, FilialClienteProtocolo, ProtocoloEnvio

User = get_user_model()


class FaturamentoProtocoloTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username='fat.user',
            password='test123',
            name='Usuário Faturamento',
            role_id='2',
            environments=['Faturamento'],
            filiais={'Faturamento': ['Ibiporã (Matriz)']},
            funcoes={'Faturamento': ['criar-protocolos', 'editar-protocolos']},
        )
        self.user_sem_funcoes = User.objects.create_user(
            username='fat.leitor',
            password='test123',
            name='Operador Sem Funções',
            role_id='2',
            environments=['Faturamento'],
            filiais={'Faturamento': ['Ibiporã (Matriz)']},
        )
        self.admin = User.objects.create_user(
            username='fat.admin.main',
            password='test123',
            name='Admin Faturamento',
            role_id='1',
            environments=['Faturamento', 'Administração'],
            filiais={'Faturamento': ['Ibiporã (Matriz)']},
        )
        self.user_excluidor = User.objects.create_user(
            username='fat.excluidor',
            password='test123',
            name='Operador com Exclusão',
            role_id='2',
            environments=['Faturamento'],
            filiais={'Faturamento': ['Ibiporã (Matriz)']},
            funcoes={'Faturamento': ['excluir-protocolos']},
        )
        self.cliente = ClienteProtocolo.objects.create(
            nome='Cliente Teste',
            requer_expedicao=True,
        )
        self.other_cliente = ClienteProtocolo.objects.create(nome='Sem Expedição')
        self.api = APIClient()

    def test_operador_nao_pode_criar_cliente(self):
        response = self.api.post(
            '/api/faturamento/protocolo-clientes/',
            {'nome': 'Novo Cliente', 'requerExpedicao': False},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(ClienteProtocolo.objects.filter(nome='Novo Cliente').exists())

    def test_operador_nao_pode_atualizar_cliente(self):
        response = self.api.patch(
            f'/api/faturamento/protocolo-clientes/{self.cliente.pk}/',
            {'nome': 'Cliente Alterado'},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)
        self.cliente.refresh_from_db()
        self.assertEqual(self.cliente.nome, 'Cliente Teste')

    def test_operador_nao_pode_excluir_cliente(self):
        response = self.api.delete(
            f'/api/faturamento/protocolo-clientes/{self.other_cliente.pk}/',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)
        self.assertTrue(ClienteProtocolo.objects.filter(pk=self.other_cliente.pk).exists())

    def test_admin_pode_criar_cliente(self):
        response = self.api.post(
            '/api/faturamento/protocolo-clientes/',
            {'nome': 'Cliente Admin', 'requerExpedicao': False},
            format='json',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 201)
        self.assertTrue(ClienteProtocolo.objects.filter(nome='Cliente Admin').exists())

    def test_operador_nao_pode_criar_filial_cliente(self):
        response = self.api.post(
            f'/api/faturamento/protocolo-clientes/{self.cliente.pk}/filiais/',
            {'nome': 'Filial SP'},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(FilialClienteProtocolo.objects.filter(cliente=self.cliente).count(), 0)

    def test_criar_protocolo_com_expedicao_obrigatoria(self):
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {
                'data': '2026-07-10',
                'clienteId': self.cliente.pk,
                'notaFiscal': '1001, 1002',
                'expedicoes': ['Transcamila Ibiporã'],
            },
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(ProtocoloEnvio.objects.count(), 1)
        protocolo = ProtocoloEnvio.objects.first()
        self.assertEqual(protocolo.nota_fiscal, '1001, 1002')
        self.assertEqual(protocolo.expedicao, 'Transcamila Ibiporã')

    def test_criar_protocolo_com_duas_expedicoes_combina_valores(self):
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {
                'data': '2026-07-10',
                'clienteId': self.cliente.pk,
                'notaFiscal': '1001',
                'expedicoes': ['Transcamila Barueri', 'Transcamila Ibiporã'],
            },
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 201)
        protocolo = ProtocoloEnvio.objects.get()
        self.assertEqual(protocolo.expedicao, 'Transcamila Barueri/Ibiporã')
        self.assertEqual(response.data['expedicoes'], ['Transcamila Barueri', 'Transcamila Ibiporã'])

    def test_criar_protocolo_com_mais_de_duas_expedicoes_retorna_400(self):
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {
                'data': '2026-07-10',
                'clienteId': self.cliente.pk,
                'notaFiscal': '1001',
                'expedicoes': [
                    'Transcamila Barueri',
                    'Transcamila Ibiporã',
                    'Transcamila Paranaguá',
                ],
            },
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(ProtocoloEnvio.objects.count(), 0)

    def test_criar_protocolo_sem_expedicao_quando_obrigatorio_retorna_400(self):
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {
                'data': '2026-07-10',
                'clienteId': self.cliente.pk,
                'notaFiscal': '1001',
            },
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(ProtocoloEnvio.objects.count(), 0)

    def test_listar_protocolos_filtra_por_nota_fiscal(self):
        ProtocoloEnvio.objects.create(
            data='2026-07-10',
            cliente=self.other_cliente,
            nota_fiscal='5555, 6666',
            usuario=self.user,
            usuario_nome=self.user.name,
        )
        response = self.api.get(
            '/api/faturamento/protocolos/?notaFiscal=5555',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 1)

    def test_excluir_cliente_com_protocolo_retorna_400(self):
        ProtocoloEnvio.objects.create(
            data='2026-07-10',
            cliente=self.cliente,
            nota_fiscal='1001',
            usuario=self.user,
            usuario_nome=self.user.name,
        )
        response = self.api.delete(
            f'/api/faturamento/protocolo-clientes/{self.cliente.pk}/',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 400)

    def test_numeracao_de_protocolo_e_independente_por_cliente(self):
        resp1 = self.api.post(
            '/api/faturamento/protocolos/',
            {'data': '2026-07-10', 'clienteId': self.other_cliente.pk, 'notaFiscal': '2001'},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        resp2 = self.api.post(
            '/api/faturamento/protocolos/',
            {
                'data': '2026-07-10', 'clienteId': self.cliente.pk, 'notaFiscal': '3001',
                'expedicoes': ['Transcamila Ibiporã'],
            },
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        resp3 = self.api.post(
            '/api/faturamento/protocolos/',
            {'data': '2026-07-10', 'clienteId': self.other_cliente.pk, 'notaFiscal': '2002'},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(resp1.data['protocoloNumero'], '2026-0001')
        self.assertEqual(resp2.data['protocoloNumero'], '2026-0001')
        self.assertEqual(resp3.data['protocoloNumero'], '2026-0002')

    def _criar_protocolo(self, nota: str):
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {'data': '2026-07-10', 'clienteId': self.other_cliente.pk, 'notaFiscal': nota},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 201)
        return response.data

    def test_nota_fiscal_com_letras_e_rejeitada(self):
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {'data': '2026-07-10', 'clienteId': self.other_cliente.pk, 'notaFiscal': '1001, NF-2'},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('apenas números', str(response.data))
        self.assertEqual(ProtocoloEnvio.objects.count(), 0)

    def test_limite_de_78_notas_por_protocolo(self):
        notas_78 = ', '.join(str(10000 + i) for i in range(78))
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {'data': '2026-07-10', 'clienteId': self.other_cliente.pk, 'notaFiscal': notas_78},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 201, response.data)

        notas_79 = ', '.join(str(20000 + i) for i in range(79))
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {'data': '2026-07-10', 'clienteId': self.other_cliente.pk, 'notaFiscal': notas_79},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('78', str(response.data))

    def test_operador_sem_funcao_nao_pode_criar_protocolo(self):
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {'data': '2026-07-10', 'clienteId': self.other_cliente.pk, 'notaFiscal': '3801'},
            format='json',
            **auth_headers(self.user_sem_funcoes, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(ProtocoloEnvio.objects.count(), 0)

    def test_operador_sem_funcao_nao_pode_editar_protocolo(self):
        p1 = self._criar_protocolo('3851')
        response = self.api.patch(
            f"/api/faturamento/protocolos/{p1['id']}/",
            {'notaFiscal': '3852'},
            format='json',
            **auth_headers(self.user_sem_funcoes, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)
        protocolo = ProtocoloEnvio.objects.get(pk=p1['id'])
        self.assertEqual(protocolo.nota_fiscal, '3851')

    def test_operador_sem_funcao_pode_listar_protocolos(self):
        self._criar_protocolo('3881')
        response = self.api.get(
            '/api/faturamento/protocolos/',
            **auth_headers(self.user_sem_funcoes, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 1)

    def test_operador_sem_funcao_nao_pode_excluir_protocolo(self):
        p1 = self._criar_protocolo('3901')
        response = self.api.delete(
            f"/api/faturamento/protocolos/{p1['id']}/",
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)

        response = self.api.post(
            '/api/faturamento/protocolos/bulk_delete/',
            {'ids': [p1['id']]},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)
        self.assertTrue(ProtocoloEnvio.objects.filter(pk=p1['id']).exists())

    def test_operador_com_funcao_gerenciar_clientes_pode_criar_cliente(self):
        operador = User.objects.create_user(
            username='fat.gestor.clientes',
            password='test123',
            name='Operador Gestor',
            role_id='2',
            environments=['Faturamento'],
            filiais={'Faturamento': ['Ibiporã (Matriz)']},
            funcoes={'Faturamento': ['gerenciar-clientes']},
        )
        response = self.api.post(
            '/api/faturamento/protocolo-clientes/',
            {'nome': 'Cliente do Operador', 'requerExpedicao': False},
            format='json',
            **auth_headers(operador, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 201)
        self.assertTrue(ClienteProtocolo.objects.filter(nome='Cliente do Operador').exists())

    def test_excluir_ultimo_protocolo_devolve_numero_para_sequencia(self):
        self._criar_protocolo('4001')
        p2 = self._criar_protocolo('4002')

        response = self.api.delete(
            f"/api/faturamento/protocolos/{p2['id']}/",
            **auth_headers(self.user_excluidor, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 204)

        self.other_cliente.refresh_from_db()
        self.assertEqual(self.other_cliente.ultimo_numero_protocolo, 1)

        p3 = self._criar_protocolo('4003')
        self.assertEqual(p3['protocoloNumero'], '2026-0002')

    def test_excluir_protocolo_do_meio_nao_devolve_numero(self):
        p1 = self._criar_protocolo('5001')
        self._criar_protocolo('5002')

        response = self.api.delete(
            f"/api/faturamento/protocolos/{p1['id']}/",
            **auth_headers(self.user_excluidor, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 204)

        self.other_cliente.refresh_from_db()
        self.assertEqual(self.other_cliente.ultimo_numero_protocolo, 2)

        p3 = self._criar_protocolo('5003')
        self.assertEqual(p3['protocoloNumero'], '2026-0003')

    def test_bulk_delete_devolve_numeros_contiguos_do_topo(self):
        p1 = self._criar_protocolo('6001')
        p2 = self._criar_protocolo('6002')
        p3 = self._criar_protocolo('6003')

        # Exclui o nº 1 (meio) e os nºs 2 e 3 (topo contíguo): contador volta a 1.
        response = self.api.post(
            '/api/faturamento/protocolos/bulk_delete/',
            {'ids': [p2['id'], p3['id']]},
            format='json',
            **auth_headers(self.user_excluidor, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200)

        self.other_cliente.refresh_from_db()
        self.assertEqual(self.other_cliente.ultimo_numero_protocolo, 1)

        response = self.api.delete(
            f"/api/faturamento/protocolos/{p1['id']}/",
            **auth_headers(self.user_excluidor, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 204)
        self.other_cliente.refresh_from_db()
        self.assertEqual(self.other_cliente.ultimo_numero_protocolo, 0)

    def test_nf_duplicada_para_mesmo_cliente_retorna_400(self):
        ProtocoloEnvio.objects.create(
            data='2026-07-10',
            cliente=self.other_cliente,
            nota_fiscal='4001, 4002',
            usuario=self.user,
            usuario_nome=self.user.name,
        )
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {'data': '2026-07-11', 'clienteId': self.other_cliente.pk, 'notaFiscal': '4002, 4003'},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(ProtocoloEnvio.objects.count(), 1)

    def test_nf_duplicada_em_cliente_diferente_e_permitida(self):
        ProtocoloEnvio.objects.create(
            data='2026-07-10',
            cliente=self.other_cliente,
            nota_fiscal='5001',
            usuario=self.user,
            usuario_nome=self.user.name,
        )
        response = self.api.post(
            '/api/faturamento/protocolos/',
            {
                'data': '2026-07-11', 'clienteId': self.cliente.pk, 'notaFiscal': '5001',
                'expedicoes': ['Transcamila Ibiporã'],
            },
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(ProtocoloEnvio.objects.count(), 2)

    def test_editar_protocolo_mantendo_mesma_nf_nao_gera_falso_duplicado(self):
        protocolo = ProtocoloEnvio.objects.create(
            data='2026-07-10',
            cliente=self.other_cliente,
            nota_fiscal='6001',
            usuario=self.user,
            usuario_nome=self.user.name,
        )
        response = self.api.patch(
            f'/api/faturamento/protocolos/{protocolo.pk}/',
            {'notaFiscal': '6001, 6002'},
            format='json',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200)

    def test_acesso_negado_sem_ambiente_faturamento(self):
        user = User.objects.create_user(
            username='fin.user',
            password='test123',
            name='Usuário Financeiro',
            role_id='2',
            environments=['Financeiro'],
            filiais={'Financeiro': ['Ibiporã (Matriz)']},
        )
        response = self.api.get(
            '/api/faturamento/protocolos/',
            **auth_headers(user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 403)

    def test_bulk_print_retorna_pdf(self):
        protocolo = ProtocoloEnvio.objects.create(
            data='2026-07-10',
            cliente=self.other_cliente,
            nota_fiscal='1001, 1002',
            numero_sequencial=1,
            usuario=self.user,
            usuario_nome='Usuário Faturamento',
        )
        response = self.api.get(
            '/api/faturamento/protocolos/bulk_print/',
            {'ids': str(protocolo.pk)},
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200, getattr(response, 'data', response.content))
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_print_pdf_individual_retorna_pdf(self):
        protocolo = ProtocoloEnvio.objects.create(
            data='2026-07-10',
            cliente=self.other_cliente,
            nota_fiscal='2002',
            numero_sequencial=2,
            usuario=self.user,
        )
        response = self.api.get(
            f'/api/faturamento/protocolos/{protocolo.pk}/print_pdf/',
            **auth_headers(self.user, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200, getattr(response, 'data', response.content))
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))


class FaturamentoProtocoloImportTests(TestCase):
    def setUp(self):
        self.operator = User.objects.create_user(
            username='fat.op',
            password='test123',
            name='Operador Faturamento',
            role_id='2',
            environments=['Faturamento'],
            filiais={'Faturamento': ['Ibiporã (Matriz)']},
        )
        self.admin = User.objects.create_user(
            username='fat.admin',
            password='test123',
            name='Admin Sistema',
            role_id='1',
            environments=['Faturamento', 'Administração'],
            filiais={'Faturamento': ['Ibiporã (Matriz)']},
        )
        self.cliente = ClienteProtocolo.objects.create(
            nome='Cliente Import',
            requer_expedicao=False,
        )
        self.api = APIClient()

    def _xlsx_file(self, rows, headers=('Data', 'Nota Fiscal'), name='protocolos.xlsx'):
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(list(headers))
        for row in rows:
            ws.append(list(row))
        buf = BytesIO()
        wb.save(buf)
        return SimpleUploadedFile(
            name,
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_operador_nao_pode_importar(self):
        response = self.api.post(
            '/api/faturamento/protocolos/import_spreadsheet/',
            {
                'file': self._xlsx_file([('10/07/2026', '9001')]),
                'clienteId': str(self.cliente.pk),
            },
            format='multipart',
            **auth_headers(self.operator, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(ProtocoloEnvio.objects.count(), 0)

    def test_admin_importa_planilha(self):
        response = self.api.post(
            '/api/faturamento/protocolos/import_spreadsheet/',
            {
                'file': self._xlsx_file([
                    ('10/07/2026', '9001'),
                    ('11/07/2026', '9002'),
                ]),
                'clienteId': str(self.cliente.pk),
            },
            format='multipart',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 2)
        self.assertEqual(ProtocoloEnvio.objects.filter(cliente=self.cliente).count(), 2)

    def test_dry_run_nao_persiste(self):
        response = self.api.post(
            '/api/faturamento/protocolos/import_spreadsheet/',
            {
                'file': self._xlsx_file([('10/07/2026', '9100')]),
                'clienteId': str(self.cliente.pk),
                'dryRun': 'true',
            },
            format='multipart',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['dryRun'])
        self.assertEqual(response.data['created'], 1)
        self.assertEqual(ProtocoloEnvio.objects.count(), 0)

    def test_importacao_permite_nf_duplicada_do_mesmo_cliente(self):
        """Na importação, NFs já cadastradas para o cliente são importadas mesmo
        assim (com aviso) — o bloqueio de duplicatas vale só para o formulário."""
        ProtocoloEnvio.objects.create(
            data='2026-07-01',
            cliente=self.cliente,
            nota_fiscal='9001',
            numero_sequencial=1,
            usuario_nome='Pré-existente',
        )
        response = self.api.post(
            '/api/faturamento/protocolos/import_spreadsheet/',
            {
                'file': self._xlsx_file([
                    ('10/07/2026', '9001'),
                    ('11/07/2026', '9002'),
                ]),
                'clienteId': str(self.cliente.pk),
            },
            format='multipart',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 2)
        self.assertEqual(response.data['ignored'], 0)
        self.assertEqual(ProtocoloEnvio.objects.filter(cliente=self.cliente).count(), 3)
        self.assertEqual(
            ProtocoloEnvio.objects.filter(cliente=self.cliente, nota_fiscal='9001').count(), 2
        )
        avisos = [w['message'] for w in response.data['warnings']]
        self.assertTrue(any('9001' in msg for msg in avisos))

    def test_importacao_rejeita_nf_com_letras(self):
        response = self.api.post(
            '/api/faturamento/protocolos/import_spreadsheet/',
            {
                'file': self._xlsx_file([('10/07/2026', 'ABC123')]),
                'clienteId': str(self.cliente.pk),
            },
            format='multipart',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(response.data['success'])
        erros = [e['message'] for e in response.data['errors']]
        self.assertTrue(any('apenas números' in msg for msg in erros))
        self.assertEqual(ProtocoloEnvio.objects.count(), 0)

    def test_importa_expedicao_e_filial_opcionais(self):
        self.cliente.requer_expedicao = True
        self.cliente.exige_filial = True
        self.cliente.save(update_fields=['requer_expedicao', 'exige_filial'])
        FilialClienteProtocolo.objects.create(cliente=self.cliente, nome='Matriz SP')
        FilialClienteProtocolo.objects.create(cliente=self.cliente, nome='Filial RJ')

        response = self.api.post(
            '/api/faturamento/protocolos/import_spreadsheet/',
            {
                'file': self._xlsx_file(
                    [
                        ('10/07/2026', '8001', 'Transcamila Ibiporã', 'Matriz SP'),
                        ('11/07/2026', '8002', 'Transcamila Barueri', 'Filial RJ'),
                    ],
                    headers=('Data', 'Nota Fiscal', 'Expedição', 'Filial'),
                ),
                'clienteId': str(self.cliente.pk),
            },
            format='multipart',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 2)

        p1 = ProtocoloEnvio.objects.get(cliente=self.cliente, nota_fiscal='8001')
        self.assertEqual(p1.expedicao, 'Transcamila Ibiporã')
        self.assertEqual(p1.notas_filiais, {'8001': 'Matriz SP'})

        p2 = ProtocoloEnvio.objects.get(cliente=self.cliente, nota_fiscal='8002')
        self.assertEqual(p2.expedicao, 'Transcamila Barueri')
        self.assertEqual(p2.notas_filiais, {'8002': 'Filial RJ'})

    def test_sem_expedicao_filial_ainda_importa_sem_aviso_de_obrigatoriedade(self):
        self.cliente.requer_expedicao = True
        self.cliente.exige_filial = True
        self.cliente.save(update_fields=['requer_expedicao', 'exige_filial'])

        response = self.api.post(
            '/api/faturamento/protocolos/import_spreadsheet/',
            {
                'file': self._xlsx_file([('10/07/2026', '8100')]),
                'clienteId': str(self.cliente.pk),
            },
            format='multipart',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 1)
        # Importação não exige expedição/filial — sem avisos de obrigatoriedade.
        self.assertEqual(response.data['warnings'], [])

        protocolo = ProtocoloEnvio.objects.get(cliente=self.cliente, nota_fiscal='8100')
        self.assertIsNone(protocolo.expedicao)
        self.assertEqual(protocolo.notas_filiais, {})

    def test_importa_formato_oficial_sem_expedicao(self):
        """Formato Ano | Numero Protocolo | Expedição | Data de envio | Cliente | Nota Fiscal."""
        response = self.api.post(
            '/api/faturamento/protocolos/import_spreadsheet/',
            {
                'file': self._xlsx_file(
                    [
                        (2025, 1, None, '04/02/2025', 'Ascenza brasil ltda.', '44'),
                        (2025, 1, None, '04/02/2025', 'Ascenza brasil ltda.', '34'),
                        (2025, 2, '', '10/02/2025', 'Ascenza brasil ltda.', '100'),
                    ],
                    headers=(
                        'Ano',
                        'Numero Protocolo',
                        'Expedição',
                        'Data de envio',
                        'Cliente',
                        'Nota Fiscal',
                    ),
                ),
                'clienteId': str(self.cliente.pk),
            },
            format='multipart',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 2)
        self.assertEqual(response.data['groupingMode'], 'grouped')

        p1 = ProtocoloEnvio.objects.get(cliente=self.cliente, numero_sequencial=1)
        self.assertEqual(p1.nota_fiscal, '44, 34')
        self.assertIsNone(p1.expedicao)
        self.assertEqual(p1.notas_filiais, {})

        p2 = ProtocoloEnvio.objects.get(cliente=self.cliente, numero_sequencial=2)
        self.assertEqual(p2.nota_fiscal, '100')
        self.assertIsNone(p2.expedicao)

    def test_operador_nao_pode_baixar_modelo(self):
        response = self.api.get(
            '/api/faturamento/protocolos/exportar_modelo/',
            **auth_headers(self.operator, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 403)

    def test_admin_baixa_modelo_referencia(self):
        import openpyxl
        from io import BytesIO

        response = self.api.get(
            '/api/faturamento/protocolos/exportar_modelo/',
            **auth_headers(self.admin, 'Faturamento'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertIn(
            'modelo_importacao_protocolos.xlsx',
            response['Content-Disposition'],
        )
        wb = openpyxl.load_workbook(BytesIO(response.content))
        self.assertIn('Importação', wb.sheetnames)
        self.assertIn('Instruções', wb.sheetnames)
        headers = [c.value for c in wb['Importação'][1]]
        self.assertEqual(
            headers,
            [
                'Ano',
                'Numero Protocolo',
                'Expedição',
                'Data de envio',
                'Cliente',
                'Nota Fiscal',
            ],
        )
