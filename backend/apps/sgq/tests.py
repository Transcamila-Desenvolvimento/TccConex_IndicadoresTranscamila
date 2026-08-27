from django.contrib.auth import get_user_model
from django.utils import timezone
from rest_framework.test import APITestCase

from apps.faturamento.models import ClienteProtocolo

from .models import PesquisaSatisfacao

User = get_user_model()

IBIPORA = 'Ibiporã (Matriz)'
RONDONOPOLIS = 'Rondonópolis'


def _headers(filial: str | None = IBIPORA) -> dict:
    headers = {'HTTP_X_PROTHON_ENVIRONMENT': 'SGQ'}
    if filial:
        headers['HTTP_X_PROTHON_FILIAL'] = filial
    return headers


ENV_HEADER = _headers()


def _payload(**overrides):
    data = {
        'motorista': 'João da Silva',
        'cte': '12345',
        'dataEntrega': '2026-07-01',
        'notaFiscal': '98765',
        'cliente': 'CCAB',
        'prazoEntrega': 'otimo',
        'condicoesMercadoria': 'otimo',
        'condicoesVeiculo': 'bom',
        'apresentacaoMotorista': 'otimo',
        'atendimentoDispensado': 'bom',
    }
    data.update(overrides)
    return data


class PesquisaSatisfacaoTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username='sgq.admin',
            password='x',
            name='Admin SGQ',
            role_id='1',
            status='ativo',
            environments=['SGQ'],
            filiais={},
        )
        self.client.force_authenticate(self.admin)

        # Operador com acesso ao SGQ nas duas filiais — usado para validar que o
        # escopo por filial (não o bypass de admin) restringe os dados corretamente.
        self.operador = User.objects.create_user(
            username='sgq.operador',
            password='x',
            name='Operador SGQ',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={'SGQ': ['criar-pesquisas', 'editar-pesquisas', 'excluir-pesquisas']},
        )
        self.operador_consulta = User.objects.create_user(
            username='sgq.consulta',
            password='x',
            name='Operador Consulta SGQ',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={},
        )
        for nome in ('CCAB', 'PRENTISS', 'ALBAUGH', 'Braskem'):
            ClienteProtocolo.objects.create(
                nome=nome,
                razao_social=nome,
                nome_interno=nome,
                considerar_pesquisa_satisfacao=True,
            )

    def _create(self, filial=IBIPORA, **overrides):
        return self.client.post(
            '/api/sgq/pesquisas-satisfacao/', _payload(**overrides), format='json', **_headers(filial)
        )

    def test_crud_pesquisa(self):
        response = self._create()
        self.assertEqual(response.status_code, 201)
        pesquisa_id = response.data['id']
        self.assertEqual(response.data['notaFiscal'], '98765')
        self.assertEqual(response.data['criadoPor'], 'Admin SGQ')
        self.assertEqual(response.data['filial'], IBIPORA)
        self.assertEqual(response.data['dataInclusao'], timezone.localdate().isoformat())
        # Cliente não pode sobrescrever dataInclusao via payload.
        self.assertEqual(
            PesquisaSatisfacao.objects.get(pk=pesquisa_id).data_inclusao,
            timezone.localdate(),
        )

        response = self.client.patch(
            f'/api/sgq/pesquisas-satisfacao/{pesquisa_id}/',
            {'prazoEntrega': 'ruim', 'analise': 'Atraso na entrega.', 'escopoAnalise': {
                'prazo_entrega': ['entregas_fora_prazo_contratual'],
            }},
            format='json',
            **ENV_HEADER,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['prazoEntrega'], 'ruim')
        self.assertEqual(response.data['escopoAnalise'], {
            'prazo_entrega': ['entregas_fora_prazo_contratual'],
        })

        response = self.client.delete(
            f'/api/sgq/pesquisas-satisfacao/{pesquisa_id}/', **ENV_HEADER
        )
        self.assertEqual(response.status_code, 204)
        self.assertEqual(PesquisaSatisfacao.objects.count(), 0)

    def test_analise_exige_escopo(self):
        response = self._create(analise='Atraso na descarga.', cte='ESCOPO-1')
        self.assertEqual(response.status_code, 400)
        self.assertIn('escopoAnalise', response.data)

        response = self._create(
            analise='Atraso na descarga.',
            escopoAnalise={'prazo_entrega': ['entregas_fora_prazo_contratual', 'motorista_recusou_ajudar_descarga']},
            cte='ESCOPO-2',
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(
            response.data['escopoAnalise'],
            {'prazo_entrega': ['entregas_fora_prazo_contratual', 'motorista_recusou_ajudar_descarga']},
        )

        response = self._create(
            analise='',
            escopoAnalise={'prazo_entrega': ['entregas_fora_prazo_contratual']},
            cte='ESCOPO-3',
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['analise'], '')
        self.assertEqual(response.data['escopoAnalise'], {})

    def test_lista_paginada_com_filtros(self):
        self._create(cte='111', cliente='CCAB', dataEntrega='2026-07-01')
        self._create(cte='222', cliente='PRENTISS', dataEntrega='2026-07-10', prazoEntrega='ruim')
        self._create(cte='333', cliente='CCAB', dataEntrega='2026-06-15')

        response = self.client.get('/api/sgq/pesquisas-satisfacao/', **ENV_HEADER)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 3)

        response = self.client.get(
            '/api/sgq/pesquisas-satisfacao/', {'cliente': 'CCAB'}, **ENV_HEADER
        )
        self.assertEqual(response.data['count'], 2)

        response = self.client.get(
            '/api/sgq/pesquisas-satisfacao/', {'search': '222'}, **ENV_HEADER
        )
        self.assertEqual(response.data['count'], 1)

        response = self.client.get(
            '/api/sgq/pesquisas-satisfacao/', {'avaliacao': 'ruim'}, **ENV_HEADER
        )
        self.assertEqual(response.data['count'], 1)

        response = self.client.get(
            '/api/sgq/pesquisas-satisfacao/',
            {'dataInicio': '2026-07-01', 'dataFim': '2026-07-31'},
            **ENV_HEADER,
        )
        self.assertEqual(response.data['count'], 2)

    def test_ordenacao_por_data_entrega(self):
        self._create(cte='OLD', dataEntrega='2026-06-01')
        self._create(cte='NEW', dataEntrega='2026-07-20')

        asc = self.client.get(
            '/api/sgq/pesquisas-satisfacao/',
            {'ordering': 'data_entrega_asc'},
            **ENV_HEADER,
        )
        self.assertEqual(asc.status_code, 200)
        self.assertEqual([r['cte'] for r in asc.data['results']], ['OLD', 'NEW'])

        desc = self.client.get(
            '/api/sgq/pesquisas-satisfacao/',
            {'ordering': 'data_entrega_desc'},
            **ENV_HEADER,
        )
        self.assertEqual([r['cte'] for r in desc.data['results']], ['NEW', 'OLD'])

    def test_filtro_por_motorista(self):
        self._create(cte='111', motorista='João da Silva')
        self._create(cte='222', motorista='Maria Souza')

        response = self.client.get(
            '/api/sgq/pesquisas-satisfacao/', {'motorista': 'joão da silva'}, **ENV_HEADER
        )
        self.assertEqual(response.data['count'], 1)
        self.assertEqual(response.data['results'][0]['cte'], '111')

    def test_stats(self):
        # 1ª pesquisa: 3 ótimo + 2 bom; 2ª: 5 ruim.
        self._create()
        self._create(
            cte='999',
            prazoEntrega='ruim',
            condicoesMercadoria='ruim',
            condicoesVeiculo='ruim',
            apresentacaoMotorista='ruim',
            atendimentoDispensado='ruim',
        )

        response = self.client.get('/api/sgq/pesquisas-satisfacao/stats/', **ENV_HEADER)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['totalPesquisas'], 2)
        self.assertEqual(response.data['totalAvaliacoes'], 10)
        self.assertEqual(response.data['contagem']['otimo'], 3)
        self.assertEqual(response.data['contagem']['ruim'], 5)
        self.assertEqual(response.data['percentual']['otimo'], 30.0)
        self.assertEqual(response.data['pontosAtencao'], 5)
        prazo = next(c for c in response.data['criterios'] if c['campo'] == 'prazo_entrega')
        self.assertEqual(prazo['otimo'], 1)
        self.assertEqual(prazo['ruim'], 1)
        self.assertEqual(prazo['score'], 2.5)

    def test_avaliacao_obrigatoria(self):
        response = self._create(prazoEntrega='')
        self.assertEqual(response.status_code, 400)

    def test_bulk_create_sucesso(self):
        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/bulk_create/',
            [_payload(cte='111'), _payload(cte='222', cliente='PRENTISS')],
            format='json',
            **ENV_HEADER,
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(len(response.data), 2)
        self.assertEqual(PesquisaSatisfacao.objects.count(), 2)
        self.assertTrue(all(item['filial'] == IBIPORA for item in response.data))

    def test_bulk_create_tudo_ou_nada(self):
        # Segunda linha inválida (sem prazoEntrega) — nenhuma das duas deve ser salva.
        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/bulk_create/',
            [_payload(cte='111'), _payload(cte='222', prazoEntrega='')],
            format='json',
            **ENV_HEADER,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('1', {str(k) for k in response.data['errors'].keys()})
        self.assertEqual(PesquisaSatisfacao.objects.count(), 0)

    def test_bulk_create_lista_vazia(self):
        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/bulk_create/', [], format='json', **ENV_HEADER
        )
        self.assertEqual(response.status_code, 400)

    def test_acesso_negado_sem_ambiente_sgq(self):
        operador = User.objects.create_user(
            username='sem.sgq',
            password='x',
            name='Sem SGQ',
            role_id='2',
            status='ativo',
            environments=['Financeiro'],
            filiais={},
        )
        self.client.force_authenticate(operador)
        response = self.client.get('/api/sgq/pesquisas-satisfacao/', **ENV_HEADER)
        self.assertEqual(response.status_code, 403)

    def test_acesso_negado_sem_filial_na_sessao(self):
        """SGQ não é mais global — exige filial na sessão, como o Indicadores."""
        response = self.client.get('/api/sgq/pesquisas-satisfacao/', **_headers(filial=None))
        self.assertEqual(response.status_code, 403)

    def test_acesso_negado_filial_fora_do_sgq(self):
        """Paranaguá existe para outros módulos, mas não é uma filial válida do SGQ."""
        response = self.client.get('/api/sgq/pesquisas-satisfacao/', **_headers(filial='Paranaguá'))
        self.assertEqual(response.status_code, 403)

    def test_filial_percent_encoded_no_header_e_aceita(self):
        """Azure/proxies podem exigir ASCII no header — frontend envia encodeURIComponent."""
        from urllib.parse import quote

        encoded = quote(IBIPORA, safe='')
        response = self.client.get(
            '/api/sgq/pesquisas-satisfacao/',
            HTTP_X_PROTHON_ENVIRONMENT='SGQ',
            HTTP_X_PROTHON_FILIAL=encoded,
        )
        self.assertEqual(response.status_code, 200)

    def test_pesquisas_isoladas_por_filial(self):
        """Pesquisa lançada em uma filial não aparece para outra filial — visão de operador."""
        self.client.force_authenticate(self.operador)
        self._create(filial=IBIPORA, cte='IBI-1')
        self._create(filial=RONDONOPOLIS, cte='RDN-1')

        resp_ibipora = self.client.get('/api/sgq/pesquisas-satisfacao/', **_headers(IBIPORA))
        self.assertEqual(resp_ibipora.data['count'], 1)
        self.assertEqual(resp_ibipora.data['results'][0]['cte'], 'IBI-1')

        resp_rondonopolis = self.client.get('/api/sgq/pesquisas-satisfacao/', **_headers(RONDONOPOLIS))
        self.assertEqual(resp_rondonopolis.data['count'], 1)
        self.assertEqual(resp_rondonopolis.data['results'][0]['cte'], 'RDN-1')

    def test_pesquisas_isoladas_por_filial_mesmo_para_admin(self):
        """Diferente de Financeiro/Indicadores, o SGQ não tem visão consolidada:
        nem o admin vê o consolidado — cada filial fica sempre segregada."""
        self._create(filial=IBIPORA, cte='IBI-ADM')
        self._create(filial=RONDONOPOLIS, cte='RDN-ADM')

        resp_ibipora = self.client.get('/api/sgq/pesquisas-satisfacao/', **_headers(IBIPORA))
        self.assertEqual(resp_ibipora.data['count'], 1)
        self.assertEqual(resp_ibipora.data['results'][0]['cte'], 'IBI-ADM')

        resp_rondonopolis = self.client.get('/api/sgq/pesquisas-satisfacao/', **_headers(RONDONOPOLIS))
        self.assertEqual(resp_rondonopolis.data['count'], 1)
        self.assertEqual(resp_rondonopolis.data['results'][0]['cte'], 'RDN-ADM')

    def test_stats_respeita_filial_da_sessao(self):
        self.client.force_authenticate(self.operador)
        self._create(filial=IBIPORA, prazoEntrega='otimo')
        self._create(
            filial=RONDONOPOLIS,
            cte='999',
            prazoEntrega='ruim',
            condicoesMercadoria='ruim',
            condicoesVeiculo='ruim',
            apresentacaoMotorista='ruim',
            atendimentoDispensado='ruim',
        )

        response = self.client.get('/api/sgq/pesquisas-satisfacao/stats/', **_headers(IBIPORA))
        self.assertEqual(response.data['totalPesquisas'], 1)
        self.assertEqual(response.data['contagem']['ruim'], 0)

    def test_operador_sem_acesso_a_filial_nao_configurada(self):
        """Operador com SGQ liberado só em Ibiporã não pode consultar com sessão Rondonópolis."""
        operador_ibipora = User.objects.create_user(
            username='sgq.ibipora',
            password='x',
            name='Operador Ibiporã',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA]},
        )
        self.client.force_authenticate(operador_ibipora)
        response = self.client.get('/api/sgq/pesquisas-satisfacao/', **_headers(RONDONOPOLIS))
        self.assertEqual(response.status_code, 403)

    def test_cliente_recusou_assinar_dispensa_avaliacoes(self):
        """Com a recusa marcada, os critérios de avaliação podem ficar vazios."""
        response = self._create(
            clienteRecusouAssinar=True,
            prazoEntrega='',
            condicoesMercadoria='',
            condicoesVeiculo='',
            apresentacaoMotorista='',
            atendimentoDispensado='',
        )
        self.assertEqual(response.status_code, 201)
        self.assertTrue(response.data['clienteRecusouAssinar'])
        self.assertEqual(response.data['prazoEntrega'], '')

    def test_avaliacoes_obrigatorias_quando_nao_recusou(self):
        """Sem a recusa marcada, os critérios continuam obrigatórios."""
        response = self._create(clienteRecusouAssinar=False, prazoEntrega='')
        self.assertEqual(response.status_code, 400)
        self.assertIn('prazoEntrega', response.data)

    def test_sugestoes_motoristas_deduplica_variacoes_de_escrita(self):
        """Nomes digitados de formas diferentes (caixa/espaços) contam como o mesmo
        motorista — a sugestão retorna a grafia mais usada, sem duplicar."""
        self._create(motorista='João da Silva', cte='1')
        self._create(motorista='João da Silva', cte='2')
        self._create(motorista='João da Silva', cte='3')
        self._create(motorista='joão da silva ', cte='4')
        self._create(motorista='Maria Souza', cte='5')

        response = self.client.get('/api/sgq/pesquisas-satisfacao/motoristas/', **ENV_HEADER)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)
        self.assertIn('João da Silva', response.data)
        self.assertIn('Maria Souza', response.data)

    def test_sugestoes_motoristas_compartilhadas_entre_filiais(self):
        self.client.force_authenticate(self.operador)
        self._create(filial=IBIPORA, motorista='Motorista Ibiporã')
        self._create(filial=RONDONOPOLIS, motorista='Motorista Rondonópolis')

        response = self.client.get('/api/sgq/pesquisas-satisfacao/motoristas/', **_headers(IBIPORA))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, ['Motorista Ibiporã', 'Motorista Rondonópolis'])

        response_rondonopolis = self.client.get(
            '/api/sgq/pesquisas-satisfacao/motoristas/', **_headers(RONDONOPOLIS)
        )
        self.assertEqual(response_rondonopolis.data, ['Motorista Ibiporã', 'Motorista Rondonópolis'])

    def test_filial_gravada_pela_sessao_e_nao_pelo_payload(self):
        """O cliente não pode escolher a filial via payload — só a sessão define."""
        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/',
            _payload(filial=RONDONOPOLIS),
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['filial'], IBIPORA)

    def test_operador_sem_funcao_pode_listar(self):
        self._create(cte='CONSULTA-1')
        self.client.force_authenticate(self.operador_consulta)
        response = self.client.get('/api/sgq/pesquisas-satisfacao/', **ENV_HEADER)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 1)

    def test_operador_sem_funcao_nao_pode_criar(self):
        self.client.force_authenticate(self.operador_consulta)
        response = self._create(cte='BLOQ-1')
        self.assertEqual(response.status_code, 403)
        self.assertEqual(PesquisaSatisfacao.objects.count(), 0)

    def test_operador_sem_funcao_nao_pode_editar(self):
        created = self._create(cte='BLOQ-ED')
        pesquisa_id = created.data['id']
        self.client.force_authenticate(self.operador_consulta)
        response = self.client.patch(
            f'/api/sgq/pesquisas-satisfacao/{pesquisa_id}/',
            {'analise': 'Tentativa sem permissão'},
            format='json',
            **ENV_HEADER,
        )
        self.assertEqual(response.status_code, 403)

    def test_operador_sem_funcao_nao_pode_excluir(self):
        created = self._create(cte='BLOQ-EX')
        pesquisa_id = created.data['id']
        self.client.force_authenticate(self.operador_consulta)
        response = self.client.delete(
            f'/api/sgq/pesquisas-satisfacao/{pesquisa_id}/', **ENV_HEADER
        )
        self.assertEqual(response.status_code, 403)
        self.assertTrue(PesquisaSatisfacao.objects.filter(pk=pesquisa_id).exists())

    def test_operador_sem_funcao_nao_pode_bulk_create(self):
        self.client.force_authenticate(self.operador_consulta)
        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/bulk_create/',
            [_payload(cte='BLOQ-Lote')],
            format='json',
            **ENV_HEADER,
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(PesquisaSatisfacao.objects.count(), 0)

    def test_operador_com_funcao_criar_pode_criar(self):
        self.client.force_authenticate(self.operador)
        response = self._create(cte='OK-CRIAR')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['criadoPor'], 'Operador SGQ')

    def test_enviar_resumo_consolida_filiais_sgq(self):
        from datetime import date

        from django.core import mail

        self._create(filial=IBIPORA, cte='111')
        self._create(
            filial=IBIPORA,
            cte='BLANK',
            clienteRecusouAssinar=True,
            prazoEntrega='',
            condicoesMercadoria='',
            condicoesVeiculo='',
            apresentacaoMotorista='',
            atendimentoDispensado='',
        )
        self._create(filial=RONDONOPOLIS, cte='222')
        PesquisaSatisfacao.objects.filter(filial=IBIPORA).update(data_inclusao=date(2026, 7, 10))
        PesquisaSatisfacao.objects.filter(filial=RONDONOPOLIS).update(data_inclusao=date(2026, 8, 5))

        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/enviar-resumo/',
            {'to': ['destino@example.com']},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(mail.outbox), 1)

        message = mail.outbox[0]
        self.assertIn('Pesquisa de Satisfação — Ibiporã e Rondonópolis', message.subject)
        self.assertIn('Admin SGQ', message.body)
        self.assertIn('indicadores/gestao-qualidade/satisfacao-clientes', message.body)
        self.assertIn('Por Filial', message.body)
        self.assertIn('Última inclusão', message.body)
        self.assertIn('10/07/2026', message.body)
        self.assertIn('05/08/2026', message.body)
        self.assertIn('Em branco', message.body)
        self.assertIn('>3<', message.body)
        self.assertIn('>1<', message.body)
        self.assertEqual(message.to, ['destino@example.com'])
        self.assertEqual(len(message.attachments), 0)

    def test_enviar_resumo_consolida_mesmo_sem_acesso_as_duas_filiais(self):
        """Operador só com Ibiporã no SGQ ainda envia o consolidado das duas unidades."""
        from django.core import mail

        self._create(filial=IBIPORA, cte='111')
        self._create(filial=RONDONOPOLIS, cte='222')

        so_ibipora = User.objects.create_user(
            username='sgq.resumo.ibipora',
            password='x',
            name='Operador Só Ibiporã',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA]},
        )
        self.client.force_authenticate(so_ibipora)
        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/enviar-resumo/',
            {'to': ['destino@example.com']},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 200, response.data)
        message = mail.outbox[0]
        self.assertIn('Pesquisa de Satisfação — Ibiporã e Rondonópolis', message.subject)
        self.assertIn('>2<', message.body)

    def test_enviar_resumo_ignora_filtros_da_tabela(self):
        from django.core import mail

        self._create(filial=IBIPORA, cte='111', cliente='CCAB')
        self._create(filial=IBIPORA, cte='222', cliente='Braskem')

        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/enviar-resumo/',
            {'to': ['destino@example.com'], 'cliente': 'CCAB'},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 200, response.data)

        message = mail.outbox[0]
        self.assertIn('>2<', message.body)

    def test_enviar_resumo_exige_destinatario(self):
        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/enviar-resumo/',
            {'to': []},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 400)

    def test_clientes_endpoint_lista_apenas_habilitados(self):
        ClienteProtocolo.objects.filter(nome__iexact='CCAB').update(considerar_pesquisa_satisfacao=False)
        response = self.client.get('/api/sgq/pesquisas-satisfacao/clientes/', **ENV_HEADER)
        self.assertEqual(response.status_code, 200)
        valores = {item['value'] for item in response.data}
        rotulos = {item['label'] for item in response.data}
        self.assertNotIn('OUTROS', valores)
        self.assertTrue(any('prentiss' in texto.casefold() for texto in valores | rotulos))
        self.assertFalse(any('ccab' in texto.casefold() for texto in valores | rotulos))

    def test_nao_cria_pesquisa_manual_com_outros(self):
        response = self._create(cliente='OUTROS')
        self.assertEqual(response.status_code, 400)

    def test_nao_cria_pesquisa_para_cliente_desabilitado(self):
        ClienteProtocolo.objects.filter(nome__iexact='CCAB').update(considerar_pesquisa_satisfacao=False)
        response = self._create(cliente='CCAB')
        self.assertEqual(response.status_code, 400)

    def test_edicao_mantem_cliente_legado_desabilitado(self):
        created = self._create(cte='legado-1', cliente='CCAB')
        self.assertEqual(created.status_code, 201, created.data)
        ClienteProtocolo.objects.filter(nome__iexact='CCAB').update(considerar_pesquisa_satisfacao=False)
        pesquisa_id = created.data['id']
        response = self.client.patch(
            f'/api/sgq/pesquisas-satisfacao/{pesquisa_id}/',
            {'motorista': 'João Atualizado'},
            format='json',
            **ENV_HEADER,
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['cliente'], 'CCAB')
        self.assertEqual(response.data['motorista'], 'João Atualizado')


class PesquisaSatisfacaoLoteDraftTests(APITestCase):
    def setUp(self):
        self.user_a = User.objects.create_user(
            username='sgq.draft.a',
            password='x',
            name='Operador A',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={'SGQ': ['criar-pesquisas']},
        )
        self.user_b = User.objects.create_user(
            username='sgq.draft.b',
            password='x',
            name='Operador B',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={'SGQ': ['criar-pesquisas']},
        )
        self.user_consulta = User.objects.create_user(
            username='sgq.draft.consulta',
            password='x',
            name='Operador Consulta Draft',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={},
        )

    def test_put_get_isolado_por_usuario(self):
        self.client.force_authenticate(self.user_a)
        response = self.client.put(
            '/api/sgq/pesquisas-satisfacao/lote-draft/',
            {'rows': [{'motorista': 'Motorista A', 'cliente': 'CCAB', 'cte': '1', 'notaFiscal': '',
                       'dataEntrega': '', 'clienteRecusouAssinar': False, 'prazoEntrega': '',
                       'condicoesMercadoria': '', 'condicoesVeiculo': '', 'apresentacaoMotorista': '',
                       'atendimentoDispensado': '', 'analise': ''}]},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['hasDraft'])
        self.assertEqual(response.data['rows'][0]['motorista'], 'Motorista A')

        self.client.force_authenticate(self.user_b)
        response_b = self.client.get('/api/sgq/pesquisas-satisfacao/lote-draft/', **_headers(IBIPORA))
        self.assertEqual(response_b.status_code, 200)
        self.assertFalse(response_b.data['hasDraft'])
        self.assertEqual(response_b.data['rows'], [])

    def test_draft_isolado_por_filial(self):
        self.client.force_authenticate(self.user_a)
        self.client.put(
            '/api/sgq/pesquisas-satisfacao/lote-draft/',
            {'rows': [{'motorista': 'Só Ibiporã', 'cliente': 'OUTROS', 'cte': '9', 'notaFiscal': '',
                       'dataEntrega': '', 'clienteRecusouAssinar': False, 'prazoEntrega': '',
                       'condicoesMercadoria': '', 'condicoesVeiculo': '', 'apresentacaoMotorista': '',
                       'atendimentoDispensado': '', 'analise': ''}]},
            format='json',
            **_headers(IBIPORA),
        )
        response = self.client.get('/api/sgq/pesquisas-satisfacao/lote-draft/', **_headers(RONDONOPOLIS))
        self.assertFalse(response.data['hasDraft'])

    def test_delete_descarta_rascunho(self):
        self.client.force_authenticate(self.user_a)
        self.client.put(
            '/api/sgq/pesquisas-satisfacao/lote-draft/',
            {'rows': [{'motorista': 'X', 'cliente': 'OUTROS', 'cte': '', 'notaFiscal': '',
                       'dataEntrega': '', 'clienteRecusouAssinar': False, 'prazoEntrega': '',
                       'condicoesMercadoria': '', 'condicoesVeiculo': '', 'apresentacaoMotorista': '',
                       'atendimentoDispensado': '', 'analise': ''}]},
            format='json',
            **_headers(IBIPORA),
        )
        deleted = self.client.delete('/api/sgq/pesquisas-satisfacao/lote-draft/', **_headers(IBIPORA))
        self.assertEqual(deleted.status_code, 204)
        response = self.client.get('/api/sgq/pesquisas-satisfacao/lote-draft/', **_headers(IBIPORA))
        self.assertFalse(response.data['hasDraft'])

    def test_operador_sem_funcao_criar_nao_pode_salvar_rascunho(self):
        self.client.force_authenticate(self.user_consulta)
        response = self.client.put(
            '/api/sgq/pesquisas-satisfacao/lote-draft/',
            {'rows': [{'motorista': 'Bloqueado', 'cliente': 'OUTROS', 'cte': '', 'notaFiscal': '',
                       'dataEntrega': '', 'clienteRecusouAssinar': False, 'prazoEntrega': '',
                       'condicoesMercadoria': '', 'condicoesVeiculo': '', 'apresentacaoMotorista': '',
                       'atendimentoDispensado': '', 'analise': ''}]},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 403)


class PesquisaSatisfacaoFormDraftTests(APITestCase):
    def setUp(self):
        self.user_a = User.objects.create_user(
            username='sgq.form.draft.a',
            password='x',
            name='Operador Form A',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={'SGQ': ['criar-pesquisas']},
        )
        self.user_b = User.objects.create_user(
            username='sgq.form.draft.b',
            password='x',
            name='Operador Form B',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={'SGQ': ['criar-pesquisas']},
        )
        self.user_consulta = User.objects.create_user(
            username='sgq.form.draft.consulta',
            password='x',
            name='Operador Consulta Form Draft',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={},
        )

    def test_put_get_isolado_por_usuario(self):
        self.client.force_authenticate(self.user_a)
        response = self.client.put(
            '/api/sgq/pesquisas-satisfacao/form-draft/',
            {'form': {'motorista': 'Motorista A', 'cliente': 'CCAB', 'cte': '1', 'notaFiscal': '',
                      'dataEntrega': '2026-08-26', 'clienteRecusouAssinar': False, 'prazoEntrega': '',
                      'condicoesMercadoria': '', 'condicoesVeiculo': '', 'apresentacaoMotorista': '',
                      'atendimentoDispensado': '', 'analise': ''}},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['hasDraft'])
        self.assertEqual(response.data['form']['motorista'], 'Motorista A')

        self.client.force_authenticate(self.user_b)
        response_b = self.client.get('/api/sgq/pesquisas-satisfacao/form-draft/', **_headers(IBIPORA))
        self.assertEqual(response_b.status_code, 200)
        self.assertFalse(response_b.data['hasDraft'])

    def test_somente_data_entrega_nao_persiste(self):
        self.client.force_authenticate(self.user_a)
        response = self.client.put(
            '/api/sgq/pesquisas-satisfacao/form-draft/',
            {'form': {'dataEntrega': '2026-08-26'}},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data['hasDraft'])

    def test_draft_isolado_por_filial(self):
        self.client.force_authenticate(self.user_a)
        self.client.put(
            '/api/sgq/pesquisas-satisfacao/form-draft/',
            {'form': {'motorista': 'Só Ibiporã', 'cliente': '', 'cte': '', 'notaFiscal': '',
                      'dataEntrega': '', 'clienteRecusouAssinar': False}},
            format='json',
            **_headers(IBIPORA),
        )
        response = self.client.get('/api/sgq/pesquisas-satisfacao/form-draft/', **_headers(RONDONOPOLIS))
        self.assertFalse(response.data['hasDraft'])

    def test_delete_descarta_rascunho(self):
        self.client.force_authenticate(self.user_a)
        self.client.put(
            '/api/sgq/pesquisas-satisfacao/form-draft/',
            {'form': {'motorista': 'X'}},
            format='json',
            **_headers(IBIPORA),
        )
        deleted = self.client.delete('/api/sgq/pesquisas-satisfacao/form-draft/', **_headers(IBIPORA))
        self.assertEqual(deleted.status_code, 204)
        response = self.client.get('/api/sgq/pesquisas-satisfacao/form-draft/', **_headers(IBIPORA))
        self.assertFalse(response.data['hasDraft'])

    def test_operador_sem_funcao_criar_nao_pode_salvar_rascunho(self):
        self.client.force_authenticate(self.user_consulta)
        response = self.client.put(
            '/api/sgq/pesquisas-satisfacao/form-draft/',
            {'form': {'motorista': 'Bloqueado'}},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 403)


class PesquisaImportacaoTests(APITestCase):
    def setUp(self):
        from .pesquisa_import_service import build_pesquisa_import_template

        self.template_bytes = build_pesquisa_import_template()
        self.importador = User.objects.create_user(
            username='sgq.importador',
            password='x',
            name='Importador SGQ',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA]},
            funcoes={'SGQ': ['importar-pesquisas']},
        )
        self.sem_importacao = User.objects.create_user(
            username='sgq.sem.import',
            password='x',
            name='Sem Import',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA]},
            funcoes={'SGQ': ['criar-pesquisas']},
        )

    def _upload(self, user, data: bytes, dry_run=False, extra=None):
        from django.core.files.uploadedfile import SimpleUploadedFile

        self.client.force_authenticate(user)
        file = SimpleUploadedFile(
            'pesquisas.xlsx',
            data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        payload = {'file': file}
        if dry_run:
            payload['dryRun'] = 'true'
        if extra:
            payload.update(extra)
        return self.client.post(
            '/api/sgq/pesquisas-satisfacao/import-spreadsheet/',
            payload,
            format='multipart',
            **_headers(IBIPORA),
        )

    def test_operador_sem_funcao_importar_recebe_403(self):
        response = self._upload(self.sem_importacao, self.template_bytes)
        self.assertEqual(response.status_code, 403)

    def test_importador_grava_com_criado_por_importacao(self):
        before = PesquisaSatisfacao.objects.filter(filial=IBIPORA).count()
        response = self._upload(self.importador, self.template_bytes)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 1)
        self.assertEqual(PesquisaSatisfacao.objects.filter(filial=IBIPORA).count(), before + 1)
        created = PesquisaSatisfacao.objects.order_by('-id').first()
        self.assertEqual(created.criado_por, 'Importação')

    def test_admin_importa_vinculando_lancado_por(self):
        admin = User.objects.create_user(
            username='sgq.import.admin',
            password='x',
            name='Admin Importador',
            role_id='1',
            status='ativo',
            environments=['SGQ'],
        )
        alvo = User.objects.create_user(
            username='miguel.ribeiro',
            password='x',
            name='Miguel Ribeiro',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA]},
        )
        response = self._upload(admin, self.template_bytes, extra={'criadoPorUserId': str(alvo.pk)})
        self.assertEqual(response.status_code, 200, response.data)
        created = PesquisaSatisfacao.objects.order_by('-id').first()
        self.assertEqual(created.criado_por, 'Miguel Ribeiro')
        self.assertEqual(response.data['criadoPor'], 'Miguel Ribeiro')

    def test_admin_sem_selecao_usa_proprio_nome(self):
        admin = User.objects.create_user(
            username='sgq.import.admin2',
            password='x',
            name='Admin Importador',
            role_id='1',
            status='ativo',
            environments=['SGQ'],
        )
        response = self._upload(admin, self.template_bytes)
        self.assertEqual(response.status_code, 200, response.data)
        created = PesquisaSatisfacao.objects.order_by('-id').first()
        self.assertEqual(created.criado_por, 'Admin Importador')

    def test_operador_ignora_criado_por_user_id(self):
        response = self._upload(
            self.importador,
            self.template_bytes,
            extra={'criadoPorUserId': str(self.importador.pk)},
        )
        self.assertEqual(response.status_code, 200, response.data)
        created = PesquisaSatisfacao.objects.order_by('-id').first()
        self.assertEqual(created.criado_por, 'Importação')

    def test_exportar_modelo_exige_funcao_importar(self):
        self.client.force_authenticate(self.sem_importacao)
        denied = self.client.get('/api/sgq/pesquisas-satisfacao/exportar-modelo/', **_headers(IBIPORA))
        self.assertEqual(denied.status_code, 403)

        self.client.force_authenticate(self.importador)
        ok = self.client.get('/api/sgq/pesquisas-satisfacao/exportar-modelo/', **_headers(IBIPORA))
        self.assertEqual(ok.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ok['Content-Type'],
        )

    def test_importa_planilha_com_linhas_vazias_antes_do_cabecalho(self):
        import openpyxl
        from io import BytesIO

        from .pesquisa_import_service import _TEMPLATE_HEADERS, _TEMPLATE_SAMPLE_ROW

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([])
        ws.append(['Relatório de pesquisas'])
        ws.append([])
        ws.append(_TEMPLATE_HEADERS)
        ws.append(_TEMPLATE_SAMPLE_ROW)
        buffer = BytesIO()
        wb.save(buffer)

        response = self._upload(self.importador, buffer.getvalue())
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['created'], 1)

    def test_preview_retorna_linhas_antes_de_importar(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        self.client.force_authenticate(self.importador)
        file = SimpleUploadedFile(
            'pesquisas.xlsx',
            self.template_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response = self.client.post(
            '/api/sgq/pesquisas-satisfacao/import-preview/',
            {'file': file},
            format='multipart',
            **_headers(IBIPORA),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['validRows'], 1)
        self.assertEqual(len(response.data['rows']), 1)
        self.assertEqual(response.data['rows'][0]['motorista'], 'EVALDO JOSE')
        stats = response.data['stats']
        self.assertEqual(stats['validRows'], 1)
        self.assertEqual(stats['processedRows'], 1)
        self.assertEqual(stats['validRate'], 100.0)
        self.assertTrue(stats['readyToImport'])
        self.assertEqual(stats['uniqueMotoristas'], 1)
        self.assertEqual(len(stats['byCliente']), 1)

    def test_importa_pesquisa_em_branco_como_recusa_de_avaliacao(self):
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile

        from .pesquisa_import_service import _TEMPLATE_HEADERS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(_TEMPLATE_HEADERS)
        ws.append([
            '10/08/2026',
            'JOAO SILVA',
            '99001',
            '03/08/2026',
            '7777',
            '',
            '',
            '',
            '',
            '',
            'OUTROS',
            'Cliente não assinou a pesquisa',
            'Prazo de Entrega: Entregas fora do prazo contratual',
        ])
        buffer = BytesIO()
        wb.save(buffer)
        data = buffer.getvalue()

        self.client.force_authenticate(self.importador)
        preview_file = SimpleUploadedFile(
            'pesquisas.xlsx',
            data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        preview = self.client.post(
            '/api/sgq/pesquisas-satisfacao/import-preview/',
            {'file': preview_file},
            format='multipart',
            **_headers(IBIPORA),
        )
        self.assertEqual(preview.status_code, 200, preview.data)
        self.assertTrue(preview.data['success'])
        self.assertTrue(preview.data['rows'][0]['clienteRecusouAssinar'])
        self.assertEqual(preview.data['stats']['rowsClienteRecusou'], 1)

        response = self._upload(self.importador, data)
        self.assertEqual(response.status_code, 200, response.data)
        created = PesquisaSatisfacao.objects.order_by('-id').first()
        self.assertTrue(created.cliente_recusou_assinar)
        self.assertEqual(created.prazo_entrega, '')
        self.assertEqual(created.analise, 'Cliente não assinou a pesquisa')
        self.assertEqual(created.escopo_analise, {'prazo_entrega': ['entregas_fora_prazo_contratual']})
        self.assertEqual(created.motorista, 'JOAO SILVA')
        self.assertEqual(created.cte, '99001')

    def test_importa_analise_sem_escopo(self):
        import openpyxl
        from io import BytesIO

        from .pesquisa_import_service import _TEMPLATE_HEADERS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(_TEMPLATE_HEADERS)
        ws.append([
            '10/08/2026',
            'MARIA SOUZA',
            '99002',
            '03/08/2026',
            '8888',
            'BOM',
            'OTIMO',
            'OTIMO',
            'OTIMO',
            'OTIMO',
            'OUTROS',
            'Atraso pontual no descarregamento.',
            '',
        ])
        buffer = BytesIO()
        wb.save(buffer)

        response = self._upload(self.importador, buffer.getvalue())
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data['success'])
        created = PesquisaSatisfacao.objects.order_by('-id').first()
        self.assertEqual(created.analise, 'Atraso pontual no descarregamento.')
        self.assertEqual(created.escopo_analise, {})


class EscopoAnaliseCadastroTests(APITestCase):
    def setUp(self):
        self.operador = User.objects.create_user(
            username='sgq.escopo.op',
            password='x',
            name='Operador Escopo',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={'SGQ': ['gerenciar-escopos']},
        )
        self.consulta = User.objects.create_user(
            username='sgq.escopo.consulta',
            password='x',
            name='Consulta Escopo',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA, RONDONOPOLIS]},
            funcoes={},
        )

    def test_catalogo_igual_nas_duas_filiais(self):
        self.client.force_authenticate(self.operador)
        ibi = self.client.get('/api/sgq/escopos-analise/', **_headers(IBIPORA))
        ron = self.client.get('/api/sgq/escopos-analise/', **_headers(RONDONOPOLIS))
        self.assertEqual(ibi.status_code, 200, ibi.data)
        self.assertEqual(ron.status_code, 200, ron.data)
        self.assertGreaterEqual(len(ibi.data), 5)
        self.assertEqual(ibi.data, ron.data)

    def test_opcao_criada_em_ibipora_aparece_em_rondonopolis(self):
        self.client.force_authenticate(self.operador)
        catalogo = self.client.get('/api/sgq/escopos-analise/', **_headers(IBIPORA))
        escopo_id = catalogo.data[0]['id']
        created = self.client.post(
            '/api/sgq/escopos-analise-opcoes/',
            {'escopoId': escopo_id, 'label': 'Atraso no pátio do cliente'},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(created.status_code, 201, created.data)

        ron = self.client.get('/api/sgq/escopos-analise/', **_headers(RONDONOPOLIS))
        labels = [opcao['label'] for grupo in ron.data for opcao in grupo['opcoes']]
        self.assertIn('Atraso no pátio do cliente', labels)

    def test_consulta_pode_ler_mas_nao_alterar(self):
        self.client.force_authenticate(self.consulta)
        listed = self.client.get('/api/sgq/escopos-analise/', **_headers(IBIPORA))
        self.assertEqual(listed.status_code, 200)
        denied = self.client.post(
            '/api/sgq/escopos-analise/',
            {'label': 'Novo grupo'},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(denied.status_code, 403)

    def test_criar_pesquisas_nao_libera_cadastro_de_escopos(self):
        so_criar = User.objects.create_user(
            username='sgq.escopo.criar',
            password='x',
            name='Só criar pesquisas',
            role_id='2',
            status='ativo',
            environments=['SGQ'],
            filiais={'SGQ': [IBIPORA]},
            funcoes={'SGQ': ['criar-pesquisas']},
        )
        self.client.force_authenticate(so_criar)
        denied = self.client.post(
            '/api/sgq/escopos-analise/',
            {'label': 'Novo grupo'},
            format='json',
            **_headers(IBIPORA),
        )
        self.assertEqual(denied.status_code, 403)

    def test_nao_exclui_opcao_ja_usada_em_pesquisa(self):
        from .models import EscopoAnaliseOpcao

        opcao = EscopoAnaliseOpcao.objects.get(escopo__chave='condicoes_mercadoria', chave='pallets_tombaram')
        PesquisaSatisfacao.objects.create(
            filial=IBIPORA,
            motorista='X',
            cte='ESC-USO',
            data_entrega=timezone.localdate(),
            nota_fiscal='1',
            cliente='OUTROS',
            analise='Pallets tombados.',
            escopo_analise={'condicoes_mercadoria': ['pallets_tombaram']},
        )
        self.client.force_authenticate(self.operador)
        denied = self.client.delete(
            f'/api/sgq/escopos-analise-opcoes/{opcao.pk}/',
            **_headers(IBIPORA),
        )
        self.assertEqual(denied.status_code, 400, denied.data)
        self.assertTrue(EscopoAnaliseOpcao.objects.filter(pk=opcao.pk).exists())
