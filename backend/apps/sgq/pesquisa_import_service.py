"""Importação de pesquisas de satisfação a partir de planilha .xlsx."""

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from io import BytesIO

import openpyxl
from django.db import transaction
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill

from .models import CLIENTE_CHOICES, PesquisaSatisfacao
from .serializers import PesquisaSatisfacaoSerializer

CRIADO_POR_IMPORTACAO = 'Importação'
MAX_IMPORT_ROWS = 500

# Layout oficial da planilha de importação (mesma ordem usada pela operação).
_TEMPLATE_HEADERS = [
    'DATA DE ENVIO',
    'MOTORISTA',
    'CHTO',
    'DATA ENTREGA',
    'Nºs NOTAS FISCAL',
    'PRAZO DE ENTREGA',
    'CONDIÇÕES DA MERCADORIA',
    'CONDIÇÕES DO VEÍCULO',
    'APRESENTAÇÃO DO MOTORISTA',
    'ATENDIMENTO DISPENSADO',
    'CLIENTE',
    'ANÁLISE, TRATATIVA E JUSTIFICATIVA',
]

_TEMPLATE_SAMPLE_ROW = [
    '10/08/2026',
    'EVALDO JOSE',
    '55232',
    '03/08/2026',
    '4290',
    'BOM',
    'OTIMO',
    'OTIMO',
    'OTIMO',
    'OTIMO',
    'OUTROS',
    '',
]

_COLUMN_ALIASES: dict[str, list[str]] = {
    'dataEnvio': ['data de envio', 'data envio'],
    'motorista': ['motorista'],
    'cte': ['chto', 'ct-e', 'cte', 'ct e'],
    'dataEntrega': ['data entrega', 'data de entrega', 'data_entrega'],
    'notaFiscal': [
        'ns notas fiscal',
        'nos notas fiscal',
        'notas fiscal',
        'nota fiscal',
        'nf',
        'nota_fiscal',
    ],
    'cliente': ['cliente'],
    'clienteRecusouAssinar': [
        'cliente recusou assinar',
        'recusou assinar',
        'recusa assinatura',
    ],
    'prazoEntrega': ['prazo de entrega', 'prazo'],
    'condicoesMercadoria': ['condicoes da mercadoria', 'condições da mercadoria', 'mercadoria'],
    'condicoesVeiculo': ['condicoes do veiculo', 'condições do veículo', 'veiculo', 'veículo'],
    'apresentacaoMotorista': ['apresentacao do motorista', 'apresentação do motorista', 'apresentacao'],
    'atendimentoDispensado': ['atendimento dispensado', 'atendimento'],
    'analise': [
        'analise tratativa e justificativa',
        'analise, tratativa e justificativa',
        'analise',
        'tratativa',
    ],
}

_HEADER_PROBE_FIELDS = ('dataEntrega', 'motorista', 'cte', 'notaFiscal', 'cliente')

_AVALIACAO_MAP = {
    'otimo': 'otimo',
    'bom': 'bom',
    'regular': 'regular',
    'ruim': 'ruim',
}

_CLIENTE_VALUES = {choice[0].upper(): choice[0] for choice in CLIENTE_CHOICES}


class PesquisaImportError(Exception):
    """Erro estrutural da planilha (cabeçalho inválido, arquivo corrompido)."""


def _norm(value) -> str:
    text = str(value or '').strip().lower()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _parse_date(value) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    raw = _norm(value)
    return raw in ('sim', 's', 'yes', 'true', '1', 'x')


def _parse_avaliacao(value) -> str:
    if value is None or value == '':
        return ''
    return _AVALIACAO_MAP.get(_norm(value), '')


def _parse_cliente(value) -> str:
    raw = str(value or '').strip().upper()
    if not raw:
        return 'OUTROS'
    return _CLIENTE_VALUES.get(raw, '')


def _parse_text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _find_column(normalized: list[str], aliases: list[str]) -> int | None:
    for alias in aliases:
        if alias in normalized:
            return normalized.index(alias)
    for idx, header in enumerate(normalized):
        for alias in aliases:
            if len(alias) >= 4 and alias in header:
                return idx
    return None


def _detect_columns(headers: list) -> dict[str, int]:
    normalized = [_norm(header) for header in headers]
    mapping: dict[str, int] = {}
    for field, aliases in _COLUMN_ALIASES.items():
        index = _find_column(normalized, aliases)
        if index is not None:
            mapping[field] = index

    missing = set(_HEADER_PROBE_FIELDS) - set(mapping)
    if missing:
        labels = ', '.join(sorted(missing))
        raise PesquisaImportError(
            f'Colunas obrigatórias não encontradas: {labels}. '
            'Use o layout oficial (DATA DE ENVIO, MOTORISTA, CHTO, DATA ENTREGA, '
            'Nºs NOTAS FISCAL, critérios, CLIENTE).',
        )
    return mapping


def _find_header_row(rows: list[tuple]) -> int:
    for idx, row in enumerate(rows[:25]):
        normalized_row = [_norm(cell) for cell in row if cell is not None]
        if not normalized_row:
            continue
        hits = sum(
            1 for field in _HEADER_PROBE_FIELDS
            if _find_column(normalized_row, _COLUMN_ALIASES[field]) is not None
        )
        if hits >= 3:
            return idx
    raise PesquisaImportError(
        'Cabeçalho da planilha não encontrado. Verifique se a linha de títulos contém '
        'DATA ENTREGA, MOTORISTA, CHTO e Nºs NOTAS FISCAL.',
    )


def _cell(row: tuple, index: int | None):
    if index is None:
        return None
    if index >= len(row):
        return None
    return row[index]


def _row_has_data(row: tuple) -> bool:
    return any(str(cell or '').strip() for cell in row)


def build_pesquisa_import_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pesquisas'

    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.append(_TEMPLATE_HEADERS)
    for col_idx in range(1, len(_TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.append(_TEMPLATE_SAMPLE_ROW)
    ws.freeze_panes = 'A2'

    widths = {
        'A': 14,
        'B': 28,
        'C': 10,
        'D': 14,
        'E': 16,
        'F': 14,
        'G': 16,
        'H': 16,
        'I': 18,
        'J': 18,
        'K': 14,
        'L': 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _format_date_br(value: date | None) -> str:
    if not value:
        return ''
    return value.strftime('%d/%m/%Y')


def _format_avaliacao_display(value: str) -> str:
    labels = {
        'otimo': 'OTIMO',
        'bom': 'BOM',
        'regular': 'REGULAR',
        'ruim': 'RUIM',
    }
    return labels.get(value, str(value or '').upper())


def parse_pesquisas_spreadsheet(file_bytes: bytes) -> dict:
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise PesquisaImportError(f'Erro ao abrir o arquivo: {exc}') from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise PesquisaImportError('Planilha vazia.')

    header_row_index = _find_header_row(rows)
    columns = _detect_columns(list(rows[header_row_index]))
    data_rows = rows[header_row_index + 1:]
    if not any(_row_has_data(row) for row in data_rows):
        raise PesquisaImportError('Nenhuma linha de dados encontrada na planilha.')

    errors: list[dict] = []
    payloads: list[dict] = []
    inclusion_dates: list[date] = []
    preview_rows: list[dict] = []
    skipped = 0
    default_inclusion = timezone.localdate()

    for offset, row in enumerate(data_rows, start=header_row_index + 2):
        if not _row_has_data(row):
            skipped += 1
            continue
        if len(payloads) >= MAX_IMPORT_ROWS:
            message = f'Limite de {MAX_IMPORT_ROWS} linhas por importação excedido.'
            errors.append({'row': offset, 'message': message})
            preview_rows.append({
                'row': offset,
                'valid': False,
                'message': message,
                'dataEnvio': _parse_text(_cell(row, columns.get('dataEnvio'))),
                'motorista': _parse_text(_cell(row, columns.get('motorista'))),
                'cte': _parse_text(_cell(row, columns.get('cte'))),
                'dataEntrega': _parse_text(_cell(row, columns.get('dataEntrega'))),
                'notaFiscal': _parse_text(_cell(row, columns.get('notaFiscal'))),
                'cliente': _parse_text(_cell(row, columns.get('cliente'))),
                'prazoEntrega': _parse_text(_cell(row, columns.get('prazoEntrega'))),
                'condicoesMercadoria': _parse_text(_cell(row, columns.get('condicoesMercadoria'))),
                'condicoesVeiculo': _parse_text(_cell(row, columns.get('condicoesVeiculo'))),
                'apresentacaoMotorista': _parse_text(_cell(row, columns.get('apresentacaoMotorista'))),
                'atendimentoDispensado': _parse_text(_cell(row, columns.get('atendimentoDispensado'))),
                'analise': _parse_text(_cell(row, columns.get('analise'))),
            })
            break

        raw_data_envio = _cell(row, columns.get('dataEnvio'))
        raw_data_entrega = _cell(row, columns.get('dataEntrega'))
        cliente_raw = _parse_text(_cell(row, columns.get('cliente')))
        cliente = _parse_cliente(cliente_raw)

        prazo = _parse_avaliacao(_cell(row, columns.get('prazoEntrega')))
        mercadoria = _parse_avaliacao(_cell(row, columns.get('condicoesMercadoria')))
        veiculo = _parse_avaliacao(_cell(row, columns.get('condicoesVeiculo')))
        apresentacao = _parse_avaliacao(_cell(row, columns.get('apresentacaoMotorista')))
        atendimento = _parse_avaliacao(_cell(row, columns.get('atendimentoDispensado')))

        preview_base = {
            'row': offset,
            'dataEnvio': _format_date_br(_parse_date(raw_data_envio)) or _parse_text(raw_data_envio),
            'motorista': _parse_text(_cell(row, columns.get('motorista'))),
            'cte': _parse_text(_cell(row, columns.get('cte'))),
            'dataEntrega': _format_date_br(_parse_date(raw_data_entrega)) or _parse_text(raw_data_entrega),
            'notaFiscal': _parse_text(_cell(row, columns.get('notaFiscal'))),
            'cliente': cliente_raw or 'OUTROS',
            'prazoEntrega': _format_avaliacao_display(prazo) or _parse_text(_cell(row, columns.get('prazoEntrega'))),
            'condicoesMercadoria': _format_avaliacao_display(mercadoria) or _parse_text(_cell(row, columns.get('condicoesMercadoria'))),
            'condicoesVeiculo': _format_avaliacao_display(veiculo) or _parse_text(_cell(row, columns.get('condicoesVeiculo'))),
            'apresentacaoMotorista': _format_avaliacao_display(apresentacao) or _parse_text(_cell(row, columns.get('apresentacaoMotorista'))),
            'atendimentoDispensado': _format_avaliacao_display(atendimento) or _parse_text(_cell(row, columns.get('atendimentoDispensado'))),
            'analise': _parse_text(_cell(row, columns.get('analise'))),
        }

        row_error = None
        if cliente_raw and not cliente:
            row_error = f'Cliente inválido: "{cliente_raw}". Use CCAB, PRENTISS, ALBAUGH ou OUTROS.'

        data_entrega = _parse_date(raw_data_entrega)
        if not row_error and not data_entrega:
            row_error = 'DATA ENTREGA inválida ou ausente.'

        payload = {
            'motorista': preview_base['motorista'],
            'cte': preview_base['cte'],
            'notaFiscal': preview_base['notaFiscal'],
            'cliente': cliente or 'OUTROS',
            'clienteRecusouAssinar': _parse_bool(_cell(row, columns.get('clienteRecusouAssinar'))),
            'prazoEntrega': prazo,
            'condicoesMercadoria': mercadoria,
            'condicoesVeiculo': veiculo,
            'apresentacaoMotorista': apresentacao,
            'atendimentoDispensado': atendimento,
            'analise': preview_base['analise'],
        }
        if data_entrega:
            payload['dataEntrega'] = data_entrega.isoformat()

        if not row_error:
            serializer = PesquisaSatisfacaoSerializer(data=payload)
            if not serializer.is_valid():
                parts = []
                for field_errors in serializer.errors.values():
                    if isinstance(field_errors, list):
                        parts.extend(str(item) for item in field_errors)
                    else:
                        parts.append(str(field_errors))
                row_error = '; '.join(parts) or 'Dados inválidos.'

        if row_error:
            errors.append({'row': offset, 'message': row_error})
            preview_rows.append({**preview_base, 'valid': False, 'message': row_error})
            continue

        data_envio = _parse_date(raw_data_envio)
        inclusion_dates.append(data_envio or default_inclusion)
        payloads.append(serializer.validated_data)
        preview_rows.append({**preview_base, 'valid': True, 'message': ''})

    valid_rows = len(payloads)
    invalid_rows = len(errors)
    success = invalid_rows == 0 and valid_rows > 0
    stats = _build_preview_stats(
        preview_rows,
        payloads,
        inclusion_dates,
        errors,
        skipped,
        success,
    )

    return {
        'success': success,
        'totalRows': len(preview_rows),
        'validRows': valid_rows,
        'invalidRows': invalid_rows,
        'skipped': skipped,
        'rows': preview_rows,
        'payloads': payloads,
        'inclusionDates': inclusion_dates,
        'errors': errors,
        'stats': stats,
        'detail': 'Corrija os erros na planilha e tente novamente.' if errors else None,
    }


def _build_preview_stats(
    preview_rows: list[dict],
    payloads: list[dict],
    inclusion_dates: list[date],
    errors: list[dict],
    skipped: int,
    success: bool,
) -> dict:
    from collections import Counter

    valid_preview = [row for row in preview_rows if row.get('valid')]
    processed = len(preview_rows)

    by_cliente = Counter(item.get('cliente') or 'OUTROS' for item in payloads)
    motoristas = {
        (item.get('motorista') or '').strip().upper()
        for item in payloads
        if (item.get('motorista') or '').strip()
    }

    entrega_dates = [item['data_entrega'] for item in payloads if item.get('data_entrega')]
    envio_dates = [value for value in inclusion_dates if value]

    avaliacao_counter: Counter = Counter()
    criterio_fields = (
        'prazo_entrega',
        'condicoes_mercadoria',
        'condicoes_veiculo',
        'apresentacao_motorista',
        'atendimento_dispensado',
    )
    for item in payloads:
        for field in criterio_fields:
            value = item.get(field)
            if value:
                avaliacao_counter[_format_avaliacao_display(value)] += 1

    duplicate_map: dict[tuple[str, str, str], list[int]] = {}
    for row in valid_preview:
        key = (row.get('cte') or '', row.get('notaFiscal') or '', row.get('dataEntrega') or '')
        duplicate_map.setdefault(key, []).append(row['row'])
    duplicate_groups = [
        {
            'cte': key[0],
            'notaFiscal': key[1],
            'dataEntrega': key[2],
            'rows': rows,
            'count': len(rows),
        }
        for key, rows in duplicate_map.items()
        if len(rows) > 1 and any(key)
    ]
    duplicate_groups.sort(key=lambda item: item['count'], reverse=True)

    error_counter = Counter(error['message'] for error in errors)
    rows_with_analise = sum(1 for item in payloads if (item.get('analise') or '').strip())

    return {
        'processedRows': processed,
        'validRows': len(valid_preview),
        'invalidRows': len(errors),
        'skippedEmptyRows': skipped,
        'validRate': round((len(valid_preview) / processed) * 100, 1) if processed else 0.0,
        'readyToImport': success,
        'uniqueMotoristas': len(motoristas),
        'rowsWithAnalise': rows_with_analise,
        'duplicateRowCount': sum(len(group['rows']) - 1 for group in duplicate_groups),
        'duplicateGroupCount': len(duplicate_groups),
        'byCliente': [
            {'cliente': cliente, 'count': count}
            for cliente, count in sorted(by_cliente.items(), key=lambda item: (-item[1], item[0]))
        ],
        'deliveryDateRange': {
            'min': _format_date_br(min(entrega_dates)) if entrega_dates else '',
            'max': _format_date_br(max(entrega_dates)) if entrega_dates else '',
        },
        'inclusionDateRange': {
            'min': _format_date_br(min(envio_dates)) if envio_dates else '',
            'max': _format_date_br(max(envio_dates)) if envio_dates else '',
        },
        'avaliacaoCounts': [
            {'label': label, 'count': count}
            for label, count in avaliacao_counter.most_common()
        ],
        'errorSummary': [
            {'message': message, 'count': count}
            for message, count in error_counter.most_common()
        ],
        'duplicateGroups': duplicate_groups[:8],
    }


def preview_pesquisas_from_spreadsheet(file_bytes: bytes) -> dict:
    parsed = parse_pesquisas_spreadsheet(file_bytes)
    return {
        'success': parsed['success'],
        'totalRows': parsed['totalRows'],
        'validRows': parsed['validRows'],
        'invalidRows': parsed['invalidRows'],
        'skipped': parsed['skipped'],
        'rows': parsed['rows'],
        'errors': parsed['errors'],
        'stats': parsed['stats'],
        'detail': parsed.get('detail'),
    }


def import_pesquisas_from_spreadsheet(
    file_bytes: bytes,
    *,
    filial: str,
    dry_run: bool = False,
) -> dict:
    parsed = parse_pesquisas_spreadsheet(file_bytes)
    skipped = parsed['skipped']

    if parsed['errors']:
        return {
            'success': False,
            'dryRun': dry_run,
            'created': 0,
            'skipped': skipped,
            'errors': parsed['errors'],
            'detail': parsed.get('detail'),
        }

    if dry_run:
        return {
            'success': True,
            'dryRun': True,
            'created': parsed['validRows'],
            'skipped': skipped,
            'errors': [],
        }

    with transaction.atomic():
        created_objects = []
        for item, data_inclusao in zip(parsed['payloads'], parsed['inclusionDates'], strict=True):
            created_objects.append(
                PesquisaSatisfacao.objects.create(
                    filial=filial,
                    data_inclusao=data_inclusao,
                    criado_por=CRIADO_POR_IMPORTACAO,
                    **item,
                )
            )

    return {
        'success': True,
        'dryRun': False,
        'created': len(created_objects),
        'skipped': skipped,
        'errors': [],
    }
