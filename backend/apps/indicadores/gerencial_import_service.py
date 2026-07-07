"""Importação de histórico gerencial a partir da planilha Gerenciais_Analitico.xlsx."""

from __future__ import annotations

import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from apps.financeiro.models import ReportBatch

from .cashflow_utils import cashflow_start_date, gerencial_pagar_cutoff
from .gerencial_email_service import save_gerencial_snapshot
from .models import GerencialSnapshot

ROW_FIELD_MAP = {
    'FATURAMENTO DO DIA': 'fat_dia',
    'FATURAMENTO DO MES': 'fat_mes',
    'FATURAMENTO DO ANO': 'fat_ano',
    'SALDO EM BANCO': 'saldo_banco',
    'DUPLICATAS A RECEBER': 'duplicatas_a_receber',
    'CTAS A REC ATRASADAS': 'ctas_rec_atrasadas',
    'CTES EMITIDOS': 'ctes_emitidos',
    'A DISPONIBILIZAR': 'a_disponibilizar',
    'CONTAS A PAGAR ATE CORTE': 'contas_pagar_ate_corte',
    'CONTAS A PAGAR 30 DIAS': 'contas_pagar_ate_corte',
    'CTAS A PAG ATRASADAS': 'ctas_pag_atrasadas',
    'SAIDAS PREVISTAS': 'saidas_previstas',
    'POSICAO GERENCIAL': 'posicao_gerencial',
}


def _norm_label(value) -> str:
    text = unicodedata.normalize('NFKD', str(value or ''))
    text = text.encode('ascii', 'ignore').decode().upper().strip()
    return text


def _parse_header_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value) -> Decimal:
    if value is None or value == '':
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace('.', '').replace(',', '.')
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal('0')


def _batch_label_for(reference: date) -> str:
    batch = (
        ReportBatch.objects.filter(reference_date=reference)
        .order_by('-created_at')
        .values_list('label', flat=True)
        .first()
    )
    return batch or ''


def parse_gerencial_history_workbook(file_path: str | Path | BinaryIO) -> list[dict]:
    """Lê a aba de histórico gerencial (formato exportado no e-mail)."""
    source = str(file_path) if isinstance(file_path, (str, Path)) else file_path
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return []

    header = rows[0]
    date_columns: list[tuple[int, date]] = []
    for col_idx, header_value in enumerate(header):
        if col_idx == 0:
            continue
        ref = _parse_header_date(header_value)
        if ref:
            date_columns.append((col_idx, ref))

    row_map: dict[str, int] = {}
    for row_idx, row in enumerate(rows):
        field = ROW_FIELD_MAP.get(_norm_label(row[0] if row else ''))
        if field:
            row_map[field] = row_idx

    snapshots: list[dict] = []
    for col_idx, reference in date_columns:
        data = {
            'reference_date': reference,
            'batch_label': _batch_label_for(reference),
            'pagar_cutoff_date': gerencial_pagar_cutoff(reference),
            'caixa_positivo_ate': '',
            'sent_by': 'import.planilha',
        }
        for field, row_idx in row_map.items():
            row = rows[row_idx]
            data[field] = _parse_amount(row[col_idx] if col_idx < len(row) else None)
        snapshots.append(data)

    return snapshots


def import_gerencial_history_file(
    file_path: str | Path,
    *,
    only_dates: set[date] | None = None,
    skip_existing: bool = False,
) -> dict:
    """
    Importa snapshots da planilha.
    Retorna resumo: created, updated, skipped, dates.
    """
    parsed = parse_gerencial_history_workbook(file_path)
    if only_dates is not None:
        parsed = [item for item in parsed if item['reference_date'] in only_dates]

    created = 0
    updated = 0
    skipped = 0
    imported_dates: list[date] = []

    for item in parsed:
        ref = item['reference_date']
        exists = GerencialSnapshot.objects.filter(reference_date=ref).exists()
        if exists and skip_existing:
            skipped += 1
            continue

        save_gerencial_snapshot(item)
        imported_dates.append(ref)
        if exists:
            updated += 1
        else:
            created += 1

    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'dates': sorted(imported_dates),
        'total_in_file': len(parsed),
    }
