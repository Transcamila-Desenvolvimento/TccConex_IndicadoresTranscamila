"""Indicador "Movimentação de RH": evolução mensal de quantitativo e folha
salarial a partir dos lotes de importação já existentes em apps.rh.

Cada `LoteMovimentacaoRH` é um snapshot mensal de colaboradores. Não existe
data de desligamento própria — admissões/desligamentos são inferidos
comparando o conjunto de CPFs de um lote com o do lote cronologicamente
anterior (mesma lógica usada em apps.rh.views.dashboard_summary), mas aqui
repetida para cada mês da série, não só para o último lote.
"""

from __future__ import annotations

import io
import re
import unicodedata
from decimal import Decimal

from django.db.models import Q

from apps.rh.models import Colaborador, LoteMovimentacaoRH, MovimentacaoColaborador

_MESES_ABREV = [
    '', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
    'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez',
]

_CATEGORIAS_VALIDAS = ('ADMINISTRATIVO', 'OPERACIONAL', 'MOTORISTA')
_CATEGORIA_PARA_CHAVE = {
    'ADMINISTRATIVO': 'administrativo',
    'OPERACIONAL': 'operacional',
    'MOTORISTA': 'motorista',
}
_CHAVE_NAO_MAPEADO = 'naoMapeado'

# Sem filtro de período explícito, mostra só os últimos N lotes — evolução
# recente é o caso de uso principal; histórico completo fica disponível via
# filtro "De".
JANELA_PADRAO_MESES = 12


def _label(mes: int, ano: int) -> str:
    return f'{_MESES_ABREV[mes]}/{ano}'


def _chave_periodo(ano: int, mes: int) -> int:
    return ano * 12 + mes


def _parse_periodo(valor) -> tuple[int, int] | None:
    """Converte "YYYY-MM" em (ano, mes). Retorna None se ausente/inválido."""
    if not valor:
        return None
    partes = str(valor).strip().split('-')
    if len(partes) != 2:
        return None
    try:
        ano, mes = int(partes[0]), int(partes[1])
    except ValueError:
        return None
    if mes < 1 or mes > 12:
        return None
    return ano, mes


def _status_bucket() -> dict:
    return {'count': 0, 'payroll': Decimal('0')}


def _bucket_categoria() -> dict:
    return {
        'count': 0,
        'payroll': Decimal('0'),
        'ativos': _status_bucket(),
        'afastados': _status_bucket(),
        'ferias': _status_bucket(),
    }


def _bucket_categorias() -> dict:
    return {
        'administrativo': _bucket_categoria(),
        'operacional': _bucket_categoria(),
        'motorista': _bucket_categoria(),
        'naoMapeado': _bucket_categoria(),
    }


# Grupos analíticos na folha:
# - afastado = situação contendo "AFASTADO TEMP."
# - férias = situação contendo "FERIAS" / "FÉRIAS" (com ou sem acento)
# - ativo = demais (SITUACAO NORMAL, ATIVO, AFASTADO INSS, etc.)
_AFASTADO_TEMP_TOKEN = 'AFASTADO TEMP'
_FERIAS_TOKENS = ('FERIAS', 'FÉRIAS', 'FERIA')
_FERIAS_Q = Q()
for _token in _FERIAS_TOKENS:
    _FERIAS_Q |= Q(situacao__icontains=_token)


def _normalizar_situacao_texto(situacao: str | None) -> str:
    if not situacao:
        return ''
    texto = unicodedata.normalize('NFKD', str(situacao))
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', texto.upper()).strip()


def _is_afastado(situacao: str | None) -> bool:
    return _AFASTADO_TEMP_TOKEN in _normalizar_situacao_texto(situacao)


def _is_ferias(situacao: str | None) -> bool:
    normalizado = _normalizar_situacao_texto(situacao)
    if not normalizado:
        return False
    return bool(re.search(r'\bFERIAS?\b', normalizado))


def _classificar_situacao(situacao: str | None) -> str:
    if _is_afastado(situacao):
        return 'afastados'
    if _is_ferias(situacao):
        return 'ferias'
    return 'ativos'


def _bucket_com_percentual(bucket: dict, total: int) -> dict:
    resultado = {}
    for chave, valores in bucket.items():
        percentual = round((valores['count'] / total) * 100, 1) if total else 0.0
        resultado[chave] = {**valores, 'percentual': percentual}
    return resultado


# Filtro de situação do indicador (query param `situacaoGrupo`).
_SITUACAO_GRUPO_TODOS = ''
_SITUACAO_GRUPO_AFASTADOS = 'AFASTADOS'
_SITUACAO_GRUPO_FERIAS = 'FERIAS'
_SITUACAO_GRUPO_NORMAL = 'SITUACAO_NORMAL'
_SITUACAO_GRUPOS_VALIDOS = {
    _SITUACAO_GRUPO_TODOS,
    _SITUACAO_GRUPO_AFASTADOS,
    _SITUACAO_GRUPO_FERIAS,
    _SITUACAO_GRUPO_NORMAL,
}


def _parse_situacao_grupo(valor) -> str:
    chave = (valor or '').strip().upper().replace(' ', '_')
    # Aliases vindos do frontend/legado.
    if chave in ('TODAS', 'TODOS', 'ALL'):
        return _SITUACAO_GRUPO_TODOS
    if chave in ('AFASTADO', 'AFASTADOS'):
        return _SITUACAO_GRUPO_AFASTADOS
    if chave in ('FERIAS', 'FÉRIAS', 'FERIA'):
        return _SITUACAO_GRUPO_FERIAS
    if chave in ('SITUACAO_NORMAL', 'SITUACAONORMAL', 'NORMAL'):
        return _SITUACAO_GRUPO_NORMAL
    if chave in _SITUACAO_GRUPOS_VALIDOS:
        return chave
    return _SITUACAO_GRUPO_TODOS


def _colaboradores_filtrados(
    lote: LoteMovimentacaoRH,
    *,
    filial: str,
    categoria: str,
    excluidos: set[str],
    situacao_grupo: str = _SITUACAO_GRUPO_TODOS,
):
    qs = lote.colaboradores.exclude(cpf__in=excluidos)
    if filial:
        qs = qs.filter(filial=filial)
    if categoria:
        qs = qs.filter(categoria=categoria)
    if situacao_grupo == _SITUACAO_GRUPO_AFASTADOS:
        qs = qs.filter(situacao__icontains=_AFASTADO_TEMP_TOKEN)
    elif situacao_grupo == _SITUACAO_GRUPO_FERIAS:
        qs = qs.filter(_FERIAS_Q)
    elif situacao_grupo == _SITUACAO_GRUPO_NORMAL:
        qs = qs.exclude(situacao__icontains=_AFASTADO_TEMP_TOKEN).exclude(_FERIAS_Q)
    return qs


def _variacao_percentual(atual, anterior):
    if not anterior:
        return None
    return round(float((atual - anterior) / anterior) * 100, 1)


def _empty_summary() -> dict:
    return {
        'totalColaboradores': 0,
        'payrollTotal': Decimal('0'),
        'salarioMedio': Decimal('0'),
        'feriasAtual': 0,
        'admitidosPeriodo': 0,
        'desligadosPeriodo': 0,
        'turnoverPercentual': 0.0,
        'variacaoHeadcountPercentual': None,
        'variacaoPayrollPercentual': None,
        'porCategoriaAtual': _bucket_com_percentual(_bucket_categorias(), 0),
    }


def build_rh_movimentacao_payload(params) -> dict:
    filial = (params.get('filial') or '').strip()
    categoria = (params.get('categoria') or '').strip().upper()
    if categoria not in _CATEGORIAS_VALIDAS:
        categoria = ''
    raw_situacao = params.get('situacaoGrupo')
    if raw_situacao is None:
        raw_situacao = params.get('situacao_grupo')
    situacao_grupo = _parse_situacao_grupo(raw_situacao)

    todos_lotes = list(LoteMovimentacaoRH.objects.order_by('ano', 'mes'))
    lotes_disponiveis = [
        {'mes': lote.mes, 'ano': lote.ano, 'label': _label(lote.mes, lote.ano)}
        for lote in todos_lotes
    ]
    filiais_disponiveis = sorted({
        f for f in MovimentacaoColaborador.objects
        .exclude(filial__isnull=True).exclude(filial='')
        .values_list('filial', flat=True).distinct()
    })

    meta_base = {
        'filiaisDisponiveis': filiais_disponiveis,
        'lotesDisponiveis': lotes_disponiveis,
    }

    if not todos_lotes:
        return {
            'meta': {**meta_base, 'periodoInicio': None, 'periodoFim': None},
            'summary': _empty_summary(),
            'series': [],
        }

    inicio = _parse_periodo(params.get('start'))
    fim = _parse_periodo(params.get('end'))

    lotes_periodo = todos_lotes
    if fim:
        fim_chave = _chave_periodo(*fim)
        lotes_periodo = [l for l in lotes_periodo if _chave_periodo(l.ano, l.mes) <= fim_chave]
    if inicio:
        inicio_chave = _chave_periodo(*inicio)
        lotes_periodo = [l for l in lotes_periodo if _chave_periodo(l.ano, l.mes) >= inicio_chave]
    if not inicio and not fim:
        lotes_periodo = lotes_periodo[-JANELA_PADRAO_MESES:]

    if not lotes_periodo:
        return {
            'meta': {**meta_base, 'periodoInicio': None, 'periodoFim': None},
            'summary': _empty_summary(),
            'series': [],
        }

    # Mapa lote -> lote cronologicamente anterior (pode estar fora do período
    # filtrado — necessário para calcular corretamente admitidos/desligados
    # do primeiro ponto da série exibida).
    lote_anterior_por_id: dict[int, LoteMovimentacaoRH | None] = {}
    for indice, lote in enumerate(todos_lotes):
        lote_anterior_por_id[lote.id] = todos_lotes[indice - 1] if indice > 0 else None

    excluidos = set(
        Colaborador.objects.filter(desconsiderado=True).values_list('cpf', flat=True)
    )

    filtro_lote = {
        'filial': filial,
        'categoria': categoria,
        'excluidos': excluidos,
        'situacao_grupo': situacao_grupo,
    }

    series = []
    for lote in lotes_periodo:
        colaboradores_lote = list(_colaboradores_filtrados(lote, **filtro_lote))
        cpfs_atual = {c.cpf for c in colaboradores_lote}

        lote_ant = lote_anterior_por_id.get(lote.id)
        if lote_ant is not None:
            cpfs_anterior = set(
                _colaboradores_filtrados(lote_ant, **filtro_lote)
                .values_list('cpf', flat=True)
            )
            admitidos = len(cpfs_atual - cpfs_anterior)
            desligados = len(cpfs_anterior - cpfs_atual)
        else:
            # Primeiro lote da história — não há base de comparação.
            admitidos = 0
            desligados = 0

        bucket = _bucket_categorias()
        payroll_total = Decimal('0')
        for c in colaboradores_lote:
            chave = _CATEGORIA_PARA_CHAVE.get((c.categoria or '').upper(), _CHAVE_NAO_MAPEADO)
            salario = c.salario or Decimal('0')
            bucket[chave]['count'] += 1
            bucket[chave]['payroll'] += salario
            status = _classificar_situacao(c.situacao)
            bucket[chave][status]['count'] += 1
            bucket[chave][status]['payroll'] += salario
            payroll_total += salario

        series.append({
            'mes': lote.mes,
            'ano': lote.ano,
            'label': _label(lote.mes, lote.ano),
            'headcount': len(colaboradores_lote),
            'payroll': payroll_total,
            'admitidos': admitidos,
            'desligados': desligados,
            'porCategoria': bucket,
        })

    primeiro, ultimo = series[0], series[-1]
    total_colaboradores = ultimo['headcount']
    payroll_total_atual = ultimo['payroll']
    salario_medio = (payroll_total_atual / total_colaboradores) if total_colaboradores else Decimal('0')

    admitidos_periodo = sum(p['admitidos'] for p in series)
    desligados_periodo = sum(p['desligados'] for p in series)
    headcount_medio = sum(p['headcount'] for p in series) / len(series)
    turnover_percentual = round((desligados_periodo / headcount_medio) * 100, 1) if headcount_medio else 0.0

    por_categoria_atual = _bucket_com_percentual(ultimo['porCategoria'], total_colaboradores)
    ferias_atual = sum(
        por_categoria_atual[chave]['ferias']['count']
        for chave in por_categoria_atual
    )

    summary = {
        'totalColaboradores': total_colaboradores,
        'payrollTotal': payroll_total_atual,
        'salarioMedio': round(salario_medio, 2),
        'feriasAtual': ferias_atual,
        'admitidosPeriodo': admitidos_periodo,
        'desligadosPeriodo': desligados_periodo,
        'turnoverPercentual': turnover_percentual,
        'variacaoHeadcountPercentual': (
            _variacao_percentual(ultimo['headcount'], primeiro['headcount']) if len(series) > 1 else None
        ),
        'variacaoPayrollPercentual': (
            _variacao_percentual(ultimo['payroll'], primeiro['payroll']) if len(series) > 1 else None
        ),
        'porCategoriaAtual': por_categoria_atual,
    }

    return {
        'meta': {**meta_base, 'periodoInicio': primeiro['label'], 'periodoFim': ultimo['label']},
        'summary': summary,
        'series': series,
    }


_GRUPO_SITUACAO_LABEL = {
    'ativos': 'Ativo',
    'afastados': 'Afastado',
    'ferias': 'Férias',
}


def parse_mes_ano_export_params(params) -> tuple[int, int]:
    """Resolve mes/ano a partir de query params (mes+ano ou referencia YYYY-MM)."""
    referencia = (params.get('referencia') or '').strip()
    if referencia:
        parsed = _parse_periodo(referencia)
        if not parsed:
            raise ValueError('Referência inválida. Use o formato YYYY-MM.')
        return parsed

    mes_raw = params.get('mes')
    ano_raw = params.get('ano')
    if mes_raw is None or ano_raw is None or mes_raw == '' or ano_raw == '':
        raise ValueError('Informe mes e ano, ou referencia no formato YYYY-MM.')

    try:
        mes, ano = int(mes_raw), int(ano_raw)
    except (TypeError, ValueError) as exc:
        raise ValueError('Mês e ano inválidos.') from exc

    if mes < 1 or mes > 12:
        raise ValueError('Mês inválido.')
    return ano, mes


def build_rh_movimentacao_export(mes: int, ano: int) -> tuple[bytes, str]:
    """Gera planilha Excel com dados brutos do lote mensal (base do indicador)."""
    lote = LoteMovimentacaoRH.objects.filter(mes=mes, ano=ano).first()
    if not lote:
        raise ValueError('Nenhum lote encontrado para o período informado.')

    excluidos = set(
        Colaborador.objects.filter(desconsiderado=True).values_list('cpf', flat=True)
    )
    colaboradores = (
        lote.colaboradores
        .exclude(cpf__in=excluidos)
        .order_by('filial', 'categoria', 'nome')
    )

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dados brutos'

    headers = [
        'Filial', 'Nome', 'CPF', 'Categoria', 'Situação', 'Grupo indicador',
        'Salário', 'Função', 'Admissão', 'Nascimento', 'UF', 'PIS/PASEP', 'RG',
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color='118CC4', end_color='118CC4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for colaborador in colaboradores:
        grupo = _classificar_situacao(colaborador.situacao)
        ws.append([
            colaborador.filial or 'Não Informada',
            (colaborador.nome or '').upper(),
            colaborador.cpf,
            colaborador.categoria or '-',
            colaborador.situacao or '-',
            _GRUPO_SITUACAO_LABEL.get(grupo, grupo),
            float(colaborador.salario) if colaborador.salario is not None else None,
            colaborador.funcao or '-',
            colaborador.data_admissao.strftime('%d/%m/%Y') if colaborador.data_admissao else '-',
            colaborador.data_nascimento.strftime('%d/%m/%Y') if colaborador.data_nascimento else '-',
            colaborador.uf_estado or '-',
            colaborador.pis_pasep or '-',
            colaborador.rg or '-',
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'Indicador_RH_Movimentacao_{mes:02d}_{ano}.xlsx'
    return output.read(), filename
