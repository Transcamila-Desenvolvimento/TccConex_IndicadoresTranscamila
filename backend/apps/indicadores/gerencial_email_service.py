"""Envio do relatório gerencial por e-mail com snapshot histórico."""

from __future__ import annotations

import io
import re
from datetime import date, timedelta
from decimal import Decimal

from django.conf import settings
from django.core.mail import EmailMessage
from django.db.models import Sum
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from apps.financeiro.models import BillingRecord, ReceberTitulo, ReportBatch
from apps.financeiro.report_filters import active_batch

from .cashflow_service import build_cashflow_payload
from .cashflow_utils import (
    cashflow_start_date,
    fmt_br,
    gerencial_fat_hoje_dates,
    gerencial_pagar_cutoff,
    parse_br_date,
)
from .gerencial_service import build_gerencial_panel
from .models import GerencialSnapshot

HISTORY_DAYS = 30
CLIENT_NAME_MAX = 30
CLIENT_NAME_TRIM = (
    ' LTDA', ' S.A.', ' ME', ' EPP', ' S/A',
    ' - EM RECUPERACAO JUDICIAL', ' EIRELI',
)


def _parse_reference(value: str | None) -> date:
    text = (value or '').strip()
    if not text:
        return timezone.localdate()
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError('Data inválida.') from exc


def _parse_emails(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        parts = value
    else:
        parts = re.split(r'[,;]+', str(value))
    return [p.strip() for p in parts if p and '@' in p]


def _fmt_currency(value) -> str:
    try:
        amount = float(value)
    except (TypeError, ValueError):
        amount = 0.0
    negative = amount < 0
    formatted = f'R$ {abs(amount):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f'({formatted})' if negative else formatted


def _fat_dia_periodo(reference: date) -> str:
    """Data(s) que o valor de "Faturamento do Dia" realmente cobre.

    O faturamento é sempre defasado: dias comuns mostram o dia anterior;
    segunda-feira acumula sexta a domingo (mesmas datas usadas no cálculo do
    valor, em ``gerencial_fat_hoje_dates``). Exibe só as datas com lançamento
    (ex.: segunda vira "sexta e sábado" quando domingo não tem faturamento).
    """
    dates = gerencial_fat_hoje_dates(reference)
    # set() dedupe: o Meta.ordering do BillingRecord (por data e filial) faz o
    # .distinct() do banco considerar a filial, repetindo a mesma data.
    com_lancamento = sorted(set(
        BillingRecord.objects.filter(reference_date__in=dates)
        .values_list('reference_date', flat=True)
    ))
    efetivas = com_lancamento or [dates[-1]]
    if len(efetivas) == 1:
        return fmt_br(efetivas[0])
    if len(efetivas) == 2:
        return f'{efetivas[0]:%d/%m} e {fmt_br(efetivas[-1])}'
    return f'{efetivas[0]:%d/%m} a {fmt_br(efetivas[-1])}'


def _clean_client_name(name: str) -> str:
    cleaned = (name or '').upper().strip()
    for term in CLIENT_NAME_TRIM:
        cleaned = cleaned.replace(term, '')
    return cleaned[:CLIENT_NAME_MAX] or '—'


def _panel_item(panel: dict, group_title: str, label_part: str) -> float:
    for group in panel.get('groups', []):
        if group.get('title') != group_title:
            continue
        for item in group.get('items', []):
            if label_part in item.get('label', ''):
                return float(item.get('value', 0))
    return 0.0


def _highlight_value(panel: dict, title_part: str) -> float:
    for card in panel.get('highlights', []):
        if title_part in card.get('title', ''):
            return float(card.get('value', 0))
    return 0.0


def _fat_ano(reference: date) -> Decimal:
    total = BillingRecord.objects.filter(
        reference_date__year=reference.year,
    ).aggregate(t=Sum('value'))['t']
    return Decimal(str(total or 0))


def _aggregate_receber_by_client(qs, *, reference: date, mode: str) -> list[dict]:
    buckets: dict[str, dict] = {}
    for row in qs.iterator():
        saldo = float(row.saldo or 0)
        if saldo <= 0:
            continue
        due = parse_br_date(row.vencimento_real)
        if not due:
            continue
        if mode == 'overdue':
            if due >= reference:
                continue
        elif mode == 'today':
            if due != reference:
                continue
        else:
            continue
        key = row.cod_cliente or row.cliente or row.titulo
        if key not in buckets:
            buckets[key] = {'client': row.cliente.strip(), 'value': 0.0}
        buckets[key]['value'] += saldo
    return sorted(buckets.values(), key=lambda x: -x['value'])


def build_snapshot_from_panel(
    panel: dict,
    *,
    reference: date,
    batch: ReportBatch | None,
    fat_ano: Decimal,
    caixa_positivo_ate: str,
    sent_by: str,
) -> dict:
    raw_ref = reference
    ref_flow = cashflow_start_date(reference)
    cutoff = gerencial_pagar_cutoff(raw_ref)
    return {
        'reference_date': raw_ref,
        'batch_label': batch.label if batch else '—',
        'fat_dia': Decimal(str(_panel_item(panel, 'Disponibilidade', 'Fat. Hoje'))),
        'fat_mes': Decimal(str(_panel_item(panel, 'Disponibilidade', 'Fat. Mês'))),
        'fat_ano': fat_ano,
        'saldo_banco': Decimal(str(_panel_item(panel, 'Disponibilidade', 'Saldo em Bancos'))),
        'duplicatas_a_receber': Decimal(str(_panel_item(panel, 'Futuro e Recebíveis', 'Duplicatas'))),
        'ctas_rec_atrasadas': Decimal(str(_panel_item(panel, 'Futuro e Recebíveis', 'Rec. Atraso'))),
        'ctes_emitidos': Decimal(str(_panel_item(panel, 'Futuro e Recebíveis', 'CTEs'))),
        'a_disponibilizar': Decimal(str(_highlight_value(panel, 'Disponibilizar'))),
        'contas_pagar_ate_corte': Decimal(str(_panel_item(panel, 'Compromissos', 'Contas a Pagar até'))),
        'pagar_cutoff_date': cutoff,
        'ctas_pag_atrasadas': Decimal(str(_panel_item(panel, 'Compromissos', 'Atrasadas'))),
        'saidas_previstas': Decimal(str(_highlight_value(panel, 'Saídas Previstas'))),
        'posicao_gerencial': Decimal(str(_highlight_value(panel, 'Posição Gerencial'))),
        'caixa_positivo_ate': caixa_positivo_ate,
        'sent_by': sent_by,
    }


def save_gerencial_snapshot(snapshot_data: dict) -> GerencialSnapshot:
    ref = snapshot_data['reference_date']
    obj, _ = GerencialSnapshot.objects.update_or_create(
        reference_date=ref,
        defaults=snapshot_data,
    )
    return obj


def build_history_workbook(reference: date) -> bytes:
    history = list(
        GerencialSnapshot.objects.order_by('-reference_date')[:HISTORY_DAYS]
    )
    history.reverse()

    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Histórico Gerencial'

    if not history:
        ws.cell(row=1, column=1, value='Sem snapshots enviados ainda.')
        wb.save(buffer)
        return buffer.getvalue()

    fill_blue = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    fill_green = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    fill_orange = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    font_bold = Font(bold=True)

    ws.cell(row=1, column=1, value=f'GERENCIAL {reference.year}').font = font_bold
    for col_idx, snap in enumerate(history, start=2):
        cell = ws.cell(row=1, column=col_idx, value=snap.reference_date.strftime('%d/%m/%y'))
        cell.font = font_bold
        cell.alignment = Alignment(horizontal='center')

    rows = [
        ('FATURAMENTO DO DIA', 'fat_dia', fill_blue),
        ('FATURAMENTO DO MÊS', 'fat_mes', fill_blue),
        ('FATURAMENTO DO ANO', 'fat_ano', fill_blue),
        ('SALDO EM BANCO', 'saldo_banco', None),
        ('DUPLICATAS À RECEBER', 'duplicatas_a_receber', None),
        ('CTAS A REC ATRASADAS', 'ctas_rec_atrasadas', None),
        ('CTEs EMITIDOS', 'ctes_emitidos', None),
        ('À DISPONIBILIZAR', 'a_disponibilizar', fill_green),
        ('CONTAS A PAGAR ATÉ CORTE', 'contas_pagar_ate_corte', None),
        ('CTAS A PAG ATRASADAS', 'ctas_pag_atrasadas', None),
        ('SAÍDAS PREVISTAS', 'saidas_previstas', fill_orange),
        ('POSIÇÃO GERENCIAL', 'posicao_gerencial', fill_yellow),
    ]

    for row_idx, (label, attr, fill) in enumerate(rows, start=2):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        if fill:
            label_cell.fill = fill
            label_cell.font = font_bold
        for col_idx, snap in enumerate(history, start=2):
            val = float(getattr(snap, attr))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = '#,##0.00'
            if fill:
                cell.fill = fill
                cell.font = font_bold

    perf_start = len(rows) + 2
    ws.cell(row=perf_start, column=1, value='PERFORMANCE DO DIA').font = font_bold
    ws.cell(row=perf_start + 1, column=1, value='PERFORMANCE DO MÊS ATUAL').font = font_bold
    ws.cell(row=perf_start + 2, column=1, value='PERFORMANCE DO ANO').font = font_bold

    for col_idx, snap in enumerate(history, start=2):
        prev = GerencialSnapshot.objects.filter(reference_date__lt=snap.reference_date).order_by('-reference_date').first()
        perf_dia = float(snap.posicao_gerencial - prev.posicao_gerencial) if prev else 0.0
        cell_d = ws.cell(row=perf_start, column=col_idx, value=perf_dia)
        cell_d.number_format = '#,##0.00'
        if perf_dia < 0:
            cell_d.font = Font(color='FF0000')

        last_day_prev_month = snap.reference_date.replace(day=1) - timedelta(days=1)
        start_month = GerencialSnapshot.objects.filter(reference_date__lte=last_day_prev_month).order_by('-reference_date').first()
        perf_mes = float(snap.posicao_gerencial - start_month.posicao_gerencial) if start_month else perf_dia
        cell_m = ws.cell(row=perf_start + 1, column=col_idx, value=perf_mes)
        cell_m.number_format = '#,##0.00'
        if perf_mes < 0:
            cell_m.font = Font(color='FF0000')

        last_day_prev_year = snap.reference_date.replace(month=1, day=1) - timedelta(days=1)
        start_year = GerencialSnapshot.objects.filter(reference_date__lte=last_day_prev_year).order_by('-reference_date').first()
        perf_ano = float(snap.posicao_gerencial - start_year.posicao_gerencial) if start_year else perf_mes
        cell_a = ws.cell(row=perf_start + 2, column=col_idx, value=perf_ano)
        cell_a.number_format = '#,##0.00'
        if perf_ano < 0:
            cell_a.font = Font(color='FF0000')

    ws.column_dimensions['A'].width = 32
    for col in range(2, len(history) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(buffer)
    return buffer.getvalue()


def send_gerencial_email(
    user,
    request,
    *,
    reference: date,
    to_emails: list[str],
    cc_emails: list[str] | None = None,
) -> GerencialSnapshot:
    if not to_emails:
        raise ValueError('Informe ao menos um destinatário.')

    batch = active_batch()
    if not batch:
        raise ValueError('Nenhum lote ativo para gerar o relatório gerencial.')

    iso = reference.isoformat()
    panel = build_gerencial_panel(batch, [], reference)
    cashflow = build_cashflow_payload(user, request, {'gerencialDate': iso})
    caixa_positivo = cashflow.get('summary', {}).get('caixaPositivoAte', '—')

    receber_qs = ReceberTitulo.objects.filter(batch=batch)
    ref_flow = cashflow_start_date(reference)
    overdue_list = _aggregate_receber_by_client(receber_qs, reference=ref_flow, mode='overdue')
    today_list = _aggregate_receber_by_client(receber_qs, reference=reference, mode='today')

    snapshot_data = build_snapshot_from_panel(
        panel,
        reference=reference,
        batch=batch,
        fat_ano=_fat_ano(reference),
        caixa_positivo_ate=caixa_positivo,
        sent_by=getattr(user, 'username', '') or 'Sistema',
    )
    snapshot = save_gerencial_snapshot(snapshot_data)

    cutoff = snapshot.pagar_cutoff_date or gerencial_pagar_cutoff(reference)
    dashboard_base = getattr(settings, 'FRONTEND_BASE_URL', 'http://localhost:5173').rstrip('/')
    dashboard_url = f'{dashboard_base}/indicadores/fluxo-de-caixa'

    email_context = {
        'batch_label': snapshot.batch_label,
        'ref_date': reference,
        'fat_dia': _fmt_currency(snapshot.fat_dia),
        'fat_dia_date': _fat_dia_periodo(reference),
        'fat_mes': _fmt_currency(snapshot.fat_mes),
        'fat_ano': _fmt_currency(snapshot.fat_ano),
        'saldo_banco': _fmt_currency(snapshot.saldo_banco),
        'duplicatas_a_receber': _fmt_currency(snapshot.duplicatas_a_receber),
        'contas_receber_atrasadas': _fmt_currency(snapshot.ctas_rec_atrasadas),
        'ctes_emitidos': _fmt_currency(snapshot.ctes_emitidos),
        'total_a_disponibilizar': _fmt_currency(snapshot.a_disponibilizar),
        'contas_pagar_ate': _fmt_currency(snapshot.contas_pagar_ate_corte),
        'contas_pagar_atrasadas': _fmt_currency(snapshot.ctas_pag_atrasadas),
        'total_saidas_previstas': _fmt_currency(snapshot.saidas_previstas),
        'posicao_gerencial': _fmt_currency(snapshot.posicao_gerencial),
        'last_positive_day': snapshot.caixa_positivo_ate or caixa_positivo,
        'date_limit': cutoff,
        'overdue_list': [
            {'client': _clean_client_name(x['client']), 'value': _fmt_currency(x['value'])}
            for x in overdue_list[:50]
        ],
        'today_list': [
            {'client': _clean_client_name(x['client']), 'value': _fmt_currency(x['value'])}
            for x in today_list[:50]
        ],
        'total_vencido': _fmt_currency(sum(x['value'] for x in overdue_list)),
        'total_hoje': _fmt_currency(sum(x['value'] for x in today_list)),
        'dashboard_url': dashboard_url,
    }

    html_body = render_to_string('indicadores/emails/relatorio_gerencial.html', email_context)
    excel_bytes = build_history_workbook(reference)

    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'digitalmidia@transcamila.com.br')
    subject = f'RELATÓRIO GERENCIAL - {reference:%d/%m/%Y} - Lote {snapshot.batch_label}'

    message = EmailMessage(
        subject=subject,
        body=html_body,
        from_email=from_email,
        to=to_emails,
        cc=cc_emails or [],
    )
    message.content_subtype = 'html'
    message.attach('Gerenciais_Analitico.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    message.send(fail_silently=False)

    return snapshot
