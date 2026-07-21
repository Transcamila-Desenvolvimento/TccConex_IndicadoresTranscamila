from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient

from apps.accounts.tests import auth_headers
from apps.audit.models import AuditLog
from apps.financeiro.models import PagarTitulo, ReceberTitulo, ReportBatch
from apps.financeiro.import_service import _normalize_filial, _normalize_origem, import_report_file

User = get_user_model()


class FinanceiroReportTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='fin_reports',
            password='fin123',
            role_id='2',
            environments=['Financeiro', 'Faturamento'],
            filiais={
                'Financeiro': ['Ibiporã (Matriz)', 'Rondonópolis'],
                'Faturamento': ['Ibiporã (Matriz)', 'Rondonópolis'],
            },
        )
        self.batch = ReportBatch.objects.create(
            label='##T01',
            reference_date=date(2026, 6, 10),
            is_active=True,
            imported_pagar=True,
        )
        for index in range(15):
            PagarTitulo.objects.create(
                batch=self.batch,
                filial='01' if index % 2 == 0 else '05',
                cod_forn=f'F{index:03d}',
                fornecedor=f'Fornecedor {index}',
                titulo=f'T{index:04d}',
                tipo='NF',
                emissao='01/06/2026',
                vencimento='10/06/2026',
                vencimento_real='10/06/2026',
                valor=Decimal('100.00'),
                saldo=Decimal('100.00'),
                historico='',
            )

    def test_pagar_report_returns_paginated_shape(self):
        response = self.client.get(
            '/api/financeiro/reports/pagar/?page=1&page_size=10',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 15)
        self.assertEqual(len(response.data['results']), 10)
        self.assertIn('codForn', response.data['results'][0])

    def test_finalize_batch_creates_audit_log(self):
        other = ReportBatch.objects.create(
            label='##T02',
            reference_date=date(2026, 6, 9),
            is_active=False,
        )
        before = AuditLog.objects.count()
        response = self.client.post(
            f'/api/financeiro/batches/{other.pk}/finalize/',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(AuditLog.objects.count(), before + 1)
        self.assertTrue(AuditLog.objects.filter(action='financeiro.lote.ativado').exists())

    def test_prepare_import_creates_first_batch(self):
        ReportBatch.objects.all().delete()
        response = self.client.post(
            '/api/financeiro/batches/prepare_import/',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['label'], '##001')
        self.assertFalse(response.data['isActive'])
        self.assertEqual(ReportBatch.objects.count(), 1)

    def test_prepare_import_reuses_same_day_batch(self):
        ReportBatch.objects.all().delete()
        first = self.client.post(
            '/api/financeiro/batches/prepare_import/',
            **auth_headers(self.user, 'Financeiro'),
        )
        second = self.client.post(
            '/api/financeiro/batches/prepare_import/',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(first.data['id'], second.data['id'])
        self.assertEqual(ReportBatch.objects.count(), 1)

    def test_batches_list_includes_max_batches(self):
        response = self.client.get(
            '/api/financeiro/batches/',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['maxBatches'], 5)
        self.assertIn('results', response.data)

    def test_cash_adjustment_create_creates_audit_log(self):
        before = AuditLog.objects.count()
        response = self.client.post(
            '/api/financeiro/adjustments/',
            {
                'date': '2026-06-10',
                'type': 'Entrada',
                'value': 1500,
                'observation': 'Teste auditoria',
                'user': 'fin_reports',
            },
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(AuditLog.objects.count(), before + 1)
        self.assertTrue(AuditLog.objects.filter(action='financeiro.ajuste.criado').exists())

    def test_billing_list_returns_paginated_shape(self):
        from apps.financeiro.models import BillingRecord
        for index in range(12):
            BillingRecord.objects.create(
                reference_date=date(2026, 6, index + 1),
                branch='Ibiporã (Matriz)',
                value=Decimal('1000.00') + index,
                notes_count=index,
            )
        response = self.client.get(
            '/api/financeiro/billing/?page=1&page_size=10',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 12)
        self.assertEqual(len(response.data['results']), 10)
        self.assertIn('trend', response.data['results'][0])

    def test_billing_manual_create(self):
        response = self.client.post(
            '/api/financeiro/billing/',
            {
                'date': '2026-07-09',
                'branch': 'Ibiporã',
                'value': '1500.50',
                'notesCount': 12,
            },
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['branch'], 'Ibiporã')
        self.assertTrue(AuditLog.objects.filter(action='financeiro.faturamento.criado').exists())

    def test_billing_manual_create_rejects_duplicate_branch_date(self):
        from apps.financeiro.models import BillingRecord
        BillingRecord.objects.create(
            reference_date=date(2026, 7, 9),
            branch='Barueri',
            value=Decimal('100'),
            notes_count=1,
        )
        response = self.client.post(
            '/api/financeiro/billing/',
            {
                'date': '2026-07-09',
                'branch': 'Barueri',
                'value': '200',
                'notesCount': 2,
            },
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 400)

    def test_normalize_filial_preserves_code(self):
        self.assertEqual(_normalize_filial(1), '01')
        self.assertEqual(_normalize_filial('03'), '03')
        self.assertEqual(_normalize_filial(''), '')

    def test_normalize_origem_preserves_code(self):
        self.assertEqual(_normalize_origem(9), '9')
        self.assertEqual(_normalize_origem('11'), '11')
        self.assertEqual(_normalize_origem(''), '')

    def test_import_rejects_wrong_report_type(self):
        batch = ReportBatch.objects.create(
            label='##IMP',
            reference_date=date(2026, 6, 11),
        )
        # Cabeçalho mínimo de pagar — não deve passar como receber
        from io import BytesIO
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Filial Orig.', 'No. Titulo', 'Tipo', 'Fornecedor', 'Nome Fornece', 'Vlr.Titulo'])
        ws.append(['01', 'T001', 'NF', 'F001', 'Fornecedor Teste', 100])
        buf = BytesIO()
        wb.save(buf)
        result = import_report_file(batch, 'receber', buf.getvalue(), 'pagar.xlsx')
        self.assertFalse(result['success'])
        self.assertTrue(any(i['severity'] == 'error' for i in result['issues']))
        batch.delete()

    def test_pagar_empty_filial_with_pa_tipo_uses_pa(self):
        batch = ReportBatch.objects.create(
            label='##PA',
            reference_date=date(2026, 6, 11),
        )
        from io import BytesIO
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            'Filial Orig.', 'No. Titulo', 'Tipo', 'Fornecedor', 'Nome Fornece',
            'DT Emissao', 'Vencimento', 'Vencto Real', 'Vlr.Titulo', 'Historico', 'Saldo',
        ])
        ws.append(['', '003420', 'PA', 'AM0229', 'CARLOS CESAR NUNES', '24/03/2025', '24/03/2025', '24/03/2025', 100, '', 100])
        buf = BytesIO()
        wb.save(buf)
        result = import_report_file(batch, 'pagar', buf.getvalue(), 'pagar_pa.xlsx')
        self.assertTrue(result['success'])
        titulo = PagarTitulo.objects.get(batch=batch)
        self.assertEqual(titulo.filial, 'PA')
        self.assertEqual(titulo.tipo, 'PA')
        batch.delete()

    def test_import_receber_does_not_double_vencido_saldo(self):
        batch = ReportBatch.objects.create(
            label='##RCV',
            reference_date=date(2026, 6, 11),
        )
        from io import BytesIO
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            'Filial Orig.', 'No. Titulo', 'Natureza', 'Cliente', 'Nome Cliente',
            'Data de Emissao', 'Vencto Orig', 'Vencto Real', 'Vlr.Titulo', 'Historico',
            'Titulos a Vencer Valor Atual', 'Tit Vencidos Valor Corrigido',
        ])
        ws.append([
            '01', '000123', 'NF', 'C1', 'Cliente Teste',
            '01/05/2026', '01/05/2026', '01/05/2026', 1776.45, '',
            1776.45, 1776.45,
        ])
        buf = BytesIO()
        wb.save(buf)
        result = import_report_file(batch, 'receber', buf.getvalue(), 'receber.xlsx')
        self.assertTrue(result['success'])
        titulo = ReceberTitulo.objects.get(batch=batch)
        self.assertEqual(titulo.saldo, Decimal('1776.45'))
        batch.delete()

    def test_adjustments_list_returns_paginated_shape(self):
        from apps.financeiro.models import CashAdjustment
        for index in range(11):
            CashAdjustment.objects.create(
                reference_date=date(2026, 6, 10),
                adjustment_type='Entrada' if index % 2 == 0 else 'Saída',
                value=Decimal('100.00'),
                observation=f'Ajuste {index}',
                created_by='fin_reports',
            )
        response = self.client.get(
            '/api/financeiro/adjustments/?page=1&page_size=10',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 11)
        self.assertEqual(len(response.data['results']), 10)


class BillingGlobalAccessTests(TestCase):
    """Faturamento diário é consolidado: operadores veem todas as filiais de billing."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='fat_scoped',
            password='fat123',
            role_id='2',
            environments=['Financeiro'],
            filiais={'Financeiro': ['Ibiporã (Matriz)']},
        )
        from apps.financeiro.models import BillingRecord
        BillingRecord.objects.create(
            reference_date=date(2026, 6, 1), branch='Ibiporã', value=Decimal('1000'), notes_count=1,
        )
        BillingRecord.objects.create(
            reference_date=date(2026, 6, 1), branch='Rondonópolis', value=Decimal('2000'), notes_count=2,
        )
        BillingRecord.objects.create(
            reference_date=date(2026, 6, 1), branch='Barueri', value=Decimal('3000'), notes_count=3,
        )
        BillingRecord.objects.create(
            reference_date=date(2026, 6, 1), branch='Armazém', value=Decimal('4000'), notes_count=4,
        )

    def test_operator_sees_all_billing_branches(self):
        response = self.client.get(
            '/api/financeiro/billing/?pagination=none',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        branches = {row['branch'] for row in response.data}
        self.assertEqual(branches, {'Ibiporã', 'Rondonópolis', 'Barueri', 'Armazém'})

    def test_admin_sees_all_billing_branches(self):
        admin = User.objects.create_user(
            username='fat_admin_scope',
            password='admin123',
            role_id='1',
            environments=['Administração/Manutenção', 'Financeiro'],
            filiais={'Financeiro': ['Ibiporã (Matriz)']},
        )
        response = self.client.get(
            '/api/financeiro/billing/?pagination=none',
            **auth_headers(admin, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        branches = {row['branch'] for row in response.data}
        self.assertEqual(branches, {'Ibiporã', 'Rondonópolis', 'Barueri', 'Armazém'})


class PrAnalysisTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='fin_pr',
            password='fin123',
            role_id='2',
            environments=['Financeiro'],
            filiais={'Financeiro': ['Ibiporã (Matriz)']},
        )
        self.batch = ReportBatch.objects.create(
            label='##PR1',
            reference_date=date(2026, 6, 12),
            is_active=True,
            imported_pagar=True,
        )
        PagarTitulo.objects.create(
            batch=self.batch,
            filial='01',
            cod_forn='F1',
            fornecedor='Fornecedor A',
            titulo='PR001',
            tipo='PR',
            emissao='01/06/2026',
            vencimento='12/06/2026',
            vencimento_real='12/06/2026',
            valor=Decimal('500.00'),
            saldo=Decimal('500.00'),
            historico='',
        )
        PagarTitulo.objects.create(
            batch=self.batch,
            filial='01',
            cod_forn='F1',
            fornecedor='Fornecedor A',
            titulo='NF001',
            tipo='NF',
            emissao='01/06/2026',
            vencimento='12/06/2026',
            vencimento_real='12/06/2026',
            valor=Decimal('500.00'),
            saldo=Decimal('500.00'),
            historico='',
        )

    def test_analise_prs_finds_duplicate(self):
        response = self.client.get(
            '/api/financeiro/reports/analise-prs/',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['totalDuplicates'], 1)
        self.assertEqual(response.data['duplicates'][0]['titulo'], 'PR001')
        self.assertEqual(len(response.data['duplicates'][0]['matches']), 1)

    def test_pr_action_ignore_creates_audit_log(self):
        pr = PagarTitulo.objects.get(titulo='PR001')
        before = AuditLog.objects.count()
        response = self.client.post(
            '/api/financeiro/reports/pr-action/',
            {'ids': [pr.id], 'action': 'ignore'},
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(AuditLog.objects.count(), before + 1)
        self.assertTrue(AuditLog.objects.filter(action='financeiro.pr.ignorado').exists())

    def test_pr_action_ignore_excludes_from_cashflow(self):
        pr = PagarTitulo.objects.get(titulo='PR001')
        response = self.client.post(
            '/api/financeiro/reports/pr-action/',
            {'ids': [pr.id], 'action': 'ignore'},
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        pr.refresh_from_db()
        self.assertTrue(pr.pr_desconsiderada)

        from apps.indicadores.cashflow_service import build_cashflow_payload
        from types import SimpleNamespace

        request = SimpleNamespace(META={})
        payload = build_cashflow_payload(self.user, request, {'start': '2026-06-12', 'end': '2026-06-12'})
        day = payload['daily'][0]
        self.assertEqual(day['saidas'], 500.0)

    def test_pr_action_propagates_to_recent_batches(self):
        older = ReportBatch.objects.create(
            label='##PR0',
            reference_date=date(2026, 6, 11),
            is_active=False,
            imported_pagar=True,
        )
        PagarTitulo.objects.create(
            batch=older,
            filial='01',
            cod_forn='F1',
            fornecedor='Fornecedor A',
            titulo='PR001',
            tipo='PR',
            emissao='01/06/2026',
            vencimento='12/06/2026',
            vencimento_real='12/06/2026',
            valor=Decimal('500.00'),
            saldo=Decimal('500.00'),
            historico='',
        )
        pr = PagarTitulo.objects.get(batch=self.batch, titulo='PR001')
        response = self.client.post(
            '/api/financeiro/reports/pr-action/',
            {'ids': [pr.id], 'action': 'ignore'},
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        older_pr = PagarTitulo.objects.get(batch=older, titulo='PR001')
        self.assertTrue(older_pr.pr_desconsiderada)

    def test_import_carries_ignored_pr_to_new_batch(self):
        from io import BytesIO

        import openpyxl

        pr = PagarTitulo.objects.get(batch=self.batch, titulo='PR001')
        self.client.post(
            '/api/financeiro/reports/pr-action/',
            {'ids': [pr.id], 'action': 'ignore'},
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )

        newer = ReportBatch.objects.create(
            label='##PR2',
            reference_date=date(2026, 6, 13),
            is_active=True,
            imported_pagar=False,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            'Filial Orig.', 'No. Titulo', 'Tipo', 'Fornecedor', 'Nome Fornece',
            'DT Emissao', 'Vencimento', 'Vencto Real', 'Vlr.Titulo', 'Historico', 'Saldo',
        ])
        ws.append(['01', 'PR001', 'PR', 'F1', 'Fornecedor A', '01/06/2026', '12/06/2026', '12/06/2026', 500, '', 500])
        buf = BytesIO()
        wb.save(buf)
        result = import_report_file(newer, 'pagar', buf.getvalue(), 'pagar_pr.xlsx')
        self.assertTrue(result['success'])
        imported = PagarTitulo.objects.get(batch=newer, titulo='PR001')
        self.assertTrue(imported.pr_desconsiderada)


class PagarDiffAnalysisTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='fin_diff',
            password='fin123',
            role_id='2',
            environments=['Financeiro'],
            filiais={'Financeiro': ['Ibiporã (Matriz)']},
        )
        self.previous = ReportBatch.objects.create(
            label='##D01',
            reference_date=date(2026, 6, 10),
            is_active=False,
            imported_pagar=True,
        )
        self.current = ReportBatch.objects.create(
            label='##D02',
            reference_date=date(2026, 6, 12),
            is_active=True,
            imported_pagar=True,
        )
        PagarTitulo.objects.create(
            batch=self.previous,
            filial='01',
            cod_forn='F1',
            fornecedor='Fornecedor A',
            titulo='T-OLD',
            tipo='NF',
            emissao='01/06/2026',
            vencimento='12/06/2026',
            vencimento_real='12/06/2026',
            valor=Decimal('100.00'),
            saldo=Decimal('100.00'),
            historico='',
        )
        PagarTitulo.objects.create(
            batch=self.current,
            filial='01',
            cod_forn='F1',
            fornecedor='Fornecedor A',
            titulo='T-NEW',
            tipo='PR',
            emissao='01/06/2026',
            vencimento='12/06/2026',
            vencimento_real='12/06/2026',
            valor=Decimal('250.00'),
            saldo=Decimal('250.00'),
            historico='',
        )
        PagarTitulo.objects.create(
            batch=self.current,
            filial='01',
            cod_forn='F2',
            fornecedor='Fornecedor B',
            titulo='T-REP',
            tipo='NF',
            emissao='01/06/2026',
            vencimento='15/06/2026',
            vencimento_real='15/06/2026',
            valor=Decimal('300.00'),
            saldo=Decimal('300.00'),
            historico='',
        )
        PagarTitulo.objects.create(
            batch=self.previous,
            filial='01',
            cod_forn='F2',
            fornecedor='Fornecedor B',
            titulo='T-REP',
            tipo='NF',
            emissao='01/06/2026',
            vencimento='10/06/2026',
            vencimento_real='10/06/2026',
            valor=Decimal('300.00'),
            saldo=Decimal('300.00'),
            historico='',
        )

    def test_pagar_diff_lists_new_titles_and_reprogrammed(self):
        response = self.client.get(
            '/api/financeiro/reports/pagar/diff/?start=2026-06-12&end=2026-06-12',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['currentBatch']['label'], '##D02')
        self.assertEqual(response.data['previousBatch']['label'], '##D01')
        self.assertEqual(response.data['summary']['novosTitulos'], 250.0)
        day = next(item for item in response.data['days'] if item['date'] == '2026-06-12')
        self.assertEqual(len(day['novosTitulos']), 1)
        self.assertEqual(day['novosTitulos'][0]['titulo'], 'T-NEW')

    def test_pagar_diff_monday_rollup_and_reprogram_from_position(self):
        """Coluna do lote anterior: rollup de fim de semana na segunda + reprogramação da posição."""
        previous = ReportBatch.objects.create(
            label='##D10',
            reference_date=date(2026, 6, 19),
            is_active=False,
            imported_pagar=True,
        )
        current = ReportBatch.objects.create(
            label='##D11',
            reference_date=date(2026, 6, 22),
            is_active=True,
            imported_pagar=True,
        )
        PagarTitulo.objects.create(
            batch=previous, filial='01', cod_forn='F1', fornecedor='A', titulo='T-SUN',
            tipo='OP', emissao='01/06/2026', vencimento='21/06/2026', vencimento_real='21/06/2026',
            valor=Decimal('300.00'), saldo=Decimal('300.00'), historico='',
        )
        PagarTitulo.objects.create(
            batch=previous, filial='01', cod_forn='F2', fornecedor='B', titulo='T-PR',
            tipo='PR', emissao='01/06/2026', vencimento='19/06/2026', vencimento_real='19/06/2026',
            valor=Decimal('1000.00'), saldo=Decimal('1000.00'), historico='',
        )
        PagarTitulo.objects.create(
            batch=current, filial='01', cod_forn='F2', fornecedor='B', titulo='T-PR',
            tipo='PR', emissao='01/06/2026', vencimento='26/06/2026', vencimento_real='26/06/2026',
            valor=Decimal('1000.00'), saldo=Decimal('1000.00'), historico='',
        )

        response = self.client.get(
            f'/api/financeiro/reports/pagar/diff/?batchId={current.pk}&start=2026-06-22&end=2026-06-26',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        monday = next(item for item in response.data['days'] if item['date'] == '2026-06-22')
        friday = next(item for item in response.data['days'] if item['date'] == '2026-06-26')
        self.assertEqual(monday['totalPrevious'], 300.0)
        self.assertEqual(friday['totalPrevious'], 1000.0)

    def test_pagar_diff_previous_day_skips_extra_when_title_already_on_same_due(self):
        """Não duplica saldo se o título já existia no mesmo vencimento no lote anterior."""
        previous = ReportBatch.objects.create(
            label='##D20',
            reference_date=date(2026, 6, 19),
            is_active=False,
            imported_pagar=True,
        )
        current = ReportBatch.objects.create(
            label='##D21',
            reference_date=date(2026, 6, 22),
            is_active=True,
            imported_pagar=True,
        )
        for batch, due in ((previous, '19/06/2026'), (previous, '17/07/2026'), (current, '17/07/2026')):
            PagarTitulo.objects.create(
                batch=batch,
                filial='01',
                cod_forn='FM0552',
                fornecedor='Fornecedor',
                titulo='000195535',
                tipo='NF',
                emissao='01/06/2026',
                vencimento=due,
                vencimento_real=due,
                valor=Decimal('537.05'),
                saldo=Decimal('537.05'),
                historico='',
            )

        response = self.client.get(
            f'/api/financeiro/reports/pagar/diff/?batchId={current.pk}&start=2026-07-17&end=2026-07-17',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        day = response.data['days'][0]
        self.assertEqual(day['totalCurrent'], 537.05)
        self.assertEqual(day['totalPrevious'], 537.05)
        self.assertEqual(day['diff'], 0.0)

    def test_pagar_diff_requires_pagar_import(self):
        batch = ReportBatch.objects.create(
            label='##D03',
            reference_date=date(2026, 6, 13),
            is_active=False,
            imported_pagar=False,
        )
        response = self.client.get(
            f'/api/financeiro/reports/pagar/diff/?batchId={batch.pk}',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 400)

    def test_pagar_diff_default_range_is_batch_date_plus_one_month(self):
        response = self.client.get(
            '/api/financeiro/reports/pagar/diff/',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['dateStart'], '2026-06-12')
        self.assertEqual(response.data['dateEnd'], '2026-07-12')


class BillingImportBranchTests(TestCase):
    def _row_xml(self, transportadora: str, filial_emissao: str = '') -> str:
        filial_tag = (
            f'<Filial_x0020_emissão>{filial_emissao}</Filial_x0020_emissão>'
            if filial_emissao
            else ''
        )
        return (
            '<Row>'
            '<Emissão_x0020_CT_x002d_e>10/06/2026</Emissão_x0020_CT_x002d_e>'
            f'<Transportadora>{transportadora}</Transportadora>'
            f'{filial_tag}'
            '<Usuário_x0020_cadastro>OPERADOR</Usuário_x0020_cadastro>'
            '<Valor_x0020_Frete>1.000,00</Valor_x0020_Frete>'
            '</Row>'
        )

    def test_armazem_code_10_not_grouped_as_rondonopolis(self):
        from apps.financeiro.billing_import_service import _detect_branch, parse_billing_xml

        self.assertEqual(
            _detect_branch('10 - TRANSCAMILA - RONDONOPOLIS', ''),
            'Armazém',
        )
        parsed = parse_billing_xml(self._row_xml('10 - TRANSCAMILA - RONDONOPOLIS'))
        self.assertEqual(parsed['10/06/2026']['Armazém']['count'], 1)
        self.assertNotIn('Rondonópolis', parsed['10/06/2026'])

    def test_rondonopolis_code_5_maps_correctly(self):
        from apps.financeiro.billing_import_service import _detect_branch, parse_billing_xml

        self.assertEqual(
            _detect_branch('5 - TRANSCAMILA - RONDONOPOLIS', ''),
            'Rondonópolis',
        )
        parsed = parse_billing_xml(self._row_xml('5 - TRANSCAMILA - RONDONOPOLIS'))
        self.assertEqual(parsed['10/06/2026']['Rondonópolis']['count'], 1)

    def test_official_filial_code_mapping(self):
        from apps.financeiro.billing_import_service import _detect_branch

        cases = {
            '1 - TRANSCAMILA - IBIPORA': 'Ibiporã',
            '11 - TRANSCAMILA - PARANAGUA': 'Paranaguá',
            '5 - TRANSCAMILA - RONDONOPOLIS': 'Rondonópolis',
            '10 - TRANSCAMILA - RONDONOPOLIS': 'Armazém',
            '9 - TRANSCAMILA - BARUERI': 'Barueri',
        }
        for transportadora, expected in cases.items():
            with self.subTest(transportadora=transportadora):
                self.assertEqual(_detect_branch(transportadora, ''), expected)

    def test_filial_emissao_field_takes_priority(self):
        from apps.financeiro.billing_import_service import parse_billing_xml

        parsed = parse_billing_xml(
            self._row_xml('05 - TRANSCAMILA - RONDONOPOLIS', filial_emissao='10')
        )
        self.assertEqual(parsed['10/06/2026']['Armazém']['count'], 1)

    def test_parse_billing_html_table_uses_valor_frete_and_transportadora(self):
        from apps.financeiro.billing_import_service import import_billing_file

        html = (
            '<html><table><tr>'
            '<th>Transportadora</th><th>Valor Frete</th><th>Emissão CT-e</th><th>Coleta</th><th>Usuário cadastro</th>'
            '</tr><tr>'
            '<td>09 - TRANSCAMILA - BARUERI</td><td>1.000,50</td><td>10/06/2026</td><td>10/06/2026</td><td>OPERADOR</td>'
            '</tr><tr>'
            '<td>10 - TRANSCAMILA - RONDONOPOLIS</td><td>500,00</td><td></td><td>10/06/2026</td><td>OPERADOR</td>'
            '</tr></table></html>'
        )
        parsed = import_billing_file(html.encode('latin-1'), 'relatorio.xls')
        self.assertTrue(parsed['success'])
        self.assertEqual(parsed['totalNotes'], 2)
        self.assertEqual(parsed['totalValue'], 1500.5)
        from apps.financeiro.models import BillingRecord
        barueri = BillingRecord.objects.get(reference_date=date(2026, 6, 10), branch='Barueri')
        armazem = BillingRecord.objects.get(reference_date=date(2026, 6, 10), branch='Armazém')
        self.assertEqual(barueri.value, Decimal('1000.50'))
        self.assertEqual(armazem.value, Decimal('500.00'))

    def test_coleta_date_takes_priority_over_emissao_cte(self):
        from apps.financeiro.billing_import_service import parse_billing_file

        html = (
            '<html><table><tr>'
            '<th>Transportadora</th><th>Valor Frete</th><th>Coleta</th><th>Emissão CT-e</th>'
            '</tr><tr>'
            '<td>09 - TRANSCAMILA - BARUERI</td><td>100,00</td><td>07/07/2026</td><td>08/07/2026</td>'
            '</tr></table></html>'
        )
        parsed = parse_billing_file(html.encode('latin-1'), 'relatorio.xls')
        self.assertEqual(parsed['07/07/2026']['Barueri']['value'], Decimal('100'))
        self.assertNotIn('08/07/2026', parsed)

    def test_import_teste_xls_matches_erp_coleta_totals_for_08_07(self):
        from pathlib import Path

        from apps.financeiro.billing_import_service import parse_billing_file

        sample = Path(r'S:\teste.xls')
        if not sample.exists():
            self.skipTest('S:\\teste.xls ausente neste ambiente')

        parsed = parse_billing_file(sample.read_bytes(), 'teste.xls')
        day = parsed['08/07/2026']
        total_value = sum(day[branch]['value'] for branch in day)
        total_notes = sum(day[branch]['count'] for branch in day)
        self.assertEqual(total_notes, 25)
        self.assertEqual(total_value, Decimal('215021.34'))
        self.assertEqual(day['Paranaguá']['value'], Decimal('40831.12'))
        self.assertEqual(day['Barueri']['count'], 22)


class CeleryImportTaskTests(TestCase):
    """As tasks são importadas em views.py e devem executar (via broker ou
    síncronas em .apply()) sem NameError, além de registrar auditoria."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='celery_user',
            password='cel123',
            role_id='2',
            environments=['Financeiro'],
            filiais={'Financeiro': ['Ibiporã (Matriz)']},
        )

    def test_import_report_task_runs_and_records_audit(self):
        from io import BytesIO

        import openpyxl

        from apps.financeiro.async_imports import save_upload
        from apps.financeiro.tasks import import_report_task

        batch = ReportBatch.objects.create(label='##CEL', reference_date=date(2026, 6, 20))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            'Filial Orig.', 'No. Titulo', 'Tipo', 'Fornecedor', 'Nome Fornece',
            'DT Emissao', 'Vencimento', 'Vencto Real', 'Vlr.Titulo', 'Historico', 'Saldo',
        ])
        ws.append(['01', 'T900', 'NF', 'F900', 'Fornecedor Celery', '01/06/2026', '20/06/2026', '20/06/2026', 100, '', 100])
        buf = BytesIO()
        wb.save(buf)
        temp_path = save_upload(buf.getvalue(), 'celery_pagar.xlsx')

        before = AuditLog.objects.count()
        result = import_report_task.apply(
            args=(str(batch.pk), 'pagar', str(temp_path), 'celery_pagar.xlsx', self.user.pk),
        ).get()

        self.assertTrue(result['success'])
        self.assertEqual(AuditLog.objects.count(), before + 1)
        self.assertTrue(AuditLog.objects.filter(action='importacao.relatorio').exists())
        batch.refresh_from_db()
        self.assertEqual(batch.updated_by_id, self.user.pk)
        self.assertFalse(temp_path.exists())

    def test_import_billing_xml_task_runs_and_records_audit(self):
        from apps.financeiro.async_imports import save_upload
        from apps.financeiro.tasks import import_billing_xml_task

        xml_bytes = self._row_xml('01 - TRANSCAMILA - IBIPORA').encode('utf-8')
        temp_path = save_upload(xml_bytes, 'celery_billing.xml')

        before = AuditLog.objects.count()
        result = import_billing_xml_task.apply(
            args=(str(temp_path), 'celery_billing.xml', self.user.pk),
        ).get()

        self.assertTrue(result['success'])
        self.assertEqual(AuditLog.objects.count(), before + 1)
        self.assertTrue(AuditLog.objects.filter(action='financeiro.faturamento.importado').exists())
        self.assertFalse(temp_path.exists())

    @staticmethod
    def _row_xml(transportadora: str) -> str:
        return (
            '<Row>'
            '<Emissão_x0020_CT_x002d_e>10/06/2026</Emissão_x0020_CT_x002d_e>'
            f'<Transportadora>{transportadora}</Transportadora>'
            '<Usuário_x0020_cadastro>OPERADOR</Usuário_x0020_cadastro>'
            '<Valor_x0020_Frete>1.000,00</Valor_x0020_Frete>'
            '</Row>'
        )


class CalendarioFinanceiroTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='fin_calendario',
            password='fin123',
            role_id='2',
            environments=['Financeiro'],
            filiais={'Financeiro': ['Ibiporã (Matriz)', 'Rondonópolis']},
        )
        self.batch = ReportBatch.objects.create(
            label='##C01',
            reference_date=date(2026, 7, 1),
            is_active=True,
            imported_pagar=True,
            imported_receber=True,
        )
        # Dois títulos do mesmo fornecedor no mesmo dia → agrupados em 1 evento
        for titulo in ('P0001', 'P0002'):
            PagarTitulo.objects.create(
                batch=self.batch,
                filial='01',
                cod_forn='F001',
                fornecedor='Fornecedor Alpha',
                titulo=titulo,
                tipo='NF',
                emissao='01/07/2026',
                vencimento='15/07/2026',
                vencimento_real='15/07/2026',
                valor=Decimal('100.00'),
                saldo=Decimal('100.00'),
                historico='',
            )
        ReceberTitulo.objects.create(
            batch=self.batch,
            filial='01',
            cod_cliente='C001',
            cliente='Cliente Beta',
            titulo='R0001',
            natureza='DUP',
            emissao='01/07/2026',
            vencimento='20/07/2026',
            vencimento_real='20/07/2026',
            valor=Decimal('250.00'),
            saldo=Decimal('250.00'),
            historico='',
        )

    def test_sistema_agrupa_por_fornecedor_e_dia(self):
        response = self.client.get(
            '/api/financeiro/calendario/sistema/?start=2026-07-01&end=2026-07-31',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['batchLabel'], '##C01')
        events = response.data['events']

        pagar_dia = events['2026-07-15']
        self.assertEqual(len(pagar_dia), 1)
        self.assertEqual(pagar_dia[0]['type'], 'pagar')
        self.assertEqual(pagar_dia[0]['count'], 2)
        self.assertEqual(pagar_dia[0]['amount'], 200.0)
        self.assertEqual(len(pagar_dia[0]['titulos']), 2)

        receber_dia = events['2026-07-20']
        self.assertEqual(receber_dia[0]['type'], 'receber')
        self.assertEqual(receber_dia[0]['amount'], 250.0)

    def test_sistema_fora_do_intervalo_fica_vazio(self):
        response = self.client.get(
            '/api/financeiro/calendario/sistema/?start=2026-08-01&end=2026-08-31',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['events'], {})

    def test_sistema_exige_start_e_end(self):
        response = self.client.get(
            '/api/financeiro/calendario/sistema/',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(response.status_code, 400)

    def test_evento_pessoal_crud_escopo_por_usuario(self):
        create = self.client.post(
            '/api/financeiro/calendario/eventos/',
            {'date': '2026-07-10', 'title': 'Reunião banco', 'description': 'Levar contratos', 'color': 'verde'},
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(create.status_code, 201)
        evento_id = create.data['id']

        listagem = self.client.get(
            '/api/financeiro/calendario/eventos/?start=2026-07-01&end=2026-07-31',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(listagem.status_code, 200)
        results = listagem.data if isinstance(listagem.data, list) else listagem.data['results']
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['title'], 'Reunião banco')

        # Outro usuário não vê o evento
        outro = User.objects.create_user(
            username='fin_outro',
            password='fin123',
            role_id='2',
            environments=['Financeiro'],
            filiais={'Financeiro': ['Ibiporã (Matriz)']},
        )
        listagem_outro = self.client.get(
            '/api/financeiro/calendario/eventos/?start=2026-07-01&end=2026-07-31',
            **auth_headers(outro, 'Financeiro'),
        )
        results_outro = listagem_outro.data if isinstance(listagem_outro.data, list) else listagem_outro.data['results']
        self.assertEqual(len(results_outro), 0)

        update = self.client.patch(
            f'/api/financeiro/calendario/eventos/{evento_id}/',
            {'date': '2026-07-11'},
            format='json',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(update.status_code, 200)
        self.assertEqual(update.data['date'], '2026-07-11')

        delete = self.client.delete(
            f'/api/financeiro/calendario/eventos/{evento_id}/',
            **auth_headers(self.user, 'Financeiro'),
        )
        self.assertEqual(delete.status_code, 204)
