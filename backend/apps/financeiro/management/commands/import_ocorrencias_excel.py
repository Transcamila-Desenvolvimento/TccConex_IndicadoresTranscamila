"""Importa OPs recebidas e GNRE-ICMS a partir da planilha de Indicadores Financeiro."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.financeiro.models import GnreIcmsOcorrencia, OpsRecebidaOcorrencia

def _ascii_key(value: str) -> str:
    table = str.maketrans({
        'Á': 'A', 'À': 'A', 'Â': 'A', 'Ã': 'A',
        'É': 'E', 'Ê': 'E',
        'Í': 'I',
        'Ó': 'O', 'Ô': 'O', 'Õ': 'O',
        'Ú': 'U',
        'Ç': 'C',
    })
    return value.upper().translate(table)


FILIAL_ALIASES = {
    'IBIPORA': 'Ibiporã',
    'RONDONOPOLIS': 'Rondonópolis',
    'PARANAGUA': 'Paranaguá',
    'BARUERI': 'Barueri',
    'PORTO ALEGRE': 'Porto Alegre',
    'ARMAZEM': 'Armazém',
}


def _norm_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _norm_filial(value) -> str | None:
    raw = _norm_text(value)
    if not raw:
        return None
    return FILIAL_ALIASES.get(_ascii_key(raw))


def _parse_bool_sim_nao(value) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = _norm_text(value).upper()
    text = (
        text.replace('Ã', 'A')
        .replace('Á', 'A')
        .replace('À', 'A')
        .replace('Â', 'A')
    )
    if text in ('SIM', 'S', '1', 'TRUE', 'YES'):
        return True
    if text in ('NAO', 'NÃO', 'N', '0', 'FALSE', 'NO'):
        return False
    # encoding-broken "NÃO" often becomes "N?O" or "NO"
    if text.startswith('N') and 'O' in text and 'SIM' not in text:
        return False
    return None


def _parse_date(value) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _norm_text(value)
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_periodo(value) -> str:
    if value is None or value == '':
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m')
    if isinstance(value, date):
        return value.strftime('%Y-%m')
    text = _norm_text(value).lower()
    months = {
        'janeiro': '01', 'fevereiro': '02', 'marco': '03', 'março': '03',
        'abril': '04', 'maio': '05', 'junho': '06', 'julho': '07',
        'agosto': '08', 'setembro': '09', 'outubro': '10',
        'novembro': '11', 'dezembro': '12',
    }
    for name, num in months.items():
        if name in text:
            digits = ''.join(ch for ch in text if ch.isdigit())
            year = digits[-4:] if len(digits) >= 4 else ''
            if year:
                return f'{year}-{num}'
    if len(text) >= 7 and text[4] == '-':
        return text[:7]
    return _norm_text(value)[:20]


def _parse_doc(value) -> str:
    if value is None or value == '':
        return ''
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip('0').rstrip('.')
    if isinstance(value, int):
        return str(value)
    return _norm_text(value)


def _parse_money(value) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = _norm_text(value).replace('R$', '').replace(' ', '')
    if ',' in text and '.' in text:
        text = text.replace('.', '').replace(',', '.')
    elif ',' in text:
        text = text.replace(',', '.')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = 'Importa OPs recebidas e GNRE-ICMS da planilha Indicadores FINANCEIRO.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            'path',
            nargs='?',
            default=r'c:\Users\Adm-ibi\Downloads\Indicadores - 01.08.2025 - FINANCEIRO.xlsx',
            help='Caminho do arquivo .xlsx',
        )
        parser.add_argument(
            '--replace',
            action='store_true',
            help='Apaga registros existentes de OPS e GNRE antes de importar',
        )

    def handle(self, *args, **options):
        path = Path(options['path'])
        if not path.exists():
            raise CommandError(f'Arquivo não encontrado: {path}')

        self.stdout.write(f'Lendo {path} ...')
        wb = load_workbook(path, data_only=True, read_only=True)

        ops_sheet = None
        gnre_sheet = None
        for name in wb.sheetnames:
            key = name.strip().upper().replace('  ', ' ')
            if 'OPS' in key and 'RECEBID' in key:
                ops_sheet = name
            if 'GNRE' in key:
                gnre_sheet = name

        if not ops_sheet or not gnre_sheet:
            raise CommandError(
                f'Abas OPS/GNRE não encontradas. Abas: {wb.sheetnames}'
            )

        ops_rows = self._read_ops(wb[ops_sheet])
        gnre_rows = self._read_gnre(wb[gnre_sheet])
        wb.close()

        with transaction.atomic():
            if options['replace']:
                deleted_ops, _ = OpsRecebidaOcorrencia.objects.all().delete()
                deleted_gnre, _ = GnreIcmsOcorrencia.objects.all().delete()
                self.stdout.write(f'Removidos: {deleted_ops} OPS, {deleted_gnre} GNRE')

            OpsRecebidaOcorrencia.objects.bulk_create(ops_rows, batch_size=500)
            GnreIcmsOcorrencia.objects.bulk_create(gnre_rows, batch_size=500)

        self.stdout.write(self.style.SUCCESS(
            f'Importação concluída: {len(ops_rows)} OPS, {len(gnre_rows)} GNRE.'
        ))

    def _read_ops(self, ws) -> list[OpsRecebidaOcorrencia]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        created: list[OpsRecebidaOcorrencia] = []
        skipped = 0
        for raw in rows[1:]:
            if not raw or not any(raw):
                continue
            filial = _norm_filial(raw[0] if len(raw) > 0 else None)
            contrato = _parse_doc(raw[1] if len(raw) > 1 else None)
            data_pag = _parse_date(raw[2] if len(raw) > 2 else None)
            mdfe = _parse_bool_sim_nao(raw[3] if len(raw) > 3 else None)
            if not filial or not contrato or not data_pag or mdfe is None:
                skipped += 1
                continue
            created.append(OpsRecebidaOcorrencia(
                filial=filial,
                contrato=contrato[:50],
                data_pagamento=data_pag,
                mdfe_encerrado=mdfe,
            ))
        if skipped:
            self.stdout.write(self.style.WARNING(f'OPS: {skipped} linha(s) ignorada(s).'))
        return created

    def _read_gnre(self, ws) -> list[GnreIcmsOcorrencia]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        created: list[GnreIcmsOcorrencia] = []
        skipped = 0
        for raw in rows[1:]:
            if not raw or not any(raw):
                continue
            filial = _norm_filial(raw[0] if len(raw) > 0 else None)
            cte = _parse_doc(raw[1] if len(raw) > 1 else None)
            valor = _parse_money(raw[2] if len(raw) > 2 else None)
            periodo = _parse_periodo(raw[3] if len(raw) > 3 else None)
            data_pag = _parse_date(raw[4] if len(raw) > 4 else None)
            validada = _parse_bool_sim_nao(raw[5] if len(raw) > 5 else None)
            if not filial or not cte or valor is None or not periodo or not data_pag or validada is None:
                skipped += 1
                continue
            created.append(GnreIcmsOcorrencia(
                filial=filial,
                cte=cte[:50],
                valor_guia=valor,
                periodo_referencia=periodo[:20],
                data_pagamento=data_pag,
                validada=validada,
            ))
        if skipped:
            self.stdout.write(self.style.WARNING(f'GNRE: {skipped} linha(s) ignorada(s).'))
        return created
