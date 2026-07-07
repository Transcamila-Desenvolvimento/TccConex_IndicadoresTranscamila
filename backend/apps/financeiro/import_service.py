import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from django.utils import timezone

from .models import AgingTitulo, PagarTitulo, ReceberTitulo, ReportBatch
from .pr_ignore_service import mark_ignored_prs_on_import

REPORT_KEYWORDS = {
    'pagar': ['filial orig', 'fornecedor', 'no. titulo'],
    'receber': ['filial orig', 'nome cliente', 'cliente'],
    'aging': ['cod.cliente', 'docto', 'total'],
}


def _normalize_header(value) -> str:
    text = str(value or '')
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if ord(c) < 128 or c.isalnum() or c.isspace())
    return re.sub(r'\s+', ' ', text.lower().strip())


def _format_date(value) -> str:
    if value is None or value == '':
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    text = str(value).strip()
    if re.match(r'^\d{2}/\d{2}/\d{4}$', text):
        return text
    return text


def _parse_number(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).replace('.', '').replace(',', '.').strip()
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _normalize_filial(value) -> str:
    """Preserva o código da filial como no Excel (ex.: 01, 03, 05)."""
    raw = str(value or '').strip()
    if not raw:
        return ''
    if raw.isdigit():
        return raw.zfill(2)
    return raw


def _normalize_origem(value) -> str:
    """Preserva o código de origem como no Aging (ex.: 1, 5, 9, 11)."""
    raw = str(value or '').strip()
    if not raw:
        return ''
    if raw.replace('.', '', 1).isdigit():
        return str(int(float(raw)))
    return raw


def _read_rows(file_bytes: bytes, report_type: str) -> tuple[list, int]:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    best_sheet = workbook[workbook.sheetnames[0]]
    best_score = 0
    keywords = REPORT_KEYWORDS[report_type]

    for name in workbook.sheetnames:
        sheet = workbook[name]
        sample = []
        for i, row in enumerate(sheet.iter_rows(max_row=15, values_only=True)):
            if i >= 15:
                break
            sample.extend(row or [])
        flat = ' '.join(_normalize_header(c) for c in sample)
        score = sum(1 for k in keywords if k in flat)
        if score > best_score:
            best_score = score
            best_sheet = sheet

    return list(best_sheet.iter_rows(values_only=True)), best_score


def _find_header_row(rows, report_type: str) -> int:
    required = REPORT_KEYWORDS[report_type]
    for i, row in enumerate(rows[:15]):
        normalized = [_normalize_header(c) for c in (row or [])]
        matches = sum(1 for key in required if any(key in cell for cell in normalized))
        if matches >= len(required):
            return i
    return 2 if report_type == 'pagar' else 1 if report_type == 'aging' else 0


def _map_headers(header_row):
    mapping = {}
    for idx, cell in enumerate(header_row or []):
        key = _normalize_header(cell)
        if key:
            mapping[key] = idx
    return mapping


def _col_index(headers, *candidates):
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
        for key, idx in headers.items():
            if candidate in key or key in candidate:
                return idx
    return -1


def _is_header_repeat(first_cell, report_type: str) -> bool:
    cell = _normalize_header(first_cell)
    if report_type == 'aging':
        return cell == 'cod.cliente'
    return cell in ('filial orig', 'filial orig.')


def import_report_file(batch: ReportBatch, report_type: str, file_bytes: bytes, file_name: str) -> dict:
    if not file_name.lower().endswith(('.xlsx', '.xls')):
        return {'success': False, 'rowCount': 0, 'skippedRows': 0, 'issues': [{'severity': 'error', 'message': 'Formato inválido. Use .xlsx.'}], 'data': []}

    rows, sheet_score = _read_rows(file_bytes, report_type)
    if not rows:
        return {'success': False, 'rowCount': 0, 'skippedRows': 0, 'issues': [{'severity': 'error', 'message': 'Arquivo vazio.'}], 'data': []}

    required_keywords = REPORT_KEYWORDS[report_type]
    if sheet_score < len(required_keywords):
        labels = {'pagar': 'Contas a Pagar', 'receber': 'Contas a Receber', 'aging': 'Aging Luft'}
        return {
            'success': False,
            'rowCount': 0,
            'skippedRows': 0,
            'issues': [{
                'severity': 'error',
                'message': (
                    f'O arquivo não corresponde ao relatório de {labels[report_type]}. '
                    'Verifique se selecionou o tipo correto na importação.'
                ),
            }],
            'data': [],
        }

    header_idx = _find_header_row(rows, report_type)
    headers = _map_headers(rows[header_idx])
    issues = []
    parsed = []
    skipped = 0

    if report_type == 'pagar':
        idx = {
            'filial': _col_index(headers, 'filial orig', 'filial orig.'),
            'titulo': _col_index(headers, 'no. titulo'),
            'tipo': _col_index(headers, 'tipo'),
            'cod_forn': _col_index(headers, 'fornecedor'),
            'fornecedor': _col_index(headers, 'nome fornece', 'nome fornecedor'),
            'emissao': _col_index(headers, 'dt emissao'),
            'vencimento': _col_index(headers, 'vencimento'),
            'vencimento_real': _col_index(headers, 'vencto real'),
            'valor': _col_index(headers, 'vlr.titulo', 'vlr titulo'),
            'historico': _col_index(headers, 'historico'),
            'saldo': _col_index(headers, 'saldo'),
        }
        if idx['titulo'] < 0 or idx['valor'] < 0:
            return {'success': False, 'rowCount': 0, 'skippedRows': 0, 'issues': [{'severity': 'error', 'message': 'Colunas obrigatórias ausentes no arquivo de Contas a Pagar.'}], 'data': []}

        for r, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            row = row or ()
            if not any(c is not None and str(c).strip() for c in row):
                continue
            if _is_header_repeat(row[0] if row else None, report_type):
                skipped += 1
                continue
            titulo = str(row[idx['titulo']] or '').strip()
            valor = _parse_number(row[idx['valor']] if idx['valor'] >= 0 else None)
            if not titulo:
                skipped += 1
                issues.append({'row': r, 'severity': 'warning', 'message': 'Linha ignorada: título vazio.'})
                continue
            if valor is None:
                skipped += 1
                issues.append({'row': r, 'severity': 'warning', 'message': 'Valor inválido — linha ignorada.'})
                continue
            saldo = _parse_number(row[idx['saldo']] if idx['saldo'] >= 0 else None) or valor
            tipo = str(row[idx['tipo']] or '').strip() if idx['tipo'] >= 0 else ''
            filial = _normalize_filial(row[idx['filial']] if idx['filial'] >= 0 else '')
            if not filial and tipo == 'PA':
                filial = 'PA'
            elif not filial:
                issues.append({'row': r, 'severity': 'warning', 'message': 'Filial vazia — linha importada sem código de filial.'})
            parsed.append(PagarTitulo(
                batch=batch,
                filial=filial,
                cod_forn=str(row[idx['cod_forn']] or '').strip() if idx['cod_forn'] >= 0 else '',
                fornecedor=str(row[idx['fornecedor']] or '').strip() if idx['fornecedor'] >= 0 else '',
                titulo=titulo,
                tipo=tipo,
                emissao=_format_date(row[idx['emissao']] if idx['emissao'] >= 0 else ''),
                vencimento=_format_date(row[idx['vencimento']] if idx['vencimento'] >= 0 else ''),
                vencimento_real=_format_date(row[idx['vencimento_real']] if idx['vencimento_real'] >= 0 else ''),
                valor=valor,
                saldo=saldo,
                historico=str(row[idx['historico']] or '').strip() if idx['historico'] >= 0 else '',
            ))

    elif report_type == 'receber':
        idx = {
            'filial': _col_index(headers, 'filial orig'),
            'titulo': _col_index(headers, 'no. titulo'),
            'natureza': _col_index(headers, 'natureza'),
            'cod_cliente': _col_index(headers, 'cliente'),
            'cliente': _col_index(headers, 'nome cliente'),
            'emissao': _col_index(headers, 'data de emissao'),
            'vencimento': _col_index(headers, 'vencto orig'),
            'vencimento_real': _col_index(headers, 'vencto real'),
            'valor': _col_index(headers, 'vlr.titulo', 'vlr titulo'),
            'historico': _col_index(headers, 'historico'),
            'saldo': _col_index(headers, 'saldo'),
            'saldo_atual': _col_index(headers, 'titulos a vencer valor atual'),
            'saldo_vencido': _col_index(headers, 'tit vencidos valor corrigido'),
        }
        if idx['titulo'] < 0 or idx['valor'] < 0:
            return {'success': False, 'rowCount': 0, 'skippedRows': 0, 'issues': [{'severity': 'error', 'message': 'Colunas obrigatórias ausentes no arquivo de Contas a Receber.'}], 'data': []}

        for r, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            row = row or ()
            if not any(c is not None and str(c).strip() for c in row):
                continue
            if _is_header_repeat(row[0] if row else None, report_type):
                skipped += 1
                continue
            titulo = str(row[idx['titulo']] or '').strip()
            valor = _parse_number(row[idx['valor']] if idx['valor'] >= 0 else None)
            if not titulo:
                skipped += 1
                continue
            if valor is None:
                skipped += 1
                issues.append({'row': r, 'severity': 'warning', 'message': 'Valor inválido — linha ignorada.'})
                continue
            saldo_atual = _parse_number(row[idx['saldo_atual']] if idx['saldo_atual'] >= 0 else None) or Decimal('0')
            saldo_vencido = _parse_number(row[idx['saldo_vencido']] if idx['saldo_vencido'] >= 0 else None) or Decimal('0')
            saldo_col = _parse_number(row[idx['saldo']] if idx['saldo'] >= 0 else None)
            if saldo_col is not None:
                saldo = saldo_col
            else:
                # Não somar atual+vencido — em títulos vencidos ambos repetem o mesmo saldo.
                saldo = saldo_vencido or saldo_atual or valor
            filial = _normalize_filial(row[idx['filial']] if idx['filial'] >= 0 else '')
            if not filial:
                issues.append({'row': r, 'severity': 'warning', 'message': 'Filial vazia — linha importada sem código de filial.'})
            parsed.append(ReceberTitulo(
                batch=batch,
                filial=filial,
                cod_cliente=str(row[idx['cod_cliente']] or '').strip() if idx['cod_cliente'] >= 0 else '',
                cliente=str(row[idx['cliente']] or '').strip() if idx['cliente'] >= 0 else '',
                titulo=titulo,
                natureza=str(row[idx['natureza']] or '').strip() if idx['natureza'] >= 0 else '',
                emissao=_format_date(row[idx['emissao']] if idx['emissao'] >= 0 else ''),
                vencimento=_format_date(row[idx['vencimento']] if idx['vencimento'] >= 0 else ''),
                vencimento_real=_format_date(row[idx['vencimento_real']] if idx['vencimento_real'] >= 0 else ''),
                valor=valor,
                saldo=saldo,
                historico=str(row[idx['historico']] or '').strip() if idx['historico'] >= 0 else '',
            ))

    else:
        idx = {
            'cod_cliente': _col_index(headers, 'cod.cliente'),
            'loja': _col_index(headers, 'loja'),
            'cliente': _col_index(headers, 'cliente'),
            'origem': _col_index(headers, 'origem'),
            'regiao': _col_index(headers, 'regiao'),
            'docto': _col_index(headers, 'docto'),
            'serie': _col_index(headers, 'serie/prefixo', 'serie'),
            'tipo': _col_index(headers, 'tipo'),
            'emissao': _col_index(headers, 'emissao'),
            'vencimento': _col_index(headers, 'vencto.', 'vencto'),
            'total': _col_index(headers, 'total'),
        }
        if idx['docto'] < 0 or idx['total'] < 0:
            return {'success': False, 'rowCount': 0, 'skippedRows': 0, 'issues': [{'severity': 'error', 'message': 'Colunas obrigatórias ausentes no arquivo Aging Luft.'}], 'data': []}

        for r, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            row = row or ()
            if not any(c is not None and str(c).strip() for c in row):
                continue
            if _is_header_repeat(row[0] if row else None, report_type):
                skipped += 1
                continue
            docto = str(row[idx['docto']] or '').strip()
            total = _parse_number(row[idx['total']] if idx['total'] >= 0 else None)
            if not docto:
                skipped += 1
                continue
            if total is None:
                skipped += 1
                issues.append({'row': r, 'severity': 'warning', 'message': 'Total inválido — linha ignorada.'})
                continue
            parsed.append(AgingTitulo(
                batch=batch,
                origem=_normalize_origem(row[idx['origem']] if idx['origem'] >= 0 else ''),
                cod_cliente=str(row[idx['cod_cliente']] or '').strip() if idx['cod_cliente'] >= 0 else '',
                cliente=str(row[idx['cliente']] or '').strip() if idx['cliente'] >= 0 else '',
                loja=str(row[idx['loja']] or '').strip() if idx['loja'] >= 0 else '',
                docto=docto,
                serie=str(row[idx['serie']] or '').strip() if idx['serie'] >= 0 else '',
                tipo=str(row[idx['tipo']] or '').strip() if idx['tipo'] >= 0 else '',
                emissao=_format_date(row[idx['emissao']] if idx['emissao'] >= 0 else ''),
                vencimento=_format_date(row[idx['vencimento']] if idx['vencimento'] >= 0 else ''),
                regiao=str(row[idx['regiao']] or '').strip() if idx['regiao'] >= 0 else '',
                total=total,
            ))

    if not parsed:
        issues.append({'severity': 'error', 'message': 'Nenhum registro válido encontrado.'})
        return {'success': False, 'rowCount': 0, 'skippedRows': skipped, 'issues': issues, 'data': []}

    if report_type == 'pagar':
        mark_ignored_prs_on_import(batch, parsed)
        PagarTitulo.objects.filter(batch=batch).delete()
        PagarTitulo.objects.bulk_create(parsed)
        batch.imported_pagar = True
    elif report_type == 'receber':
        ReceberTitulo.objects.filter(batch=batch).delete()
        ReceberTitulo.objects.bulk_create(parsed)
        batch.imported_receber = True
    else:
        AgingTitulo.objects.filter(batch=batch).delete()
        AgingTitulo.objects.bulk_create(parsed)
        batch.imported_aging = True

    batch.save(update_fields=['imported_pagar', 'imported_receber', 'imported_aging'])

    return {
        'success': True,
        'rowCount': len(parsed),
        'skippedRows': skipped,
        'issues': issues,
        'data': [],
    }
