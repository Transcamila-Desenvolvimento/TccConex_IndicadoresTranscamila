import re
import unicodedata
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction

from .models import (
    CustoAbastecimentoLinha,
    CustoFrotaLote,
    CustoManutencaoLinha,
    VeiculoFrota,
    format_placa,
    normalize_placa,
    placa_valida,
)

MAX_CUSTO_LOTES = 24

MESES_PT = {
    'janeiro': 1, 'fevereiro': 2, 'marco': 3, 'março': 3, 'abril': 4,
    'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8, 'setembro': 9,
    'outubro': 10, 'novembro': 11, 'dezembro': 12,
}

MESES_LABEL = {
    1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
    7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez',
}

PERIOD_RE = re.compile(
    r'(\d{2}/\d{2}/\d{4})\s+at[eéê]?\s+(\d{2}/\d{2}/\d{4})',
    re.IGNORECASE,
)
DATE_BR_RE = re.compile(r'^(\d{1,2})/(\d{1,2})/(\d{4})$')

SKIP_MANUT_PREFIXES = (
    'quebra', 'filtro', 'sub-total', 'subtotal', 'total geral', 'analise',
    'análise', 'material', 'valor total', 'servicos', 'serviços',
)


def _strip_accents(value: str) -> str:
    text = unicodedata.normalize('NFD', value or '')
    return ''.join(c for c in text if unicodedata.category(c) != 'Mn')


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _parse_br_date(value) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value).split(' ')[0]
    match = DATE_BR_RE.match(text)
    if not match:
        return None
    day, month, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_number(value) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = _cell_str(value)
    text = text.replace('R$', '').replace(' ', '').strip()
    if not text or text in ('-', 'undefined'):
        return None
    if re.match(r'^-?\d+,\d+$', text):
        text = text.replace('.', '').replace(',', '.')
    elif text.count('.') == 1 and text.count(',') == 0:
        pass
    else:
        text = text.replace('.', '').replace(',', '.')
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_int(value) -> int | None:
    """Inteiro no padrão BR: 1.238.773 ou 601.690 são milhares, não decimal."""
    if value is None or value == '':
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value >= 0 else None
    if isinstance(value, float):
        if value < 0:
            return None
        return int(round(value))
    if isinstance(value, Decimal):
        if value < 0:
            return None
        return int(value)
    text = _cell_str(value).replace('R$', '').replace(' ', '').strip()
    if not text or text in ('-', 'undefined'):
        return None
    digits = re.sub(r'\D', '', text)
    if not digits:
        return None
    return int(digits)


def _read_rows(file_bytes: bytes, file_name: str) -> list[list]:
    name = (file_name or '').lower()
    if name.endswith('.xls') and not name.endswith('.xlsx'):
        import xlrd

        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        return [
            [sheet.cell_value(row, col) for col in range(sheet.ncols)]
            for row in range(sheet.nrows)
        ]

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def detect_report_kind(rows: list[list], file_name: str) -> str | None:
    name = _strip_accents((file_name or '').lower())
    blob = _strip_accents(
        ' '.join(_cell_str(cell) for row in rows[:50] for cell in (row or [])[:18]).lower()
    )
    if 'abastec' in name or 'litragem' in blob or 'transacao' in blob:
        return 'abastecimento'
    if 'manutenc' in name or 'quebra' in blob or 'custo de manut' in blob:
        return 'manutencao'
    return None


def _period_from_filename(file_name: str) -> tuple[date, date] | None:
    name = _strip_accents((file_name or '').lower())
    year_match = re.search(r'(20\d{2})', name)
    month = None
    for label, number in MESES_PT.items():
        if _strip_accents(label) in name:
            month = number
            break
    if not month:
        return None
    year = int(year_match.group(1)) if year_match else date.today().year
    last_day = monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _period_from_text(rows: list[list]) -> tuple[date, date] | None:
    for row in rows[:20]:
        text = ' '.join(_cell_str(cell) for cell in (row or []))
        match = PERIOD_RE.search(_strip_accents(text))
        if match:
            start = _parse_br_date(match.group(1))
            end = _parse_br_date(match.group(2))
            if start and end:
                return start, end
    return None


def _period_from_dates(dates: list[date]) -> tuple[date, date] | None:
    valid = [item for item in dates if item]
    if not valid:
        return None
    return min(valid), max(valid)


def _as_month_period(inicio: date, fim: date) -> tuple[date, date]:
    if inicio.year == fim.year and inicio.month == fim.month:
        last_day = monthrange(inicio.year, inicio.month)[1]
        return date(inicio.year, inicio.month, 1), date(inicio.year, inicio.month, last_day)
    return inicio, fim


def lote_label(inicio: date, fim: date) -> str:
    if inicio.day == 1 and fim == date(inicio.year, inicio.month, monthrange(inicio.year, inicio.month)[1]):
        return f'{MESES_LABEL[inicio.month]}/{inicio.year}'
    return f'{inicio.strftime("%d/%m")}–{fim.strftime("%d/%m/%Y")}'


def trim_old_lotes(max_lotes: int = MAX_CUSTO_LOTES) -> None:
    keep_ids = list(
        CustoFrotaLote.objects.order_by('-periodo_inicio', '-created_at')
        .values_list('pk', flat=True)[:max_lotes]
    )
    if keep_ids:
        CustoFrotaLote.objects.exclude(pk__in=keep_ids).delete()


def get_or_create_lote(inicio: date, fim: date, user) -> tuple[CustoFrotaLote, bool]:
    existing = CustoFrotaLote.objects.filter(periodo_inicio=inicio, periodo_fim=fim).first()
    if existing:
        return existing, False
    lote = CustoFrotaLote.objects.create(
        label=lote_label(inicio, fim),
        periodo_inicio=inicio,
        periodo_fim=fim,
        updated_by=user if getattr(user, 'is_authenticated', False) else None,
    )
    trim_old_lotes()
    return lote, True


def _veiculos_por_placa() -> dict[str, VeiculoFrota]:
    return {item.placa: item for item in VeiculoFrota.objects.all()}


def _parse_manutencao(rows: list[list]) -> tuple[list[dict], list[str]]:
    items: list[dict] = []
    placas_encontradas: list[str] = []
    current_placa = ''
    current_grupo = ''

    for row in rows:
        cells = list(row or [])
        first = _cell_str(cells[0] if cells else '')
        if not first:
            continue
        first_norm = _strip_accents(first).lower()
        if first_norm.startswith(SKIP_MANUT_PREFIXES) or first_norm in ('(%)',):
            continue

        placa = normalize_placa(first)
        rest_empty = all(not _cell_str(cell) for cell in cells[1:5])
        has_values = any(_parse_number(cell) is not None for cell in cells[5:11])
        if placa_valida(placa) and rest_empty and not has_values:
            current_placa = placa
            current_grupo = ''
            placas_encontradas.append(placa)
            continue

        if not current_placa:
            continue

        if not has_values:
            current_grupo = first
            continue

        material = _parse_number(cells[5] if len(cells) > 5 else None) or Decimal('0')
        servicos = _parse_number(cells[7] if len(cells) > 7 else None) or Decimal('0')
        total = _parse_number(cells[9] if len(cells) > 9 else None)
        if total is None:
            total = material + servicos
        items.append({
            'placa': current_placa,
            'grupo': current_grupo,
            'item': first,
            'valor_material': material,
            'valor_servicos': servicos,
            'valor_total': total,
        })
    return items, placas_encontradas


def _is_header_row(cells: list) -> bool:
    first = _strip_accents(_cell_str(cells[0] if cells else '')).lower()
    second = _strip_accents(_cell_str(cells[1] if len(cells) > 1 else '')).lower()
    return first == 'cliente' and second == 'placa'


def _is_tx_header(cells: list) -> bool:
    first = _strip_accents(_cell_str(cells[0] if cells else '')).lower()
    return first in ('transacao', 'transação')


def _parse_abastecimento(rows: list[list]) -> tuple[list[dict], list[str], list[date]]:
    items: list[dict] = []
    placas: list[str] = []
    dates: list[date] = []
    current_placa = ''

    for row in rows:
        cells = list(row or [])
        if not any(_cell_str(cell) for cell in cells):
            continue
        if _is_header_row(cells):
            continue
        if _is_tx_header(cells):
            continue

        joined = _strip_accents(' '.join(_cell_str(cell) for cell in cells)).lower()
        if 'totais do veiculo' in joined or 'totais do veículo' in joined:
            continue

        maybe_placa = normalize_placa(_cell_str(cells[1] if len(cells) > 1 else ''))
        first = _cell_str(cells[0] if cells else '')
        if placa_valida(maybe_placa) and _strip_accents(first).lower() != 'transacao':
            current_placa = maybe_placa
            placas.append(maybe_placa)
            continue

        if not current_placa:
            continue

        data = _parse_br_date(cells[1] if len(cells) > 1 else None)
        if data is None:
            continue
        dates.append(data)
        items.append({
            'placa': current_placa,
            'transacao': first,
            'data': data,
            'hora': _cell_str(cells[2] if len(cells) > 2 else ''),
            'estabelecimento': _cell_str(cells[3] if len(cells) > 3 else ''),
            'cidade': _cell_str(cells[4] if len(cells) > 4 else ''),
            'motorista': _cell_str(cells[5] if len(cells) > 5 else ''),
            'hodometro': _parse_int(cells[6] if len(cells) > 6 else None),
            'km_trecho': _parse_int(cells[14] if len(cells) > 14 else None),
            'litragem': _parse_number(cells[8] if len(cells) > 8 else None),
            'combustivel': _cell_str(cells[9] if len(cells) > 9 else ''),
            'valor_total': _parse_number(cells[12] if len(cells) > 12 else None) or Decimal('0'),
            'numero_nfe': _cell_str(cells[13] if len(cells) > 13 else ''),
        })
    return items, placas, dates


def import_custo_file(report_type: str, file_bytes: bytes, file_name: str, user) -> dict:
    issues: list[dict] = []
    try:
        rows = _read_rows(file_bytes, file_name)
    except Exception as exc:
        return {
            'success': False,
            'rowCount': 0,
            'skippedRows': 0,
            'issues': [{'severity': 'error', 'message': f'Não foi possível ler o arquivo: {exc}'}],
            'lote': None,
            'reusedLote': False,
            'periodoInicio': None,
            'periodoFim': None,
        }

    detected = detect_report_kind(rows, file_name)
    if detected and detected != report_type:
        labels = {'manutencao': 'manutenção', 'abastecimento': 'abastecimentos'}
        return {
            'success': False,
            'rowCount': 0,
            'skippedRows': 0,
            'issues': [{
                'severity': 'error',
                'message': (
                    f'O arquivo parece ser de {labels.get(detected, detected)}, '
                    f'não de {labels.get(report_type, report_type)}.'
                ),
            }],
            'lote': None,
            'reusedLote': False,
            'periodoInicio': None,
            'periodoFim': None,
        }

    if report_type == 'manutencao':
        parsed, placas_arquivo = _parse_manutencao(rows)
        period = _period_from_text(rows) or _period_from_filename(file_name)
    elif report_type == 'abastecimento':
        parsed, placas_arquivo, dates = _parse_abastecimento(rows)
        period = _period_from_text(rows) or _period_from_dates(dates) or _period_from_filename(file_name)
    else:
        return {
            'success': False,
            'rowCount': 0,
            'skippedRows': 0,
            'issues': [{'severity': 'error', 'message': 'Tipo de relatório inválido.'}],
            'lote': None,
            'reusedLote': False,
            'periodoInicio': None,
            'periodoFim': None,
        }

    if not period:
        return {
            'success': False,
            'rowCount': 0,
            'skippedRows': 0,
            'issues': [{
                'severity': 'error',
                'message': 'Não foi possível identificar o período da planilha. Verifique o filtro de datas ou o nome do arquivo.',
            }],
            'lote': None,
            'reusedLote': False,
            'periodoInicio': None,
            'periodoFim': None,
        }

    period = _as_month_period(period[0], period[1])

    if not parsed and not placas_arquivo:
        return {
            'success': False,
            'rowCount': 0,
            'skippedRows': 0,
            'issues': [{'severity': 'error', 'message': 'Nenhum lançamento encontrado na planilha.'}],
            'lote': None,
            'reusedLote': False,
            'periodoInicio': period[0].isoformat(),
            'periodoFim': period[1].isoformat(),
        }

    veiculos = _veiculos_por_placa()
    missing = sorted({placa for placa in placas_arquivo if placa not in veiculos})
    accepted = [item for item in parsed if item['placa'] in veiculos]
    skipped = len(parsed) - len(accepted)

    for placa in missing:
        issues.append({
            'severity': 'warning',
            'message': f'Veículo {format_placa(placa)} não está cadastrado na frota. Os lançamentos dessa placa foram ignorados.',
        })

    if not accepted:
        issues.insert(0, {
            'severity': 'error',
            'message': 'Nenhum lançamento importado: cadastre os veículos da planilha antes de importar.',
        })
        return {
            'success': False,
            'rowCount': 0,
            'skippedRows': skipped,
            'issues': issues,
            'lote': None,
            'reusedLote': False,
            'periodoInicio': period[0].isoformat(),
            'periodoFim': period[1].isoformat(),
        }

    with transaction.atomic():
        lote, created = get_or_create_lote(period[0], period[1], user)
        if report_type == 'manutencao':
            lote.manutencao_linhas.all().delete()
            CustoManutencaoLinha.objects.bulk_create([
                CustoManutencaoLinha(
                    lote=lote,
                    veiculo=veiculos[item['placa']],
                    placa=item['placa'],
                    grupo=item['grupo'],
                    item=item['item'],
                    valor_material=item['valor_material'],
                    valor_servicos=item['valor_servicos'],
                    valor_total=item['valor_total'],
                )
                for item in accepted
            ])
            lote.imported_manutencao = True
        else:
            lote.abastecimento_linhas.all().delete()
            CustoAbastecimentoLinha.objects.bulk_create([
                CustoAbastecimentoLinha(
                    lote=lote,
                    veiculo=veiculos[item['placa']],
                    placa=item['placa'],
                    transacao=item['transacao'],
                    data=item['data'],
                    hora=item['hora'],
                    estabelecimento=item['estabelecimento'],
                    cidade=item['cidade'],
                    motorista=item['motorista'],
                    hodometro=item['hodometro'],
                    km_trecho=item.get('km_trecho'),
                    litragem=item['litragem'],
                    combustivel=item['combustivel'],
                    valor_total=item['valor_total'],
                    numero_nfe=item['numero_nfe'],
                )
                for item in accepted
            ])
            lote.imported_abastecimento = True
        lote.updated_by = user if getattr(user, 'is_authenticated', False) else lote.updated_by
        lote.save()

    issues.append({
        'severity': 'warning',
        'message': (
            f'Lote {lote.label} ({period[0].strftime("%d/%m/%Y")} a {period[1].strftime("%d/%m/%Y")})'
            + (
                ' reutilizado: este relatório do período foi substituído; os demais foram mantidos.'
                if not created
                else ' criado a partir do período detectado na planilha.'
            )
        ),
    })

    return {
        'success': True,
        'rowCount': len(accepted),
        'skippedRows': skipped,
        'issues': issues,
        'loteId': str(lote.pk),
        'loteLabel': lote.label,
        'reusedLote': not created,
        'periodoInicio': period[0].isoformat(),
        'periodoFim': period[1].isoformat(),
    }
